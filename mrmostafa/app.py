from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timedelta, date
import os
from functools import wraps
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from config import config
import time

# Load environment variables
load_dotenv()

# Create Flask app
app = Flask(__name__)

# Load configuration based on environment
config_name = os.environ.get('FLASK_CONFIG', 'development')
app.config.from_object(config[config_name])

# Ensure instance directory exists for SQLite
if 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', ''):
    db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
    db_dir = os.path.dirname(db_path)
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)

# Initialize SQLAlchemy
db = SQLAlchemy(app)

# Association table for many-to-many relationship between students and groups
student_groups = db.Table('student_groups',
    db.Column('student_id', db.Integer, db.ForeignKey('student.id'), primary_key=True),
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True)
)

# Association table for many-to-many relationship between groups and subjects
group_subjects = db.Table('group_subjects',
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True),
    db.Column('subject_id', db.Integer, db.ForeignKey('subject.id'), primary_key=True)
)

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='instructor')  # 'admin' or 'instructor'
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'), nullable=True)  # Link to instructor if role is instructor
    is_hidden = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    last_activity = db.Column(db.DateTime, default=datetime.utcnow)
    is_online = db.Column(db.Boolean, default=False)
    
    # Permission fields for granular access control
    can_manage_payments = db.Column(db.Boolean, default=False)  # Add/edit/view payments
    can_take_attendance = db.Column(db.Boolean, default=False)  # Take attendance
    can_view_reports = db.Column(db.Boolean, default=False)     # View reports and analytics
    can_manage_students = db.Column(db.Boolean, default=False)  # Add/edit students
    can_manage_groups = db.Column(db.Boolean, default=False)    # Add/edit groups  
    can_manage_instructors = db.Column(db.Boolean, default=False) # Add/edit instructors
    can_manage_users = db.Column(db.Boolean, default=False)     # User management (admin only)
    can_manage_subjects = db.Column(db.Boolean, default=False)  # Manage subjects
    can_export_data = db.Column(db.Boolean, default=False)     # Export to Excel
    can_import_data = db.Column(db.Boolean, default=False)     # Import from Excel
    can_manage_expenses = db.Column(db.Boolean, default=False)  # Manage expenses
    can_manage_tasks = db.Column(db.Boolean, default=False)     # Manage tasks
    
    # Relationship
    linked_instructor = db.relationship('Instructor', backref='user_account', uselist=False)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def update_activity(self):
        """Update last activity and set user as online"""
        self.last_activity = datetime.utcnow()
        self.is_online = True
        db.session.commit()
    
    def is_active_now(self):
        """Check if user is active (last activity within 5 minutes)"""
        if not self.last_activity:
            return False
        return (datetime.utcnow() - self.last_activity).total_seconds() < 300  # 5 minutes
    
    def has_permission(self, permission):
        """Check if user has a specific permission"""
        if self.role == 'admin':
            return True  # Admin has all permissions
        return getattr(self, f'can_{permission}', False)
    
    def get_permissions_list(self):
        """Get list of all permissions this user has"""
        if self.role == 'admin':
            return ['manage_payments', 'take_attendance', 'view_reports', 'manage_students', 
                   'manage_groups', 'manage_instructors', 'manage_users', 'manage_subjects',
                   'export_data', 'import_data', 'manage_expenses', 'manage_tasks']
        
        permissions = []
        permission_fields = ['manage_payments', 'take_attendance', 'view_reports', 'manage_students',
                           'manage_groups', 'manage_instructors', 'manage_subjects', 'export_data', 
                           'import_data', 'manage_expenses', 'manage_tasks']
        
        for perm in permission_fields:
            if getattr(self, f'can_{perm}', False):
                permissions.append(perm)
        
        return permissions
    
    def set_role_permissions(self, role_type):
        """Set default permissions based on role type"""
        # Reset all permissions first
        self.can_manage_payments = False
        self.can_take_attendance = False
        self.can_view_reports = False
        self.can_manage_students = False
        self.can_manage_groups = False
        self.can_manage_instructors = False
        self.can_manage_users = False
        self.can_manage_subjects = False
        self.can_export_data = False
        self.can_import_data = False
        self.can_manage_expenses = False
        self.can_manage_tasks = False
        
        if role_type == 'financial_administrator':
            # مدير مالي - Financial Administrator - ONLY PAYMENTS
            self.can_manage_payments = True
            self.can_manage_expenses = True
        elif role_type == 'attendance_coordinator':
            # منسق الحضور - Attendance Coordinator - ONLY ATTENDANCE
            self.can_take_attendance = True
        elif role_type == 'student_affairs_manager':
            # مدير شؤون الطلاب - Student Affairs Manager - ONLY STUDENTS
            self.can_manage_students = True
        elif role_type == 'academic_coordinator':
            # منسق أكاديمي - Academic Coordinator - ONLY INSTRUCTORS, GROUPS, SUBJECTS
            self.can_manage_instructors = True
            self.can_manage_groups = True
            self.can_manage_subjects = True
            self.can_manage_tasks = True
        elif role_type == 'data_analyst':
            # محلل بيانات - Data Analyst - ONLY REPORTS
            self.can_view_reports = True
        elif role_type == 'senior_instructor':
            # مدرس أول - Senior Instructor - ATTENDANCE + GROUPS + TASKS
            self.can_take_attendance = True
            self.can_manage_groups = True
            self.can_manage_tasks = True
        elif role_type == 'assistant_instructor':
            # مدرس مساعد - Assistant Instructor - ONLY ATTENDANCE
            self.can_take_attendance = True
        elif role_type == 'data_entry_specialist':
            # أخصائي إدخال بيانات - Data Entry Specialist - ONLY STUDENTS (for data entry)
            self.can_manage_students = True
        elif role_type == 'admin':
            # Admin gets all permissions automatically via has_permission method
            self.role = 'admin'
            # Set all permissions to True for admin
            self.can_manage_payments = True
            self.can_take_attendance = True
            self.can_view_reports = True
            self.can_manage_students = True
            self.can_manage_groups = True
            self.can_manage_instructors = True
            self.can_manage_users = True
            self.can_manage_subjects = True
            self.can_export_data = True
            self.can_import_data = True
            self.can_manage_expenses = True
            self.can_manage_tasks = True
    
    @staticmethod
    def get_available_role_types():
        """Get list of available professional role types"""
        return {
            'financial_administrator': {
                'name': 'مدير مالي',
                'name_en': 'Financial Administrator',
                'description': 'المدفوعات والمصروفات فقط',
                'permissions': ['manage_payments', 'manage_expenses']
            },
            'attendance_coordinator': {
                'name': 'منسق الحضور', 
                'name_en': 'Attendance Coordinator',
                'description': 'تسجيل الحضور فقط',
                'permissions': ['take_attendance']
            },
            'student_affairs_manager': {
                'name': 'مدير شؤون الطلاب',
                'name_en': 'Student Affairs Manager', 
                'description': 'إدارة الطلاب ودرجاتهم فقط',
                'permissions': ['manage_students']
            },
            'academic_coordinator': {
                'name': 'منسق أكاديمي',
                'name_en': 'Academic Coordinator',
                'description': 'إدارة المدرسين والمجموعات والمواد والمهام',
                'permissions': ['manage_instructors', 'manage_groups', 'manage_subjects', 'manage_tasks']
            },
            'data_analyst': {
                'name': 'محلل بيانات',
                'name_en': 'Data Analyst',
                'description': 'عرض التقارير فقط',
                'permissions': ['view_reports']
            },
            'senior_instructor': {
                'name': 'مدرس أول',
                'name_en': 'Senior Instructor',
                'description': 'الحضور والمجموعات والمهام',
                'permissions': ['take_attendance', 'manage_groups', 'manage_tasks']
            },
            'assistant_instructor': {
                'name': 'مدرس مساعد',
                'name_en': 'Assistant Instructor', 
                'description': 'تسجيل الحضور فقط',
                'permissions': ['take_attendance']
            },
            'data_entry_specialist': {
                'name': 'أخصائي إدخال بيانات',
                'name_en': 'Data Entry Specialist',
                'description': 'إدخال بيانات الطلاب فقط',
                'permissions': ['manage_students']
            }
        }
    
    def get_role_info(self):
        """Get professional role information for this user"""
        role_types = self.get_available_role_types()
        permissions = self.get_permissions_list()
        
        # Find matching role type based on permissions
        for role_key, role_info in role_types.items():
            if set(role_info['permissions']) == set(permissions):
                return {
                    'key': role_key,
                    'name': role_info['name'],
                    'name_en': role_info['name_en'],
                    'description': role_info['description']
                }
        
        # If no exact match found, return custom role
        if self.role == 'admin':
            return {
                'key': 'admin',
                'name': 'مدير النظام',
                'name_en': 'System Administrator',
                'description': 'صلاحيات كاملة لإدارة النظام'
            }
        
        return {
            'key': 'custom',
            'name': 'صلاحيات مخصصة',
            'name_en': 'Custom Role',
            'description': f'صلاحيات مخصصة ({len(permissions)} صلاحية)'
        }

class Instructor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    specialization = db.Column(db.String(100))
    students = db.relationship('Student', backref='instructor_ref', lazy=True)
    groups = db.relationship('Group', backref='instructor_ref', lazy=True)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    age = db.Column(db.Integer)
    location = db.Column(db.String(50))  # Changed from level to location
    grade_level = db.Column(db.String(50))  # المرحلة الدراسية (رياض الأطفال، ابتدائي، إعدادي، ثانوي)
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'))
    # Removed group_id - now using many-to-many relationship
    total_paid = db.Column(db.Float, default=0.0)
    discount = db.Column(db.Float, default=0.0)  # Discount amount in currency
    # Removed course_price - now price is per group
    registration_date = db.Column(db.DateTime, nullable=False)
    
    # Achievement Points System
    total_achievement_points = db.Column(db.Float, default=0.0)  # إجمالي نقاط الإنجاز
    attendance_points = db.Column(db.Float, default=0.0)  # نقاط الحضور
    grade_points = db.Column(db.Float, default=0.0)  # نقاط الدرجات
    bonus_points = db.Column(db.Float, default=0.0)  # نقاط إضافية (للمكافآت)
    achievement_level = db.Column(db.String(20), default='مبتدئ')  # مستوى الإنجاز
    last_points_update = db.Column(db.DateTime, default=datetime.utcnow)  # آخر تحديث للنقاط
    
    # Many-to-many relationship with groups
    groups = db.relationship('Group', secondary=student_groups, backref=db.backref('students', lazy='dynamic'))
    
    @property
    def total_course_price(self):
        """Calculate total price of all groups the student is enrolled in"""
        return sum(group.price for group in self.groups)
    
    @property
    def total_course_price_after_discount(self):
        """Calculate total price after applying discount"""
        total_price = self.total_course_price
        discounted_price = total_price - self.discount
        return max(0, discounted_price)  # Ensure price doesn't go below 0

    @property
    def remaining_balance(self):
        """Calculate remaining balance for the student after discount"""
        balance = self.total_course_price_after_discount - self.total_paid
        return max(0, balance)  # Ensure we don't return negative balance as pending payment
    
    # Achievement Points Methods
    def get_achievement_rules(self):
        """Get achievement calculation rules based on grade level"""
        rules = {
            'رياض الأطفال': {
                'attendance_weight': 0.7,  # 70% weight for attendance
                'grade_weight': 0.3,       # 30% weight for grades
                'max_attendance_points': 50,
                'max_grade_points': 30,
                'levels': {
                    'مبتدئ': 0,
                    'متقدم': 30,
                    'متفوق': 60,
                    'نجم': 80
                }
            },
            'ابتدائي': {
                'attendance_weight': 0.6,  # 60% weight for attendance
                'grade_weight': 0.4,       # 40% weight for grades
                'max_attendance_points': 60,
                'max_grade_points': 40,
                'levels': {
                    'مبتدئ': 0,
                    'متقدم': 40,
                    'متفوق': 75,
                    'نجم': 100
                }
            },
            'إعدادي': {
                'attendance_weight': 0.5,  # 50% weight for attendance
                'grade_weight': 0.5,       # 50% weight for grades
                'max_attendance_points': 70,
                'max_grade_points': 70,
                'levels': {
                    'مبتدئ': 0,
                    'متقدم': 50,
                    'متفوق': 100,
                    'نجم': 140
                }
            },
            'ثانوي': {
                'attendance_weight': 0.4,  # 40% weight for attendance
                'grade_weight': 0.6,       # 60% weight for grades
                'max_attendance_points': 80,
                'max_grade_points': 120,
                'levels': {
                    'مبتدئ': 0,
                    'متقدم': 60,
                    'متفوق': 120,
                    'نجم': 180
                }
            }
        }
        return rules.get(self.grade_level, rules['ابتدائي'])  # Default to primary if grade level not found
    
    def calculate_attendance_points(self):
        """Calculate achievement points based on attendance"""
        from datetime import datetime, timedelta
        
        # Get last 30 days attendance records
        thirty_days_ago = datetime.now() - timedelta(days=30)
        attendance_records = Attendance.query.filter(
            Attendance.student_id == self.id,
            Attendance.date >= thirty_days_ago.date()
        ).all()
        
        if not attendance_records:
            return 0
        
        # Calculate attendance statistics
        total_sessions = len(attendance_records)
        present_count = len([a for a in attendance_records if a.status == 'حاضر'])
        late_count = len([a for a in attendance_records if a.status == 'متأخر'])
        absent_count = len([a for a in attendance_records if a.status == 'غائب'])
        
        # Calculate attendance score (present = 1.0, late = 0.5, absent = 0)
        attendance_score = (present_count * 1.0 + late_count * 0.5) / total_sessions
        
        # Get achievement rules for this grade level
        rules = self.get_achievement_rules()
        max_points = rules['max_attendance_points']
        
        # Calculate points based on attendance percentage
        if attendance_score >= 0.95:  # 95%+ attendance
            points = max_points
        elif attendance_score >= 0.90:  # 90-94% attendance
            points = max_points * 0.9
        elif attendance_score >= 0.80:  # 80-89% attendance
            points = max_points * 0.7
        elif attendance_score >= 0.70:  # 70-79% attendance
            points = max_points * 0.5
        elif attendance_score >= 0.60:  # 60-69% attendance
            points = max_points * 0.3
        else:  # Below 60% attendance
            points = max_points * 0.1
        
        return round(points, 2)
    
    def calculate_grade_points(self):
        """Calculate achievement points based on grades"""
        from datetime import datetime, timedelta
        
        # Get last 30 days grades
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_grades = Grade.query.filter(
            Grade.student_id == self.id,
            Grade.created_at >= thirty_days_ago
        ).all()
        
        if not recent_grades:
            return 0
        
        # Calculate average percentage from recent grades
        valid_grades = [g for g in recent_grades if g.percentage is not None]
        if not valid_grades:
            return 0
        
        avg_percentage = sum(g.percentage for g in valid_grades) / len(valid_grades)
        
        # Get achievement rules for this grade level
        rules = self.get_achievement_rules()
        max_points = rules['max_grade_points']
        
        # Calculate points based on grade percentage
        if avg_percentage >= 95:  # A+ (95%+)
            points = max_points
        elif avg_percentage >= 90:  # A (90-94%)
            points = max_points * 0.95
        elif avg_percentage >= 85:  # B+ (85-89%)
            points = max_points * 0.85
        elif avg_percentage >= 80:  # B (80-84%)
            points = max_points * 0.75
        elif avg_percentage >= 75:  # C+ (75-79%)
            points = max_points * 0.65
        elif avg_percentage >= 70:  # C (70-74%)
            points = max_points * 0.55
        elif avg_percentage >= 65:  # D+ (65-69%)
            points = max_points * 0.4
        elif avg_percentage >= 60:  # D (60-64%)
            points = max_points * 0.25
        else:  # F (Below 60%)
            points = max_points * 0.1
        
        return round(points, 2)
    
    def update_achievement_points(self):
        """Update all achievement points and determine level"""
        # Calculate new points
        new_attendance_points = self.calculate_attendance_points()
        new_grade_points = self.calculate_grade_points()
        
        # Update points
        self.attendance_points = new_attendance_points
        self.grade_points = new_grade_points
        self.total_achievement_points = self.attendance_points + self.grade_points + self.bonus_points
        
        # Determine achievement level
        rules = self.get_achievement_rules()
        levels = rules['levels']
        
        current_points = self.total_achievement_points
        if current_points >= levels['نجم']:
            self.achievement_level = 'نجم'
        elif current_points >= levels['متفوق']:
            self.achievement_level = 'متفوق'
        elif current_points >= levels['متقدم']:
            self.achievement_level = 'متقدم'
        else:
            self.achievement_level = 'مبتدئ'
        
        # Update timestamp
        self.last_points_update = datetime.utcnow()
        
        # Save to database
        db.session.add(self)
        db.session.commit()
        
        return {
            'total_points': self.total_achievement_points,
            'attendance_points': self.attendance_points,
            'grade_points': self.grade_points,
            'bonus_points': self.bonus_points,
            'level': self.achievement_level
        }
    
    def get_achievement_level_info(self):
        """Get information about current and next achievement level"""
        rules = self.get_achievement_rules()
        levels = rules['levels']
        current_points = self.total_achievement_points
        
        # Find current level info
        current_level = self.achievement_level
        current_level_points = levels[current_level]
        
        # Find next level
        level_order = ['مبتدئ', 'متقدم', 'متفوق', 'نجم']
        current_index = level_order.index(current_level)
        
        if current_index < len(level_order) - 1:
            next_level = level_order[current_index + 1]
            next_level_points = levels[next_level]
            points_needed = next_level_points - current_points
        else:
            next_level = None
            next_level_points = None
            points_needed = 0
        
        return {
            'current_level': current_level,
            'current_points': current_points,
            'next_level': next_level,
            'points_needed': max(0, points_needed) if points_needed else 0,
            'progress_percentage': min(100, (current_points / next_level_points * 100) if next_level_points else 100)
        }
    
    @property
    def achievement_badge_color(self):
        """Get badge color for achievement level"""
        colors = {
            'مبتدئ': '#6c757d',     # Gray
            'متقدم': '#17a2b8',     # Info blue
            'متفوق': '#28a745',     # Success green  
            'نجم': '#ffc107'        # Warning gold
        }
        return colors.get(self.achievement_level, '#6c757d')

class Group(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    # Removed level field - now using subjects relationship
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'))
    max_students = db.Column(db.Integer, default=15)
    price = db.Column(db.Float, default=0.0)  # Price for this group
    # Course completion fields
    status = db.Column(db.String(20), default='active')  # active, completed
    completion_date = db.Column(db.Date, nullable=True)
    completion_notes = db.Column(db.Text, nullable=True)
    # Monthly payment settings
    monthly_payment_enabled = db.Column(db.Boolean, default=True)  # Enable monthly payments
    monthly_price = db.Column(db.Float, default=0.0)  # Monthly price (if different from total price)
    payment_due_day = db.Column(db.Integer, default=1)  # Day of month when payment is due (1-31)
    # Students relationship is now defined in Student model with secondary table
    schedules = db.relationship('Schedule', backref='group_ref', lazy=True)
    # Many-to-many relationship with subjects
    subjects = db.relationship('Subject', secondary=group_subjects, backref=db.backref('groups', lazy='dynamic'))
    
    @property
    def is_completed(self):
        """Check if the group/course is completed"""
        return self.status == 'completed'
    
    @property
    def active_students_count(self):
        """Count of students currently enrolled in this group"""
        return self.students.count()
    
    @property
    def effective_monthly_price(self):
        """Get the effective monthly price (monthly_price if set, otherwise total price)"""
        return self.monthly_price if self.monthly_price > 0 else self.price
    
    def get_monthly_payment(self, year, month):
        """Get or create monthly payment record for this group"""
        monthly_payment = MonthlyPayment.query.filter_by(
            group_id=self.id, year=year, month=month
        ).first()
        
        if not monthly_payment:
            # Create new monthly payment record
            arabic_months = {
                1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
                5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
                9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
            }
            
            # Calculate due date
            try:
                due_date = date(year, month, min(self.payment_due_day, 28))  # Use 28 to avoid month overflow
            except:
                due_date = date(year, month, 1)
            
            monthly_payment = MonthlyPayment(
                group_id=self.id,
                year=year,
                month=month,
                month_name=arabic_months.get(month, f'شهر {month}'),
                monthly_price=self.effective_monthly_price,
                due_date=due_date
            )
            db.session.add(monthly_payment)
            db.session.commit()
        
        return monthly_payment
    
    def get_current_month_payment_status(self):
        """Get payment status for current month"""
        now = datetime.now()
        monthly_payment = self.get_monthly_payment(now.year, now.month)
        return {
            'status': monthly_payment.payment_status,
            'paid': monthly_payment.total_paid,
            'remaining': monthly_payment.remaining_amount,
            'is_overdue': monthly_payment.is_overdue
        }

class Schedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'))
    day_of_week = db.Column(db.String(20))  # السبت، الأحد، الاثنين، etc.
    start_time = db.Column(db.String(10))
    end_time = db.Column(db.String(10))

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'))
    date = db.Column(db.Date)
    status = db.Column(db.String(20))  # حاضر، غائب، متأخر
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'))

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'))
    amount = db.Column(db.Float)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    month = db.Column(db.String(20))
    notes = db.Column(db.Text)

class MonthlyPayment(db.Model):
    """Track monthly payments for groups"""
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)  # 1-12
    month_name = db.Column(db.String(20), nullable=False)  # Arabic month name
    monthly_price = db.Column(db.Float, nullable=False)  # Price for this month
    total_paid = db.Column(db.Float, default=0.0)  # Total paid for this month
    payment_status = db.Column(db.String(20), default='pending')  # pending, partial, complete
    due_date = db.Column(db.Date, nullable=True)  # When this month's payment is due
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)
    
    # Relationship
    group = db.relationship('Group', backref='monthly_payments')
    
    @property
    def remaining_amount(self):
        """Calculate remaining amount for this month"""
        return max(0, self.monthly_price - self.total_paid)
    
    @property
    def is_overdue(self):
        """Check if payment is overdue"""
        if not self.due_date:
            return False
        return self.due_date < date.today() and self.payment_status != 'complete'
    
    def update_payment_status(self):
        """Update payment status based on total paid"""
        if self.total_paid == 0:
            self.payment_status = 'pending'
        elif self.total_paid >= self.monthly_price:
            self.payment_status = 'complete'
        else:
            self.payment_status = 'partial'
        self.updated_at = datetime.utcnow()

class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(100))  # رواتب، إيجار، مرافق، مستلزمات، أخرى
    date = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    priority = db.Column(db.String(20), default='متوسط')  # عالي، متوسط، منخفض
    status = db.Column(db.String(20), default='قيد التنفيذ')  # قيد التنفيذ، مكتمل، ملغي
    due_date = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    # Relationships
    creator = db.relationship('User', foreign_keys=[created_by], backref='created_tasks')
    assignee = db.relationship('User', foreign_keys=[assigned_to], backref='assigned_tasks')

    @property
    def is_overdue(self):
        """Check if task is overdue"""
        if self.due_date and self.status != 'مكتمل':
            return datetime.now().date() > self.due_date
        return False

    @property
    def days_remaining(self):
        """Calculate days remaining until due date"""
        if self.due_date and self.status != 'مكتمل':
            delta = self.due_date - datetime.now().date()
            return delta.days
        return None

class Subject(db.Model):
    """Model for subjects/courses/exams that grades are recorded for"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # اسم المادة/الاختبار
    code = db.Column(db.String(20))  # كود المادة (اختياري)
    description = db.Column(db.Text)  # وصف المادة
    max_grade = db.Column(db.Float, default=100.0)  # الدرجة القصوى
    min_grade = db.Column(db.Float, default=0.0)  # الدرجة الدنيا
    subject_type = db.Column(db.String(50), default='مادة')  # نوع (مادة، اختبار، واجب، مشروع)
    # Removed group_id - now using many-to-many relationship with groups
    instructor_id = db.Column(db.Integer, db.ForeignKey('instructor.id'), nullable=True)  # ربط بمدرس
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    instructor = db.relationship('Instructor', backref='subjects')
    grades = db.relationship('Grade', backref='subject', cascade='all, delete-orphan')
    
    @property
    def average_grade(self):
        """Calculate average grade for this subject"""
        if not self.grades:
            return 0
        return sum(grade.score for grade in self.grades if grade.score is not None) / len([g for g in self.grades if g.score is not None])
    
    @property
    def students_count(self):
        """Count of students who have grades for this subject"""
        return len(set(grade.student_id for grade in self.grades))

class Grade(db.Model):
    """Model for storing student grades"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    score = db.Column(db.Float)  # الدرجة المحصلة
    max_score = db.Column(db.Float)  # الدرجة القصوى للاختبار (قد تختلف عن المادة)
    percentage = db.Column(db.Float)  # النسبة المئوية
    grade_letter = db.Column(db.String(5))  # الدرجة بالحروف (A, B, C, D, F)
    exam_date = db.Column(db.Date)  # تاريخ الاختبار
    notes = db.Column(db.Text)  # ملاحظات
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    student = db.relationship('Student', backref='grades')
    
    def calculate_percentage(self):
        """Calculate percentage based on score and max_score"""
        if self.score is not None and self.max_score and self.max_score > 0:
            self.percentage = (self.score / self.max_score) * 100
        elif self.score is not None and self.subject and self.subject.max_grade > 0:
            self.percentage = (self.score / self.subject.max_grade) * 100
        else:
            self.percentage = 0
    
    def calculate_letter_grade(self):
        """Calculate letter grade based on percentage"""
        if self.percentage is None:
            self.calculate_percentage()
        
        if self.percentage >= 90:
            self.grade_letter = 'A'
        elif self.percentage >= 80:
            self.grade_letter = 'B'
        elif self.percentage >= 70:
            self.grade_letter = 'C'
        elif self.percentage >= 60:
            self.grade_letter = 'D'
        else:
            self.grade_letter = 'F'
    
    def save_with_calculations(self):
        """Save grade with automatic calculations"""
        self.calculate_percentage()
        self.calculate_letter_grade()
        self.updated_at = datetime.utcnow()
        db.session.add(self)
        db.session.commit()
        
        # Auto-update achievement points for the student
        try:
            if self.student:
                self.student.update_achievement_points()
        except Exception as e:
            # Log error but don't fail the grade save
            print(f"Error updating achievement points for student {self.student_id}: {e}")

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    category = db.Column(db.String(50), default='عام')  # عام، شخصي، عمل، مهم
    color = db.Column(db.String(20), default='yellow')  # yellow, blue, green, red, purple
    is_pinned = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    # Relationship
    creator = db.relationship('User', backref='notes')

    @property
    def created_ago(self):
        """Get how long ago the note was created"""
        delta = datetime.utcnow() - self.created_at
        if delta.days > 0:
            return f'منذ {delta.days} يوم'
        elif delta.seconds > 3600:
            return f'منذ {delta.seconds // 3600} ساعة'
        elif delta.seconds > 60:
            return f'منذ {delta.seconds // 60} دقيقة'
        else:
            return 'منذ لحظات'

class InstructorNote(db.Model):
    """Notes created by instructors - automatically sent to admins"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=True)  # Optional: specific student
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=True)  # Optional: specific group
    priority = db.Column(db.String(20), default='متوسط')  # عالي، متوسط، منخفض
    status = db.Column(db.String(20), default='جديد')  # جديد، قيد المراجعة، مكتمل
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # Instructor user
    reviewed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Admin who reviewed
    reviewed_at = db.Column(db.DateTime, nullable=True)
    admin_response = db.Column(db.Text, nullable=True)
    
    # Relationships
    creator = db.relationship('User', foreign_keys=[created_by], backref='instructor_notes')
    reviewer = db.relationship('User', foreign_keys=[reviewed_by])
    student = db.relationship('Student', backref='instructor_notes')
    group = db.relationship('Group', backref='instructor_notes')

    @property
    def created_ago(self):
        """Get how long ago the note was created"""
        delta = datetime.utcnow() - self.created_at
        if delta.days > 0:
            return f'منذ {delta.days} يوم'
        elif delta.seconds > 3600:
            return f'منذ {delta.seconds // 3600} ساعة'
        elif delta.seconds > 60:
            return f'منذ {delta.seconds // 60} دقيقة'
        else:
            return 'منذ لحظات'

class InstructorTodo(db.Model):
    """Todo list for instructors - personal task management"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default='مفتوح')  # مفتوح، مكتمل، ملغي
    priority = db.Column(db.String(20), default='متوسط')  # عالي، متوسط، منخفض
    category = db.Column(db.String(50), default='عام')  # عام، تحضير، حضور، متابعة، إداري
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=True)  # Optional: specific group
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=True)  # Optional: specific student
    due_date = db.Column(db.Date, nullable=True)  # Optional: due date
    reminder_date = db.Column(db.DateTime, nullable=True)  # Optional: reminder date
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    completed_at = db.Column(db.DateTime, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # Instructor user
    
    # Relationships
    creator = db.relationship('User', backref='instructor_todos')
    group = db.relationship('Group', backref='instructor_todos')
    student = db.relationship('Student', backref='instructor_todos')

    @property
    def is_overdue(self):
        """Check if todo is overdue"""
        if self.due_date and self.status == 'مفتوح':
            return datetime.now().date() > self.due_date
        return False

    @property
    def days_remaining(self):
        """Calculate days remaining until due date"""
        if self.due_date and self.status == 'مفتوح':
            delta = self.due_date - datetime.now().date()
            return delta.days
        return None

    @property
    def created_ago(self):
        """Get how long ago the todo was created"""
        delta = datetime.utcnow() - self.created_at
        if delta.days > 0:
            return f'منذ {delta.days} يوم'
        elif delta.seconds > 3600:
            return f'منذ {delta.seconds // 3600} ساعة'
        elif delta.seconds > 60:
            return f'منذ {delta.seconds // 60} دقيقة'
        else:
            return 'منذ لحظات'

# Update user activity before each request
@app.before_request
def update_user_activity():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user:
            user.update_activity()

# Authentication functions
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يرجى تسجيل الدخول للوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'error')
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

def instructor_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'error')
            return redirect(url_for('login'))
        
        user = User.query.get(session['user_id'])
        if not user or user.role not in ['admin', 'instructor']:
            flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

# Permission-based decorators
def permission_required(permission):
    """Decorator factory for permission-based access control"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('يجب تسجيل الدخول أولاً', 'error')
                return redirect(url_for('login'))
            
            user = User.query.get(session['user_id'])
            if not user or not user.has_permission(permission):
                flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Specific permission decorators
def payments_required(f):
    """Require permission to manage payments"""
    return permission_required('manage_payments')(f)

def attendance_required(f):
    """Require permission to take attendance"""
    return permission_required('take_attendance')(f)

def reports_required(f):
    """Require permission to view reports"""
    return permission_required('view_reports')(f)

def students_required(f):
    """Require permission to manage students"""
    return permission_required('manage_students')(f)

def groups_required(f):
    """Require permission to manage groups"""
    return permission_required('manage_groups')(f)

def instructors_required(f):
    """Require permission to manage instructors"""
    return permission_required('manage_instructors')(f)

def users_required(f):
    """Require permission to manage users"""
    return permission_required('manage_users')(f)

def subjects_required(f):
    """Require permission to manage subjects"""
    return permission_required('manage_subjects')(f)

def export_required(f):
    """Require permission to export data"""
    return permission_required('export_data')(f)

def import_required(f):
    """Require permission to import data"""
    return permission_required('import_data')(f)

def expenses_required(f):
    """Require permission to manage expenses"""
    return permission_required('manage_expenses')(f)

def tasks_required(f):
    """Require permission to manage tasks"""
    return permission_required('manage_tasks')(f)

def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None

# Make current user available in all templates
@app.context_processor
def inject_current_user():
    return dict(current_user=get_current_user())

def create_default_admin():
    """Create default hidden admin user if it doesn't exist"""
    admin_user = User.query.filter_by(username='araby').first()
    if not admin_user:
        admin_user = User(
            username='araby',
            full_name='System Administrator',
            role='admin',
            is_hidden=True
        )
        admin_user.set_password('92321066')
        db.session.add(admin_user)
        db.session.commit()
        print("Default admin user 'araby' created successfully!")

# Helper function to get Arabic day name
def get_arabic_day_name(date_obj):
    arabic_days = {
        'Monday': 'الاثنين',
        'Tuesday': 'الثلاثاء', 
        'Wednesday': 'الأربعاء',
        'Thursday': 'الخميس',
        'Friday': 'الجمعة',
        'Saturday': 'السبت',
        'Sunday': 'الأحد'
    }
    english_day = date_obj.strftime('%A')
    return arabic_days.get(english_day, english_day)

# Helper function to format date in Arabic
def format_arabic_date(date_obj):
    """Format date as: 4 يونيو 2025 (Day Month Year)"""
    if not date_obj:
        return ""
    
    arabic_months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    
    day = date_obj.day
    month = arabic_months[date_obj.month]
    year = date_obj.year
    
    # Format: Day Month Year (e.g., 4 يونيو 2025)
    return f"{day} {month} {year}"

# Helper function to get Arabic month name
def get_arabic_month_name(month_num):
    """Convert month number to Arabic month name"""
    arabic_months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    
    return arabic_months.get(month_num, "")

def format_time_12hour(datetime_obj):
    """Format time in 12-hour format with Arabic AM/PM"""
    if not datetime_obj:
        return ""
    
    time_str = datetime_obj.strftime('%I:%M')
    am_pm = datetime_obj.strftime('%p')
    
    # Convert AM/PM to Arabic
    if am_pm == 'AM':
        am_pm_arabic = 'ص'
    else:
        am_pm_arabic = 'م'
    
    return f"{time_str} {am_pm_arabic}"

def format_date_for_input(date_obj):
    """Format date for HTML input fields as DD-MM-YYYY"""
    if not date_obj:
        return ""
    
    return date_obj.strftime('%d-%m-%Y')

def parse_date_from_input(date_string):
    """Parse date from HTML5 date input (YYYY-MM-DD) or DD-MM-YYYY format to datetime object"""
    if not date_string:
        return None
    
    try:
        # Try YYYY-MM-DD format first (HTML5 date input format)
        return datetime.strptime(date_string, '%Y-%m-%d')
    except ValueError:
        try:
            # Fallback to DD-MM-YYYY format (manual input)
            return datetime.strptime(date_string, '%d-%m-%Y')
        except ValueError:
            # If both fail, return None
            return None

# Function to get today's schedule
def get_today_schedule():
    today = datetime.now()
    today_arabic = get_arabic_day_name(today)
    
    # Get all schedules for today
    today_schedules = Schedule.query.filter_by(day_of_week=today_arabic).all()
    
    schedule_data = []
    for schedule in today_schedules:
        group = Group.query.get(schedule.group_id)
        if group and group.instructor_ref:
            schedule_data.append({
                'group_name': group.name,
                'instructor_name': group.instructor_ref.name,
                'start_time': schedule.start_time,
                'end_time': schedule.end_time,
                'subjects': [s.name for s in group.subjects] if group.subjects else ['عام'],
                'student_count': group.students.count(),  # Use count() for dynamic relationship
                'max_students': group.max_students  # Add max_students field
            })
    
    # Sort by start time
    schedule_data.sort(key=lambda x: x['start_time'])
    return schedule_data

# Function to get weekly schedule
def get_weekly_schedule():
    """Get schedule for all days of the week"""
    arabic_days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    weekly_schedule = {}
    
    for day in arabic_days:
        # Get schedules for this day
        day_schedules = Schedule.query.filter_by(day_of_week=day).all()
        schedule_data = []
        
        for schedule in day_schedules:
            group = Group.query.get(schedule.group_id)
            # Include schedules even if group doesn't have instructor (with default values)
            if group:
                instructor_name = group.instructor_ref.name if group.instructor_ref else 'غير محدد'
                schedule_data.append({
                    'group_name': group.name,
                    'instructor_name': instructor_name,
                    'start_time': schedule.start_time,
                    'end_time': schedule.end_time,
                    'subjects': [s.name for s in group.subjects] if group.subjects else ['عام'],
                    'student_count': group.students.count(),  # Use count() for dynamic relationship
                    'max_students': group.max_students or 15,  # Default to 15 if not set
                    'group_id': group.id
                })
            else:
                # Handle orphaned schedules (group was deleted but schedule remains)
                schedule_data.append({
                    'group_name': 'مجموعة محذوفة',
                    'instructor_name': 'غير محدد',
                    'start_time': schedule.start_time,
                    'end_time': schedule.end_time,
                    'subjects': ['غير محدد'],
                    'student_count': 0,
                    'max_students': 15,
                    'group_id': 0
                })
        
        # Sort by start time
        try:
            schedule_data.sort(key=lambda x: datetime.strptime(x['start_time'], '%H:%M').time())
        except:
            # Fallback to string sorting if time parsing fails
            schedule_data.sort(key=lambda x: x['start_time'])
        
        weekly_schedule[day] = schedule_data
    
    return weekly_schedule

# Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        remember_me = 'remember_me' in request.form
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['user_role'] = user.role
            session['user_name'] = user.full_name
            
            # Set session as permanent if remember me is checked
            if remember_me:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=30)  # Remember for 30 days
            
            # Update last login and activity
            user.last_login = datetime.utcnow()
            user.update_activity()
            
            flash(f'مرحباً {user.full_name}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('تم تسجيل الخروج بنجاح! نراك قريباً 👋', 'success')
    return redirect(url_for('login'))

@app.route('/users')
@users_required
def users():
    users = User.query.filter_by(is_hidden=False).all()
    instructors = Instructor.query.all()
    current_user = get_current_user()
    return render_template('users.html', users=users, instructors=instructors, current_user=current_user)

@app.route('/add_user', methods=['POST'])
@users_required
def add_user():
    username = request.form['username'].lower()  # Ensure lowercase
    password = request.form['password']
    full_name = request.form['full_name']
    role_type = request.form.get('role_type')  # Get role type from form
    
    # Determine role based on role_type
    if role_type == 'admin':
        role = 'admin'
    else:
        role = 'user'  # All other roles are 'user' with specific permissions
    
    # Check if username already exists
    if User.query.filter_by(username=username).first():
        flash('اسم المستخدم موجود بالفعل', 'error')
        return redirect(url_for('users'))
    
    new_user = User(
        username=username,
        full_name=full_name,
        role=role,
        instructor_id=None
    )
    new_user.set_password(password)
    
    # Set permissions based on role type
    new_user.set_role_permissions(role_type)
    
    try:
        db.session.add(new_user)
        db.session.commit()
        flash('تم إضافة المستخدم بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء إضافة المستخدم', 'error')
    
    return redirect(url_for('users'))

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@users_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    current_user = get_current_user()
    
    # Prevent editing hidden admin unless you are the hidden admin
    if user.is_hidden and current_user.username != 'araby':
        flash('لا يمكن تعديل هذا المستخدم', 'error')
        return redirect(url_for('users'))
    
    user.username = request.form['username'].lower()  # Ensure lowercase
    user.full_name = request.form['full_name']
    
    # Handle role type and set permissions
    role_type = request.form.get('role_type')
    if role_type:
        # Determine role based on role_type
        if role_type == 'admin':
            user.role = 'admin'
        else:
            user.role = 'user'  # All other roles are 'user' with specific permissions
        
        user.set_role_permissions(role_type)
    
    if request.form['password']:
        user.set_password(request.form['password'])
    
    try:
        db.session.commit()
        flash('تم تحديث بيانات المستخدم بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء تحديث المستخدم', 'error')
    
    return redirect(url_for('users'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    current_user = get_current_user()
    
    # Prevent deleting hidden admin unless you are the hidden admin
    if user.is_hidden and current_user.username != 'araby':
        flash('لا يمكن حذف هذا المستخدم', 'error')
        return redirect(url_for('users'))
    
    # Prevent users from deleting themselves
    if user.id == current_user.id:
        flash('لا يمكن حذف حسابك الشخصي', 'error')
        return redirect(url_for('users'))
    
    db.session.delete(user)
    db.session.commit()
    flash('تم حذف المستخدم بنجاح', 'success')
    return redirect(url_for('users'))

@app.route('/')
@login_required
def index():
    current_user = get_current_user()
    
    # Redirect to role-specific dashboard
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif current_user.role == 'instructor':
        return redirect(url_for('instructor_dashboard'))
    else:
        # For users with specific role types, redirect to their specialized dashboard
        role_info = current_user.get_role_info()
        role_key = role_info['key']
        
        if role_key == 'financial_administrator':
            # Direct to payments page - their only function
            return redirect(url_for('payments'))
        elif role_key == 'attendance_coordinator':
            # Direct to attendance page - their only function
            return redirect(url_for('attendance'))
        elif role_key == 'student_affairs_manager':
            # Direct to students page - their only function
            return redirect(url_for('students'))
        elif role_key == 'academic_coordinator':
            # Direct to instructors page - their main function
            return redirect(url_for('instructors'))
        elif role_key == 'data_analyst':
            # Direct to reports page - their only function
            return redirect(url_for('reports'))
        elif role_key in ['senior_instructor', 'assistant_instructor']:
            # Direct to attendance page - their main function
            return redirect(url_for('attendance'))
        elif role_key == 'data_entry_specialist':
            # Direct to students page - their only function
            return redirect(url_for('students'))
        else:
            # Default dashboard for custom roles
            return redirect(url_for('default_dashboard'))

@app.route('/admin_dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard with full system overview"""
    students = Student.query.all()
    instructors = Instructor.query.all() 
    groups = Group.query.all()
    
    total_students = len(students)
    total_groups = len(groups)
    total_instructors = len(instructors)
    
    # Get today's schedule
    today_schedule = get_today_schedule()
    
    # Get weekly schedule  
    weekly_schedule = get_weekly_schedule()
    
    # Get today's Arabic day name
    today_arabic = get_arabic_day_name(datetime.now())
    
    return render_template('index.html', 
                         total_students=total_students,
                         total_groups=total_groups, 
                         total_instructors=total_instructors,
                         today_schedule=today_schedule,
                         weekly_schedule=weekly_schedule,
                         today_date=datetime.now(),
                         today_arabic=today_arabic)

@app.route('/financial_dashboard')
@payments_required
def financial_dashboard():
    """Financial Administrator dashboard - focused on payments and expenses"""
    current_user = get_current_user()
    
    # Get financial data
    total_payments = Payment.query.count()
    total_revenue = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
    total_expenses = Expense.query.count()
    total_expense_amount = db.session.query(db.func.sum(Expense.amount)).scalar() or 0
    
    # Recent payments
    recent_payments = Payment.query.order_by(Payment.date.desc()).limit(10).all()
    
    # Recent expenses  
    recent_expenses = Expense.query.order_by(Expense.date.desc()).limit(10).all()
    
    return render_template('financial_dashboard.html',
                         current_user=current_user,
                         total_payments=total_payments,
                         total_revenue=total_revenue,
                         total_expenses=total_expenses, 
                         total_expense_amount=total_expense_amount,
                         recent_payments=recent_payments,
                         recent_expenses=recent_expenses,
                         today_date=datetime.now())

@app.route('/attendance_dashboard')
@attendance_required  
def attendance_dashboard():
    """Attendance Coordinator dashboard - focused on attendance tracking"""
    current_user = get_current_user()
    
    # Get attendance data
    today = datetime.now().date()
    total_students = Student.query.count()
    today_attendance = Attendance.query.filter_by(date=today).count()
    
    # Get groups for attendance
    groups = Group.query.all()
    
    # Recent attendance records
    recent_attendance = Attendance.query.order_by(Attendance.date.desc()).limit(20).all()
    
    return render_template('attendance_dashboard.html',
                         current_user=current_user,
                         total_students=total_students,
                         today_attendance=today_attendance,
                         groups=groups,
                         recent_attendance=recent_attendance,
                         today_date=today,
                         today_schedule=get_today_schedule())

@app.route('/student_affairs_dashboard')
@students_required
def student_affairs_dashboard():
    """Student Affairs Manager dashboard - focused on student management"""
    current_user = get_current_user()
    
    # Get student data
    total_students = Student.query.count()
    new_students_this_month = Student.query.filter(
        Student.registration_date >= datetime.now().replace(day=1)
    ).count()
    
    # Students by grade level
    grade_levels = db.session.query(
        Student.grade_level, 
        db.func.count(Student.id)
    ).group_by(Student.grade_level).all()
    
    # Recent registrations
    recent_students = Student.query.order_by(Student.registration_date.desc()).limit(10).all()
    
    return render_template('student_affairs_dashboard.html',
                         current_user=current_user,
                         total_students=total_students,
                         new_students_this_month=new_students_this_month,
                         grade_levels=grade_levels,
                         recent_students=recent_students,
                         today_date=datetime.now())

@app.route('/academic_dashboard')
@groups_required
def academic_dashboard():
    """Academic Coordinator dashboard - focused on academic management"""
    current_user = get_current_user()
    
    # Get academic data
    total_instructors = Instructor.query.count()
    total_groups = Group.query.count() 
    total_subjects = Subject.query.count()
    
    # Get schedule data
    today_schedule = get_today_schedule()
    weekly_schedule = get_weekly_schedule()
    
    # Recent tasks
    recent_tasks = Task.query.order_by(Task.created_at.desc()).limit(10).all()
    
    return render_template('academic_dashboard.html',
                         current_user=current_user,
                         total_instructors=total_instructors,
                         total_groups=total_groups,
                         total_subjects=total_subjects,
                         today_schedule=today_schedule,
                         weekly_schedule=weekly_schedule,
                         recent_tasks=recent_tasks,
                         today_date=datetime.now(),
                         today_arabic=get_arabic_day_name(datetime.now()))

@app.route('/reports_dashboard')
@reports_required
def reports_dashboard():
    """Data Analyst dashboard - focused on reports and analytics"""
    current_user = get_current_user()
    
    # Redirect to existing reports page which has all the analytics
    return redirect(url_for('reports'))

@app.route('/data_entry_dashboard')
@students_required
def data_entry_dashboard():
    """Data Entry Specialist dashboard - focused on data management"""
    current_user = get_current_user()
    
    # Similar to student affairs but more focused on data entry tasks
    return redirect(url_for('students'))

@app.route('/default_dashboard')
@login_required
def default_dashboard():
    """Default dashboard for users with custom roles"""
    current_user = get_current_user()
    role_info = current_user.get_role_info()
    
    return render_template('default_dashboard.html',
                         current_user=current_user,
                         role_info=role_info,
                         permissions=current_user.get_permissions_list(),
                         today_date=datetime.now())

@app.route('/instructor_dashboard')
@instructor_required
def instructor_dashboard():
    current_user = get_current_user()
    
    if not current_user.linked_instructor:
        flash('حساب المدرس غير مرتبط بملف مدرس', 'error')
        return redirect(url_for('logout'))
    
    instructor = current_user.linked_instructor
    
    # Get instructor's groups and students
    instructor_groups = instructor.groups
    instructor_students = get_instructor_students(current_user)
    
    # Get today's schedule for this instructor
    today_schedule = []
    for group in instructor_groups:
        for schedule in group.schedules:
            if schedule.day_of_week == get_arabic_day_name(datetime.now()):
                today_schedule.append({
                    'group': group,
                    'schedule': schedule
                })
    
    # Get recent instructor notes
    recent_notes = InstructorNote.query.filter_by(created_by=current_user.id)\
                                      .order_by(InstructorNote.created_at.desc())\
                                      .limit(5).all()
    
    # Statistics
    total_students = len(instructor_students)
    total_groups = len(instructor_groups)
    
    # Get attendance statistics for instructor's groups
    total_classes_today = len(today_schedule)
    
    # Get unique ages for instructor's students
    instructor_ages = []
    for student in instructor_students:
        if student.age and student.age not in instructor_ages:
            instructor_ages.append(student.age)
    instructor_ages.sort()
    
    return render_template('instructor_dashboard.html',
                         instructor=instructor,
                         total_students=total_students,
                         total_groups=total_groups,
                         total_classes_today=total_classes_today,
                         today_schedule=today_schedule,
                         recent_notes=recent_notes,
                         instructor_groups=instructor_groups,
                         instructor_students=instructor_students,
                         instructor_ages=instructor_ages)

@app.route('/students')
@students_required
def students():
    # Get filter parameters
    group_filter = request.args.get('group_id', '')
    age_filter = request.args.get('age_range', '')
    location_filter = request.args.get('location', '')
    grade_level_filter = request.args.get('grade_level', '')
    
    # Start with base query
    query = Student.query
    
    # Apply group filter - since students can have multiple groups, we need to join
    if group_filter:
        query = query.join(Student.groups).filter(Group.id == int(group_filter))
    
    # Apply age filter
    if age_filter:
        try:
            age_value = int(age_filter)
            query = query.filter(Student.age == age_value)
        except ValueError:
            pass  # Ignore invalid age values
    
    # Apply location filter
    if location_filter:
        query = query.filter(Student.location.ilike(f'%{location_filter}%'))
    
    # Apply grade level filter
    if grade_level_filter:
        query = query.filter(Student.grade_level == grade_level_filter)
    
    students = query.all()
    instructors = Instructor.query.all()
    groups = Group.query.all()
    
    # Get all unique locations for the filter dropdown
    locations = db.session.query(Student.location).filter(Student.location.isnot(None)).distinct().all()
    locations = [loc[0] for loc in locations if loc[0] and loc[0].strip()]
    locations.sort()
    
    # Get all unique ages for the filter dropdown
    ages = db.session.query(Student.age).filter(Student.age.isnot(None)).distinct().all()
    ages = [age[0] for age in ages if age[0] is not None]
    ages.sort()
    
    # Get all unique grade levels for the filter dropdown
    grade_levels = db.session.query(Student.grade_level).filter(Student.grade_level.isnot(None)).distinct().all()
    grade_levels = [level[0] for level in grade_levels if level[0] and level[0].strip()]
    grade_levels.sort()
    
    return render_template('students.html', 
                         students=students, 
                         instructors=instructors, 
                         groups=groups,
                         locations=locations,
                         ages=ages,
                         grade_levels=grade_levels,
                         selected_group=group_filter,
                         selected_age=age_filter,
                         selected_location=location_filter,
                         selected_grade_level=grade_level_filter)

@app.route('/add_student', methods=['POST'])
def add_student():
    try:
        name = request.form['name']
        phone = request.form.get('phone', '')  # Use get() for optional fields
        age = int(request.form['age'])
        location = request.form.get('location', '')  # Use get() for optional fields
        grade_level = request.form.get('grade_level', '')  # المرحلة الدراسية
        instructor_id = None  # Make instructor optional - can be set later if needed
        registration_date = parse_date_from_input(request.form['registration_date'])
        discount = float(request.form.get('discount', 0))  # Get discount amount, default to 0
        
        if not registration_date:
            flash('يرجى اختيار تاريخ التسجيل من منتقي التاريخ', 'error')
            return redirect(url_for('students'))
        
        student = Student(
            name=name,
            phone=phone,
            age=age,
            location=location,
            grade_level=grade_level,
            instructor_id=instructor_id,
            registration_date=registration_date,
            discount=discount
        )
        
        db.session.add(student)
        db.session.commit()
        
        # Handle group selections (multiple groups allowed)
        group_ids = request.form.getlist('group_ids')  # Get list of selected group IDs
        if group_ids:
            for group_id in group_ids:
                if group_id:  # Make sure it's not empty
                    group = Group.query.get(int(group_id))
                    if group:
                        student.groups.append(group)
        
        db.session.commit()
        flash('تم إضافة الطالب بنجاح!', 'success')
        return redirect(url_for('students'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة الطالب: {str(e)}', 'error')
        return redirect(url_for('students'))

@app.route('/instructors')
@instructors_required
def instructors():
    instructors = Instructor.query.all()
    return render_template('instructors.html', instructors=instructors)

@app.route('/add_instructor', methods=['POST'])
def add_instructor():
    name = request.form['name']
    phone = request.form['phone']
    specialization = request.form['specialization']
    
    instructor = Instructor(
        name=name,
        phone=phone,
        specialization=specialization
    )
    
    db.session.add(instructor)
    db.session.commit()
    flash('تم إضافة المدرس بنجاح', 'success')
    return redirect(url_for('instructors'))

@app.route('/groups')
@groups_required
def groups():
    # Get filter parameters
    instructor_filter = request.args.get('instructor_id', type=int)
    
    # Start with base query
    groups_query = Group.query
    
    # Apply instructor filter if specified
    if instructor_filter:
        groups_query = groups_query.filter(Group.instructor_id == instructor_filter)
    
    groups = groups_query.all()
    instructors = Instructor.query.all()
    
    # Calculate total students across filtered groups
    total_students = 0
    for group in groups:
        total_students += group.students.count()
    
    return render_template('groups.html', 
                         groups=groups, 
                         instructors=instructors,
                         total_students=total_students,
                         selected_instructor=instructor_filter)

def check_instructor_schedule_conflicts(day, start_time, end_time, instructor_id, exclude_group_id=None):
    """Check for schedule conflicts for the same instructor only"""
    conflicts = []
    
    # Get all schedules for the same day and instructor
    existing_schedules = db.session.query(Schedule).join(Group).filter(
        Schedule.day_of_week == day,
        Group.instructor_id == instructor_id
    )
    
    # Exclude current group if editing
    if exclude_group_id:
        existing_schedules = existing_schedules.filter(Group.id != exclude_group_id)
    
    existing_schedules = existing_schedules.all()
    
    # Convert times to minutes for easier comparison
    def time_to_minutes(time_str):
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    
    new_start_min = time_to_minutes(start_time)
    new_end_min = time_to_minutes(end_time)
    
    for schedule in existing_schedules:
        existing_start_min = time_to_minutes(schedule.start_time)
        existing_end_min = time_to_minutes(schedule.end_time)
        
        # Check for overlap
        if (new_start_min < existing_end_min and new_end_min > existing_start_min):
            conflicts.append({
                'group_name': schedule.group_ref.name,
                'start_time': schedule.start_time,
                'end_time': schedule.end_time,
                'day': day
            })
    
    return conflicts



@app.route('/add_group', methods=['POST'])
def add_group():
    name = request.form['name']
    # Removed level - now using subjects
    instructor_id = int(request.form['instructor_id'])
    max_students = int(request.form['max_students'])
    price = float(request.form['price'])
    
    # Monthly payment fields
    monthly_price = float(request.form['monthly_price']) if request.form.get('monthly_price') else 0.0
    payment_due_day = int(request.form.get('payment_due_day', 1))
    monthly_payment_enabled = request.form.get('monthly_payment_enabled', 'true') == 'true'
    
    force_save = request.form.get('force_save', 'false') == 'true'
    
    # Collect schedule data for conflict checking
    selected_days = request.form.getlist('days[]')
    schedules_to_add = []
    
    for day in selected_days:
        day_prefix = {
            'السبت': 'sat',
            'الأحد': 'sun', 
            'الاثنين': 'mon',
            'الثلاثاء': 'tue',
            'الأربعاء': 'wed',
            'الخميس': 'thu',
            'الجمعة': 'fri'
        }.get(day)
        
        if not day_prefix:
            continue
            
        hour = request.form.get(f'{day_prefix}_hour')
        minute = request.form.get(f'{day_prefix}_minute')
        period = request.form.get(f'{day_prefix}_period')
        duration = request.form.get(f'{day_prefix}_duration')
        
        if hour and minute and period and duration:
            start_time = convert_12_to_24_hour(hour, minute, period)
            
            # Calculate end time based on duration
            duration_minutes = int(duration)
            start_total_minutes = int(start_time.split(':')[0]) * 60 + int(start_time.split(':')[1])
            end_total_minutes = start_total_minutes + duration_minutes
            end_hour = (end_total_minutes // 60) % 24
            end_minute = end_total_minutes % 60
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            schedules_to_add.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time
            })
    
    # Check for instructor schedule conflicts if not forcing save
    all_conflicts = []
    if not force_save and schedules_to_add:
        for schedule_data in schedules_to_add:
            conflicts = check_instructor_schedule_conflicts(
                schedule_data['day'], 
                schedule_data['start_time'], 
                schedule_data['end_time'], 
                instructor_id
            )
            all_conflicts.extend(conflicts)
        
        if all_conflicts:
            # Get instructor name
            instructor = Instructor.query.get(instructor_id)
            instructor_name = instructor.name if instructor else "غير محدد"
            
            # Return conflict information to frontend
            conflict_message = f"المدرس <strong>{instructor_name}</strong> لديه مجموعة أخرى في نفس التوقيت:<br>"
            for conflict in all_conflicts:
                start_12 = convert_24_to_12_hour(conflict['start_time'])
                end_12 = convert_24_to_12_hour(conflict['end_time'])
                conflict_message += f"• مجموعة {conflict['group_name']} - {conflict['day']}: {start_12['hour']}:{start_12['minute']} {start_12['period']} - {end_12['hour']}:{end_12['minute']} {end_12['period']}<br>"
            
            return jsonify({
                'has_conflicts': True,
                'message': conflict_message,
                'form_data': dict(request.form)
            })
    
    # Create and save group
    group = Group(
        name=name,
        # Removed level parameter
        instructor_id=instructor_id,
        max_students=max_students,
        price=price,
        monthly_price=monthly_price,
        payment_due_day=payment_due_day,
        monthly_payment_enabled=monthly_payment_enabled
    )
    
    db.session.add(group)
    db.session.commit()
    
    # Add schedules
    for schedule_data in schedules_to_add:
        schedule = Schedule(
            group_id=group.id,
            day_of_week=schedule_data['day'],
            start_time=schedule_data['start_time'],
            end_time=schedule_data['end_time']
        )
        db.session.add(schedule)
    
    db.session.commit()
    flash('تم إضافة المجموعة بنجاح', 'success')
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'redirect': url_for('groups')})
    return redirect(url_for('groups'))

@app.route('/attendance')
@attendance_required
def attendance():
    groups = Group.query.all()
    students = Student.query.all()
    today = datetime.now().date()
    return render_template('attendance.html', groups=groups, students=students, today=today)

@app.route('/mark_attendance', methods=['POST'])
@attendance_required
def mark_attendance():
    data = request.get_json()
    date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    group_id = data['group_id']
    
    updated_students = set()  # Track students whose attendance was updated
    
    for student_data in data['students']:
        student_id = student_data['student_id']
        status = student_data['status']
        
        # Check if attendance already exists
        existing = Attendance.query.filter_by(
            student_id=student_id,
            date=date,
            group_id=group_id
        ).first()
        
        if existing:
            existing.status = status
        else:
            attendance = Attendance(
                student_id=student_id,
                date=date,
                status=status,
                group_id=group_id
            )
            db.session.add(attendance)
        
        updated_students.add(student_id)
    
    db.session.commit()
    
    # Auto-update achievement points for affected students
    for student_id in updated_students:
        try:
            student = Student.query.get(student_id)
            if student:
                student.update_achievement_points()
        except Exception as e:
            # Log error but don't fail the attendance update
            print(f"Error updating achievement points for student {student_id}: {e}")
    
    return jsonify({'success': True, 'message': 'تم حفظ الحضور بنجاح'})

@app.route('/payments')
@payments_required
def payments():
    # Get pagination parameters
    payments_page = request.args.get('payments_page', 1, type=int)
    expenses_page = request.args.get('expenses_page', 1, type=int)
    per_page = 10  # Number of items per page
    
    # Get search parameters
    search_student = request.args.get('search_student', '')
    search_month = request.args.get('search_month', '')
    search_amount_min = request.args.get('search_amount_min', '')
    search_amount_max = request.args.get('search_amount_max', '')
    search_date_from = request.args.get('search_date_from', '')
    search_date_to = request.args.get('search_date_to', '')
    
    # Expense search parameters
    search_description = request.args.get('search_description', '')
    search_category = request.args.get('search_category', '')
    search_expense_amount_min = request.args.get('search_expense_amount_min', '')
    search_expense_amount_max = request.args.get('search_expense_amount_max', '')
    search_expense_date_from = request.args.get('search_expense_date_from', '')
    search_expense_date_to = request.args.get('search_expense_date_to', '')
    
    students = Student.query.all()
    
    # Build payment query with filters
    payment_query = Payment.query
    
    # Apply payment filters
    if search_student:
        student_ids = [s.id for s in students if search_student.lower() in s.name.lower()]
        if student_ids:
            payment_query = payment_query.filter(Payment.student_id.in_(student_ids))
        else:
            payment_query = payment_query.filter(Payment.student_id == -1)  # No results
    
    if search_month:
        payment_query = payment_query.filter(Payment.month.ilike(f'%{search_month}%'))
    
    if search_amount_min:
        try:
            payment_query = payment_query.filter(Payment.amount >= float(search_amount_min))
        except ValueError:
            pass
    
    if search_amount_max:
        try:
            payment_query = payment_query.filter(Payment.amount <= float(search_amount_max))
        except ValueError:
            pass
    
    if search_date_from:
        try:
            date_from = datetime.strptime(search_date_from, '%Y-%m-%d')
            payment_query = payment_query.filter(Payment.date >= date_from)
        except ValueError:
            pass
    
    if search_date_to:
        try:
            date_to = datetime.strptime(search_date_to, '%Y-%m-%d')
            # Add 1 day to include the end date
            date_to = date_to + timedelta(days=1)
            payment_query = payment_query.filter(Payment.date < date_to)
        except ValueError:
            pass
    
    # Build expense query with filters
    expense_query = Expense.query
    
    # Apply expense filters
    if search_description:
        expense_query = expense_query.filter(Expense.description.ilike(f'%{search_description}%'))
    
    if search_category:
        expense_query = expense_query.filter(Expense.category.ilike(f'%{search_category}%'))
    
    if search_expense_amount_min:
        try:
            expense_query = expense_query.filter(Expense.amount >= float(search_expense_amount_min))
        except ValueError:
            pass
    
    if search_expense_amount_max:
        try:
            expense_query = expense_query.filter(Expense.amount <= float(search_expense_amount_max))
        except ValueError:
            pass
    
    if search_expense_date_from:
        try:
            date_from = datetime.strptime(search_expense_date_from, '%Y-%m-%d')
            expense_query = expense_query.filter(Expense.date >= date_from)
        except ValueError:
            pass
    
    if search_expense_date_to:
        try:
            date_to = datetime.strptime(search_expense_date_to, '%Y-%m-%d')
            # Add 1 day to include the end date
            date_to = date_to + timedelta(days=1)
            expense_query = expense_query.filter(Expense.date < date_to)
        except ValueError:
            pass
    
    # Paginated payments and expenses with filters
    payments_paginated = payment_query.order_by(Payment.date.desc()).paginate(
        page=payments_page, 
        per_page=per_page, 
        error_out=False
    )
    
    expenses_paginated = expense_query.order_by(Expense.date.desc()).paginate(
        page=expenses_page, 
        per_page=per_page, 
        error_out=False
    )
    
    # All payments and expenses for calculations (without filters for totals)
    all_payments = Payment.query.all()
    all_expenses = Expense.query.all()
    
    # Calculate comprehensive statistics
    total_income = sum(payment.amount for payment in all_payments) if all_payments else 0
    total_expenses = sum(expense.amount for expense in all_expenses) if all_expenses else 0
    net_balance = total_income - total_expenses
    
    students_with_dues = sum(1 for student in students if student.remaining_balance > 0)
    recent_payments = len([p for p in all_payments if (datetime.now() - p.date).days <= 30]) if all_payments else 0
    recent_expenses = len([e for e in all_expenses if (datetime.now() - e.date).days <= 30]) if all_expenses else 0
    
    # Monthly breakdown for current year
    current_year = datetime.now().year
    monthly_income = {}
    monthly_expenses = {}
    
    # Get monthly income
    for payment in all_payments:
        if payment.date.year == current_year:
            month = payment.date.month
            monthly_income[month] = monthly_income.get(month, 0) + payment.amount
    
    # Get monthly expenses  
    for expense in all_expenses:
        if expense.date.year == current_year:
            month = expense.date.month
            monthly_expenses[month] = monthly_expenses.get(month, 0) + expense.amount
    
    # Monthly and group breakdown for revenue
    monthly_group_income = {}
    group_monthly_income = {}
    
    # Get all groups
    groups = Group.query.all()
    
    # Create monthly breakdown by groups
    for payment in all_payments:
        student = Student.query.get(payment.student_id)
        if student and payment.date.year == current_year:
            month = payment.date.month
            month_name = get_arabic_month_name(month)
            
            if month_name not in monthly_group_income:
                monthly_group_income[month_name] = {}
            
            # Add payment to each group the student belongs to
            for group in student.groups:
                if group.name not in monthly_group_income[month_name]:
                    monthly_group_income[month_name][group.name] = 0
                # Distribute payment amount evenly among student's groups
                amount_per_group = payment.amount / len(student.groups)
                monthly_group_income[month_name][group.name] += amount_per_group
                
                # Also create group-based monthly breakdown
                if group.name not in group_monthly_income:
                    group_monthly_income[group.name] = {}
                if month_name not in group_monthly_income[group.name]:
                    group_monthly_income[group.name][month_name] = 0
                group_monthly_income[group.name][month_name] += amount_per_group
    
    return render_template('payments.html', 
                         students=students, 
                         payments=payments_paginated.items,
                         expenses=expenses_paginated.items,
                         payments_pagination=payments_paginated,
                         expenses_pagination=expenses_paginated,
                         total_income=total_income,
                         total_expenses=total_expenses,
                         net_balance=net_balance,
                         students_with_dues=students_with_dues,
                         recent_payments=recent_payments,
                         recent_expenses=recent_expenses,
                         monthly_income=monthly_income,
                         monthly_expenses=monthly_expenses,
                         monthly_group_income=monthly_group_income,
                         group_monthly_income=group_monthly_income,
                         groups=groups,
                         # Model classes for template access
                         Payment=Payment,
                         Expense=Expense,
                         # Search parameters for payments
                         search_student=search_student,
                         search_month=search_month,
                         search_amount_min=search_amount_min,
                         search_amount_max=search_amount_max,
                         search_date_from=search_date_from,
                         search_date_to=search_date_to,
                         # Search parameters for expenses
                         search_description=search_description,
                         search_category=search_category,
                         search_expense_amount_min=search_expense_amount_min,
                         search_expense_amount_max=search_expense_amount_max,
                         search_expense_date_from=search_expense_date_from,
                         search_expense_date_to=search_expense_date_to)

@app.route('/add_payment', methods=['POST'])
@payments_required
def add_payment():
    student_id = int(request.form['student_id'])
    amount = float(request.form['amount'])
    month = request.form['month']
    notes = request.form['notes']
    
    payment = Payment(
        student_id=student_id,
        amount=amount,
        month=month,
        notes=notes
    )
    
    # Update student's total paid
    student = Student.query.get(student_id)
    student.total_paid += amount
    
    db.session.add(payment)
    db.session.commit()
    flash('تم إضافة الدفعة بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/edit_payment/<int:payment_id>', methods=['POST'])
@payments_required
def edit_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    old_amount = payment.amount
    old_student_id = payment.student_id
    
    # Get new values
    new_student_id = int(request.form['student_id'])
    new_amount = float(request.form['amount'])
    new_month = request.form['month']
    new_notes = request.form['notes']
    
    # Update student's total paid - subtract old amount and add new amount
    # Handle case where student might have changed
    if old_student_id == new_student_id:
        # Same student - adjust the difference
        student = Student.query.get(old_student_id)
        student.total_paid = student.total_paid - old_amount + new_amount
    else:
        # Different student - subtract from old, add to new
        old_student = Student.query.get(old_student_id)
        old_student.total_paid -= old_amount
        new_student = Student.query.get(new_student_id)
        new_student.total_paid += new_amount
    
    # Update payment
    payment.student_id = new_student_id
    payment.amount = new_amount
    payment.month = new_month
    payment.notes = new_notes
    
    db.session.commit()
    flash('تم تحديث الدفعة بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/delete_payment/<int:payment_id>', methods=['POST'])
@payments_required
def delete_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    
    # Update student's total paid - subtract the payment amount
    student = Student.query.get(payment.student_id)
    student.total_paid -= payment.amount
    
    # Delete the payment
    db.session.delete(payment)
    db.session.commit()
    
    flash('تم حذف الدفعة بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/reports')
@reports_required
def reports():
    # Attendance statistics
    total_students = Student.query.count()
    today = datetime.now().date()
    present_today = Attendance.query.filter_by(date=today, status='حاضر').count()
    absent_today = Attendance.query.filter_by(date=today, status='غائب').count()
    
    # Payment statistics
    total_revenue = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
    
    # Calculate pending payments based on group-based pricing after discounts
    pending_payments = 0
    students = Student.query.all()
    for student in students:
        if student.remaining_balance > 0:
            pending_payments += student.remaining_balance
    
    # Other statistics
    groups_count = Group.query.count()
    instructors_count = Instructor.query.count()
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    # Additional useful statistics - calculate expected revenue after discounts
    total_groups_revenue = sum(student.total_course_price_after_discount for student in Student.query.all())
    late_today = Attendance.query.filter_by(date=today, status='متأخر').count()
    
    # Monthly statistics for the current year
    current_year = datetime.now().year
    monthly_payments = {}
    monthly_expenses = {}
    
    # Get monthly payment data
    payments = Payment.query.filter(
        db.extract('year', Payment.date) == current_year
    ).all()
    
    for payment in payments:
        month = payment.date.month
        monthly_payments[month] = monthly_payments.get(month, 0) + payment.amount
    
    # Get monthly expense data
    expenses = Expense.query.filter(
        db.extract('year', Expense.date) == current_year
    ).all()
    
    for expense in expenses:
        month = expense.date.month
        monthly_expenses[month] = monthly_expenses.get(month, 0) + expense.amount
    
    # Get groups data for health check
    groups_count_list = Group.query.all()
    
    return render_template('reports.html',
                         total_students=total_students,
                         present_today=present_today,
                         absent_today=absent_today,
                         late_today=late_today,
                         total_revenue=total_revenue,
                         pending_payments=pending_payments,
                         groups_count=groups_count,
                         instructors_count=instructors_count,
                         today_date=datetime.now(),  # Pass datetime object instead of string
                         total_groups_revenue=total_groups_revenue,
                         monthly_payments=monthly_payments,
                         monthly_expenses=monthly_expenses,
                         groups_count_list=groups_count_list)

@app.route('/export_reports')
@login_required
def export_reports():
    """Export comprehensive reports to Excel file"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير شامل"
        
        # Set RTL direction for Arabic support
        ws.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        sub_header_font = Font(size=12, bold=True, color="2F5F8F")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"تقرير شامل - مركز تفرا التعليمي - {format_arabic_date(datetime.now())}"
        title_cell.font = Font(size=16, bold=True, color="2F5F8F")
        title_cell.alignment = center_alignment
        current_row += 2
        
        # Basic Statistics Section
        ws[f'A{current_row}'] = "الإحصائيات الأساسية"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Get statistics data
        total_students = Student.query.count()
        instructors_count = Instructor.query.count()
        groups_count = Group.query.count()
        today = datetime.now().date()
        present_today = Attendance.query.filter_by(date=today, status='حاضر').count()
        absent_today = Attendance.query.filter_by(date=today, status='غائب').count()
        late_today = Attendance.query.filter_by(date=today, status='متأخر').count()
        
        # Add basic statistics
        stats_data = [
            ['البيان', 'القيمة'],
            ['إجمالي الطلاب', total_students],
            ['عدد المدرسين', instructors_count],
            ['عدد المجموعات', groups_count],
            ['حاضر اليوم', present_today],
            ['غائب اليوم', absent_today],
            ['متأخر اليوم', late_today],
        ]
        
        for row_data in stats_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                if current_row == len(stats_data) + current_row - len(stats_data):  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
            current_row += 1
        
        current_row += 2
        
        # Financial Statistics Section
        ws[f'A{current_row}'] = "الإحصائيات المالية"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Get financial data
        total_revenue = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
        total_expenses = db.session.query(db.func.sum(Expense.amount)).scalar() or 0
        pending_payments = sum(student.remaining_balance for student in Student.query.all() if student.remaining_balance > 0)
        
        financial_data = [
            ['البيان المالي', 'المبلغ (ريال)'],
            ['إجمالي الإيرادات', f"{total_revenue:,.0f}"],
            ['إجمالي المصروفات', f"{total_expenses:,.0f}"],
            ['صافي الربح', f"{total_revenue - total_expenses:,.0f}"],
            ['مدفوعات معلقة', f"{pending_payments:,.0f}"],
        ]
        
        for row_data in financial_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                if current_row == len(financial_data) + current_row - len(financial_data):  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
            current_row += 1
        
        current_row += 2
        
        # Students Data Section
        ws[f'A{current_row}'] = "بيانات الطلاب"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Students headers
        student_headers = ['#', 'اسم الطالب', 'العمر', 'الموقع', 'المجموعات', 'المدفوع', 'المتبقي', 'تاريخ التسجيل']
        for col, header in enumerate(student_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # Students data
        students = Student.query.all()
        for idx, student in enumerate(students, 1):
            groups_names = ', '.join([group.name for group in student.groups])
            student_data = [
                idx,
                student.name,
                student.age or 'غير محدد',
                student.location or 'غير محدد',
                groups_names or 'لا توجد مجموعات',
                f"{student.total_paid:,.0f}",
                f"{student.remaining_balance:,.0f}",
                student.registration_date.strftime('%Y-%m-%d') if student.registration_date else 'غير محدد'
            ]
            
            for col, value in enumerate(student_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
            current_row += 1
        
        current_row += 2
        
        # Groups Data Section
        ws[f'A{current_row}'] = "بيانات المجموعات"
        ws[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        # Groups headers
        group_headers = ['#', 'اسم المجموعة', 'المستوى', 'المدرس', 'عدد الطلاب', 'الحد الأقصى', 'السعر']
        for col, header in enumerate(group_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # Groups data
        groups = Group.query.all()
        for idx, group in enumerate(groups, 1):
            instructor_name = group.instructor_ref.name if group.instructor_ref else 'غير محدد'
            group_data = [
                idx,
                group.name,
                ', '.join([s.name for s in group.subjects]) if group.subjects else 'غير محدد',
                instructor_name,
                group.students.count(),
                group.max_students,
                f"{group.price:,.0f}"
            ]
            
            for col, value in enumerate(group_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"تقرير_شامل_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/get_group_students/<int:group_id>')
def get_group_students(group_id):
    group = Group.query.get_or_404(group_id)
    student_list = []
    for student in group.students:
        student_list.append({
            'id': student.id,
            'name': student.name
        })
    return jsonify(student_list)

@app.route('/edit_student/<int:student_id>', methods=['POST'])
def edit_student(student_id):
    try:
        student = Student.query.get_or_404(student_id)
        student.name = request.form['name']
        student.phone = request.form.get('phone', '')
        student.age = int(request.form['age'])
        student.location = request.form.get('location', '')
        student.grade_level = request.form.get('grade_level', '')  # المرحلة الدراسية
        student.instructor_id = None  # Keep instructor optional
        registration_date = parse_date_from_input(request.form['registration_date'])
        student.discount = float(request.form.get('discount', 0))  # Get discount amount, default to 0
        
        if not registration_date:
            flash('يرجى اختيار تاريخ التسجيل من منتقي التاريخ', 'error')
            return redirect(url_for('students'))
            
        student.registration_date = registration_date
        
        # Clear existing group associations
        student.groups.clear()
        
        # Handle group selections (multiple groups allowed)
        group_ids = request.form.getlist('group_ids')  # Get list of selected group IDs
        if group_ids:
            for group_id in group_ids:
                if group_id:  # Make sure it's not empty
                    group = Group.query.get(int(group_id))
                    if group:
                        student.groups.append(group)
        
        db.session.commit()
        flash('تم تحديث بيانات الطالب بنجاح!', 'success')
        return redirect(url_for('students'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث الطالب: {str(e)}', 'error')
        return redirect(url_for('students'))

@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Delete related attendance records
    Attendance.query.filter_by(student_id=student_id).delete()
    # Delete related payment records
    Payment.query.filter_by(student_id=student_id).delete()
    
    db.session.delete(student)
    db.session.commit()
    flash('تم حذف الطالب بنجاح', 'success')
    return redirect(url_for('students'))

@app.route('/bulk_delete_students', methods=['POST'])
def bulk_delete_students():
    try:
        data = request.get_json()
        student_ids = data.get('student_ids', [])
        
        if not student_ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد أي طلاب'})
        
        # التحقق من وجود الطلاب وحذفهم
        students_deleted = 0
        for student_id in student_ids:
            student = Student.query.get(student_id)
            if student:
                # Delete related attendance records
                Attendance.query.filter_by(student_id=student_id).delete()
                # Delete related payment records
                Payment.query.filter_by(student_id=student_id).delete()
                # Delete the student
                db.session.delete(student)
                students_deleted += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'تم حذف {students_deleted} طالب بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ أثناء الحذف: {str(e)}'
        })

@app.route('/bulk_edit_group', methods=['POST'])
def bulk_edit_group():
    try:
        student_ids = request.form.get('student_ids', '').split(',')
        student_ids = [int(id.strip()) for id in student_ids if id.strip()]
        group_id = request.form.get('group_id')
        operation = request.form.get('operation', 'add')
        
        if not student_ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد أي طلاب'})
        
        if not group_id:
            return jsonify({'success': False, 'message': 'لم يتم تحديد مجموعة'})
        
        group = Group.query.get(group_id)
        if not group:
            return jsonify({'success': False, 'message': 'المجموعة غير موجودة'})
        
        students_updated = 0
        
        for student_id in student_ids:
            student = Student.query.get(student_id)
            if not student:
                continue
                
            if operation == 'add':
                # إضافة إلى مجموعة (إذا لم يكن مضافاً بالفعل)
                if group not in student.groups:
                    student.groups.append(group)
                    students_updated += 1
                    
            elif operation == 'remove':
                # إزالة من مجموعة
                if group in student.groups:
                    student.groups.remove(group)
                    students_updated += 1
                    
            elif operation == 'replace':
                # استبدال المجموعات (إزالة الحالية وإضافة الجديدة)
                student.groups.clear()
                student.groups.append(group)
                students_updated += 1
        
        db.session.commit()
        
        operation_messages = {
            'add': f'تم إضافة {students_updated} طالب إلى مجموعة {group.name}',
            'remove': f'تم إزالة {students_updated} طالب من مجموعة {group.name}',
            'replace': f'تم تحديث مجموعات {students_updated} طالب إلى {group.name}'
        }
        
        return jsonify({
            'success': True, 
            'message': operation_messages.get(operation, 'تم التحديث بنجاح')
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ أثناء التحديث: {str(e)}'
        })

@app.route('/edit_instructor/<int:instructor_id>', methods=['POST'])
def edit_instructor(instructor_id):
    instructor = Instructor.query.get_or_404(instructor_id)
    
    instructor.name = request.form['name']
    instructor.phone = request.form['phone']
    instructor.specialization = request.form['specialization']
    
    db.session.commit()
    flash('تم تحديث بيانات المدرس بنجاح', 'success')
    return redirect(url_for('instructors'))

@app.route('/delete_instructor/<int:instructor_id>', methods=['POST'])
def delete_instructor(instructor_id):
    instructor = Instructor.query.get_or_404(instructor_id)
    
    # Check if instructor has students or groups
    if instructor.students or instructor.groups:
        flash('لا يمكن حذف المدرس لأنه مرتبط بطلاب أو مجموعات', 'error')
        return redirect(url_for('instructors'))
    
    db.session.delete(instructor)
    db.session.commit()
    flash('تم حذف المدرس بنجاح', 'success')
    return redirect(url_for('instructors'))

@app.route('/edit_group/<int:group_id>', methods=['POST'])
def edit_group(group_id):
    group = Group.query.get_or_404(group_id)
    force_save = request.form.get('force_save', 'false') == 'true'
    
    # Update basic group information
    group.name = request.form['name']
    # Removed level update - now using subjects
    new_instructor_id = int(request.form['instructor_id'])
    group.price = float(request.form['price'])
    group.max_students = int(request.form['max_students'])
    
    # Collect schedule data for conflict checking
    selected_days = request.form.getlist('days[]')
    schedules_to_add = []
    
    for day in selected_days:
        day_prefix = {
            'السبت': 'sat',
            'الأحد': 'sun', 
            'الاثنين': 'mon',
            'الثلاثاء': 'tue',
            'الأربعاء': 'wed',
            'الخميس': 'thu',
            'الجمعة': 'fri'
        }.get(day)
        
        if not day_prefix:
            continue
            
        hour = request.form.get(f'{day_prefix}_hour')
        minute = request.form.get(f'{day_prefix}_minute')
        period = request.form.get(f'{day_prefix}_period')
        duration = request.form.get(f'{day_prefix}_duration')
        
        if hour and minute and period and duration:
            start_time = convert_12_to_24_hour(hour, minute, period)
            
            # Calculate end time based on duration
            duration_minutes = int(duration)
            start_total_minutes = int(start_time.split(':')[0]) * 60 + int(start_time.split(':')[1])
            end_total_minutes = start_total_minutes + duration_minutes
            end_hour = (end_total_minutes // 60) % 24
            end_minute = end_total_minutes % 60
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            schedules_to_add.append({
                'day': day,
                'start_time': start_time,
                'end_time': end_time
            })
    
    # Check for instructor schedule conflicts if not forcing save
    all_conflicts = []
    if not force_save and schedules_to_add:
        for schedule_data in schedules_to_add:
            conflicts = check_instructor_schedule_conflicts(
                schedule_data['day'], 
                schedule_data['start_time'], 
                schedule_data['end_time'], 
                new_instructor_id,
                exclude_group_id=group_id
            )
            all_conflicts.extend(conflicts)
        
        if all_conflicts:
            # Get instructor name
            instructor = Instructor.query.get(new_instructor_id)
            instructor_name = instructor.name if instructor else "غير محدد"
            
            # Return conflict information to frontend
            conflict_message = f"المدرس <strong>{instructor_name}</strong> لديه مجموعة أخرى في نفس التوقيت:<br>"
            for conflict in all_conflicts:
                start_12 = convert_24_to_12_hour(conflict['start_time'])
                end_12 = convert_24_to_12_hour(conflict['end_time'])
                conflict_message += f"• مجموعة {conflict['group_name']} - {conflict['day']}: {start_12['hour']}:{start_12['minute']} {start_12['period']} - {end_12['hour']}:{end_12['minute']} {end_12['period']}<br>"
            
            return jsonify({
                'has_conflicts': True,
                'message': conflict_message,
                'form_data': dict(request.form),
                'group_id': group_id
            })
    
    # Update instructor
    group.instructor_id = new_instructor_id
    
    # Delete existing schedules
    Schedule.query.filter_by(group_id=group_id).delete()
    
    # Add new schedules
    for schedule_data in schedules_to_add:
        schedule = Schedule(
            group_id=group.id,
            day_of_week=schedule_data['day'],
            start_time=schedule_data['start_time'],
            end_time=schedule_data['end_time']
        )
        db.session.add(schedule)
    
    db.session.commit()
    flash('تم تحديث بيانات المجموعة والجداول بنجاح', 'success')
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'redirect': url_for('groups')})
    return redirect(url_for('groups'))

@app.route('/delete_group/<int:group_id>', methods=['POST'])
def delete_group(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check if group has students using the new many-to-many relationship
    student_count = group.students.count()
    if student_count > 0:
        flash('لا يمكن حذف المجموعة لأنها تحتوي على طلاب', 'error')
        return redirect(url_for('groups'))
    
    # Delete related schedules
    Schedule.query.filter_by(group_id=group_id).delete()
    # Delete related attendance records
    Attendance.query.filter_by(group_id=group_id).delete()
    
    db.session.delete(group)
    db.session.commit()
    flash('تم حذف المجموعة بنجاح', 'success')
    return redirect(url_for('groups'))

@app.route('/get_group_details/<int:group_id>')
def get_group_details(group_id):
    group = Group.query.get_or_404(group_id)
    schedules = []
    for schedule in group.schedules:
        schedules.append({
            'day': schedule.day_of_week,
            'start_time': schedule.start_time,
            'end_time': schedule.end_time
        })
    
    return jsonify({
        'id': group.id,
        'name': group.name,
        'subjects': [{'id': s.id, 'name': s.name, 'type': s.subject_type} for s in group.subjects],
        'instructor_id': group.instructor_id,
        'max_students': group.max_students,
        'price': group.price,
        'status': group.status,
        'completion_date': group.completion_date.strftime('%Y-%m-%d') if group.completion_date else None,
        'completion_notes': group.completion_notes,
        'schedules': schedules
    })

@app.route('/complete_group/<int:group_id>', methods=['POST'])
@login_required
def complete_group(group_id):
    group = Group.query.get_or_404(group_id)
    
    if group.status == 'completed':
        flash('هذه المجموعة مكتملة بالفعل', 'warning')
        return redirect(url_for('groups'))
    
    # Get completion data
    completion_notes = request.form.get('completion_notes', '')
    completion_date = request.form.get('completion_date')
    
    # Parse completion date
    if completion_date:
        try:
            completion_date = datetime.strptime(completion_date, '%Y-%m-%d').date()
        except ValueError:
            completion_date = datetime.now().date()
    else:
        completion_date = datetime.now().date()
    
    # Update group status
    group.status = 'completed'
    group.completion_date = completion_date
    group.completion_notes = completion_notes
    
    db.session.commit()
    flash(f'تم إنهاء المجموعة "{group.name}" بنجاح', 'success')
    return redirect(url_for('groups'))

@app.route('/activate_group/<int:group_id>', methods=['POST'])
@login_required
def activate_group(group_id):
    group = Group.query.get_or_404(group_id)
    
    if group.status == 'active':
        flash('هذه المجموعة نشطة بالفعل', 'warning')
        return redirect(url_for('groups'))
    
    # Reactivate group
    group.status = 'active'
    group.completion_date = None
    group.completion_notes = None
    
    db.session.commit()
    flash(f'تم تفعيل المجموعة "{group.name}" بنجاح', 'success')
    return redirect(url_for('groups'))

@app.route('/update_group_completion/<int:group_id>', methods=['POST'])
@login_required
def update_group_completion(group_id):
    group = Group.query.get_or_404(group_id)
    
    if group.status != 'completed':
        flash('هذه المجموعة غير مكتملة', 'error')
        return redirect(url_for('groups'))
    
    # Update completion details
    completion_notes = request.form.get('completion_notes', '')
    completion_date = request.form.get('completion_date')
    
    if completion_date:
        try:
            completion_date = datetime.strptime(completion_date, '%Y-%m-%d').date()
        except ValueError:
            completion_date = group.completion_date
    else:
        completion_date = group.completion_date
    
    group.completion_date = completion_date
    group.completion_notes = completion_notes
    
    db.session.commit()
    flash(f'تم تحديث تفاصيل إنهاء المجموعة "{group.name}" بنجاح', 'success')
    return redirect(url_for('groups'))

@app.route('/add_expense', methods=['POST'])
def add_expense():
    description = request.form['description']
    amount = float(request.form['amount'])
    category = request.form['category']
    notes = request.form.get('notes', '')
    
    expense = Expense(
        description=description,
        amount=amount,
        category=category,
        notes=notes
    )
    
    db.session.add(expense)
    db.session.commit()
    flash('تم إضافة المصروف بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/edit_expense/<int:expense_id>', methods=['POST'])
@login_required
def edit_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    
    # Update expense details
    expense.description = request.form['description']
    expense.amount = float(request.form['amount'])
    expense.category = request.form['category']
    expense.notes = request.form.get('notes', '')
    
    db.session.commit()
    flash('تم تحديث المصروف بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/delete_expense/<int:expense_id>', methods=['POST'])
@login_required
def delete_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    
    db.session.delete(expense)
    db.session.commit()
    
    flash('تم حذف المصروف بنجاح', 'success')
    return redirect(url_for('payments'))

@app.route('/bulk_delete_payments', methods=['POST'])
@login_required
def bulk_delete_payments():
    payment_ids = request.form.get('bulk_delete_ids', '')
    
    if not payment_ids:
        flash('لم يتم تحديد أي مدفوعات للحذف', 'error')
        return redirect(url_for('payments'))
    
    try:
        # Convert comma-separated IDs to list of integers
        ids_list = [int(id.strip()) for id in payment_ids.split(',') if id.strip()]
        
        if not ids_list:
            flash('لم يتم تحديد أي مدفوعات للحذف', 'error')
            return redirect(url_for('payments'))
        
        # Get all payments to be deleted
        payments_to_delete = Payment.query.filter(Payment.id.in_(ids_list)).all()
        
        if not payments_to_delete:
            flash('لم يتم العثور على المدفوعات المحددة', 'error')
            return redirect(url_for('payments'))
        
        # Update students' total_paid before deleting payments
        for payment in payments_to_delete:
            student = Student.query.get(payment.student_id)
            if student:
                student.total_paid -= payment.amount
        
        # Delete all selected payments
        Payment.query.filter(Payment.id.in_(ids_list)).delete(synchronize_session=False)
        
        db.session.commit()
        flash(f'تم حذف {len(payments_to_delete)} مدفوعة بنجاح', 'success')
        
    except ValueError:
        flash('خطأ في معرفات المدفوعات', 'error')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء حذف المدفوعات', 'error')
    
    return redirect(url_for('payments'))

@app.route('/bulk_delete_expenses', methods=['POST'])
@login_required
def bulk_delete_expenses():
    expense_ids = request.form.get('bulk_delete_ids', '')
    
    if not expense_ids:
        flash('لم يتم تحديد أي مصروفات للحذف', 'error')
        return redirect(url_for('payments'))
    
    try:
        # Convert comma-separated IDs to list of integers
        ids_list = [int(id.strip()) for id in expense_ids.split(',') if id.strip()]
        
        if not ids_list:
            flash('لم يتم تحديد أي مصروفات للحذف', 'error')
            return redirect(url_for('payments'))
        
        # Get all expenses to be deleted for counting
        expenses_to_delete = Expense.query.filter(Expense.id.in_(ids_list)).all()
        
        if not expenses_to_delete:
            flash('لم يتم العثور على المصروفات المحددة', 'error')
            return redirect(url_for('payments'))
        
        # Delete all selected expenses
        Expense.query.filter(Expense.id.in_(ids_list)).delete(synchronize_session=False)
        
        db.session.commit()
        flash(f'تم حذف {len(expenses_to_delete)} مصروف بنجاح', 'success')
        
    except ValueError:
        flash('خطأ في معرفات المصروفات', 'error')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء حذف المصروفات', 'error')
    
    return redirect(url_for('payments'))

@app.route('/group_details/<int:group_id>')
@login_required
def group_details(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Get all students in this group
    students = group.students.all()
    
    # Get all attendance records for this group
    attendance_records = Attendance.query.filter_by(group_id=group_id).all()
    
    # Calculate attendance statistics for the group
    total_sessions = len(set(record.date for record in attendance_records))
    total_attendances = len([record for record in attendance_records if record.status == 'حاضر'])
    total_absences = len([record for record in attendance_records if record.status == 'غائب'])
    total_late = len([record for record in attendance_records if record.status == 'متأخر'])
    
    # Calculate attendance percentage
    attendance_percentage = (total_attendances / len(attendance_records) * 100) if attendance_records else 0
    
    # Get unique dates when sessions happened
    session_dates = sorted(set(record.date for record in attendance_records), reverse=True)
    
    # Create attendance matrix for each student
    student_attendance = {}
    for student in students:
        student_records = [record for record in attendance_records if record.student_id == student.id]
        
        # Calculate student statistics
        student_present = len([r for r in student_records if r.status == 'حاضر'])
        student_absent = len([r for r in student_records if r.status == 'غائب'])
        student_late = len([r for r in student_records if r.status == 'متأخر'])
        student_percentage = (student_present / len(student_records) * 100) if student_records else 0
        
        # Create date-wise attendance - convert dates to strings for JSON serialization
        attendance_by_date = {}
        for record in student_records:
            attendance_by_date[record.date.strftime('%Y-%m-%d')] = record.status
        
        student_attendance[student.id] = {
            'student': {
                'id': student.id,
                'name': student.name,
                'phone': student.phone,
                'location': student.location
            },
            'total_present': student_present,
            'total_absent': student_absent,
            'total_late': student_late,
            'total_sessions': len(student_records),
            'percentage': round(student_percentage, 1),
            'attendance_by_date': attendance_by_date
        }
    
    # Get recent payments for this group's students
    student_ids = [s.id for s in students]
    recent_payments = Payment.query.filter(
        Payment.student_id.in_(student_ids)
    ).order_by(Payment.date.desc()).limit(10).all() if student_ids else []
    
    # Calculate financial statistics - use prices after discount
    total_expected_revenue = sum(student.total_course_price_after_discount for student in students)
    total_received_revenue = sum(student.total_paid for student in students)
    pending_revenue = total_expected_revenue - total_received_revenue
    
    return render_template('group_details.html',
                         group=group,
                         students=students,
                         session_dates=[date.strftime('%Y-%m-%d') for date in session_dates],
                         student_attendance=student_attendance,
                         total_sessions=total_sessions,
                         total_attendances=total_attendances,
                         total_absences=total_absences,
                         total_late=total_late,
                         attendance_percentage=round(attendance_percentage, 1),
                         recent_payments=recent_payments,
                         total_expected_revenue=total_expected_revenue,
                         total_received_revenue=total_received_revenue,
                         pending_revenue=pending_revenue)

@app.route('/add_sample_attendance')
@admin_required
def add_sample_attendance():
    """Add sample attendance data for testing - Admin only"""
    from datetime import date, timedelta
    
    # Get all groups and their students
    groups = Group.query.all()
    
    # Generate attendance for the last 30 days
    start_date = date.today() - timedelta(days=30)
    
    for group in groups:
        students = group.students.all()
        if not students:
            continue
            
        # Generate attendance for each day in the last 30 days
        for i in range(30):
            current_date = start_date + timedelta(days=i)
            
            # Skip weekends (Friday and Saturday in Middle East)
            if current_date.weekday() in [4, 5]:  # Friday and Saturday
                continue
                
            for student in students:
                # Check if attendance already exists
                existing = Attendance.query.filter_by(
                    student_id=student.id,
                    date=current_date,
                    group_id=group.id
                ).first()
                
                if not existing:
                    # Generate random attendance status
                    # 70% present, 20% absent, 10% late
                    rand = random.random()
                    if rand < 0.7:
                        status = 'حاضر'
                    elif rand < 0.9:
                        status = 'غائب'
                    else:
                        status = 'متأخر'
                    
                    attendance = Attendance(
                        student_id=student.id,
                        date=current_date,
                        status=status,
                        group_id=group.id
                    )
                    db.session.add(attendance)
    
    db.session.commit()
    flash('تم إضافة بيانات الحضور التجريبية بنجاح!', 'success')
    return redirect(url_for('groups'))

def convert_12_to_24_hour(hour, minute, period):
    """Convert 12-hour format to 24-hour format"""
    hour = int(hour)
    minute = int(minute)
    
    if period == 'AM':
        if hour == 12:
            hour = 0
    else:  # PM
        if hour != 12:
            hour += 12
    
    return f"{hour:02d}:{minute:02d}"

def convert_24_to_12_hour(time_24):
    """Convert 24-hour format to 12-hour format"""
    if not time_24 or ':' not in time_24:
        return {'hour': '12', 'minute': '00', 'period': 'AM'}
    
    hour, minute = time_24.split(':')
    hour = int(hour)
    period = 'AM'
    
    if hour == 0:
        hour = 12
    elif hour == 12:
        period = 'PM'
    elif hour > 12:
        hour = hour - 12
        period = 'PM'
    
    return {'hour': str(hour), 'minute': minute, 'period': period}

# Add the function to Jinja2 template context
@app.context_processor
def utility_processor():
    def get_new_instructor_notes_count():
        """Get count of new instructor notes for admin notification"""
        current_user = get_current_user()
        if current_user and current_user.role == 'admin':
            return InstructorNote.query.filter_by(status='جديد').count()
        return 0
    
    return dict(
        convert_24_to_12_hour=convert_24_to_12_hour,
        get_arabic_day_name=get_arabic_day_name,
        format_arabic_date=format_arabic_date,
        format_time_12hour=format_time_12hour,
        format_date_for_input=format_date_for_input,
        get_new_instructor_notes_count=get_new_instructor_notes_count
    )

def init_db():
    """Initialize database and create default admin"""
    with app.app_context():
        db.create_all()
        create_default_admin()

@app.route('/debug')
@login_required
def debug_prices():
    """Debug page to test price calculation"""
    return render_template('debug.html')

@app.route('/tasks')
@tasks_required
def tasks():
    """Display tasks and notes management page"""
    filter_status = request.args.get('status', 'all')
    filter_priority = request.args.get('priority', 'all')
    filter_category = request.args.get('category', 'all')
    
    # Build query based on filters for tasks
    query = Task.query
    
    if filter_status != 'all':
        query = query.filter_by(status=filter_status)
    
    if filter_priority != 'all':
        query = query.filter_by(priority=filter_priority)
    
    # Order by priority and creation date
    priority_order = {'عالي': 3, 'متوسط': 2, 'منخفض': 1}
    tasks = query.all()
    tasks.sort(key=lambda x: (priority_order.get(x.priority, 0), x.created_at), reverse=True)
    
    # Get notes and filter them
    notes_query = Note.query
    
    if filter_category != 'all':
        notes_query = notes_query.filter_by(category=filter_category)
    
    # Order notes by pinned status and creation date
    notes = notes_query.order_by(Note.is_pinned.desc(), Note.updated_at.desc()).all()
    
    # Get instructor notes (for admins only)
    instructor_notes = []
    current_user = get_current_user()
    if current_user.role == 'admin':
        instructor_notes_query = InstructorNote.query
        
        if filter_status != 'all':
            instructor_notes_query = instructor_notes_query.filter_by(status=filter_status)
        
        if filter_priority != 'all':
            instructor_notes_query = instructor_notes_query.filter_by(priority=filter_priority)
        
        instructor_notes = instructor_notes_query.order_by(InstructorNote.created_at.desc()).all()
    
    # Get statistics
    total_tasks = Task.query.count()
    completed_tasks = Task.query.filter_by(status='مكتمل').count()
    pending_tasks = Task.query.filter(Task.status.in_(['قيد التنفيذ'])).count()
    overdue_tasks = len([t for t in Task.query.all() if t.is_overdue])
    
    # Notes statistics
    total_notes = Note.query.count()
    pinned_notes = Note.query.filter_by(is_pinned=True).count()
    
    # Instructor notes statistics (for admins)
    total_instructor_notes = 0
    new_instructor_notes = 0
    if current_user.role == 'admin':
        total_instructor_notes = InstructorNote.query.count()
        new_instructor_notes = InstructorNote.query.filter_by(status='جديد').count()
    
    users = User.query.all()
    
    return render_template('tasks.html',
                         tasks=tasks,
                         notes=notes,
                         instructor_notes=instructor_notes,
                         users=users,
                         current_user=current_user,
                         total_tasks=total_tasks,
                         completed_tasks=completed_tasks,
                         pending_tasks=pending_tasks,
                         overdue_tasks=overdue_tasks,
                         total_notes=total_notes,
                         pinned_notes=pinned_notes,
                         total_instructor_notes=total_instructor_notes,
                         new_instructor_notes=new_instructor_notes,
                         filter_status=filter_status,
                         filter_priority=filter_priority,
                         filter_category=filter_category)

@app.route('/add_task', methods=['POST'])
@login_required
def add_task():
    """Add a new task"""
    try:
        title = request.form['title']
        description = request.form.get('description', '')
        priority = request.form['priority']
        due_date_str = request.form.get('due_date')
        assigned_to = request.form.get('assigned_to')
        
        # Parse due date
        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        
        # Convert assigned_to to int if provided
        assigned_to_id = None
        if assigned_to and assigned_to != '':
            assigned_to_id = int(assigned_to)
        
        current_user = get_current_user()
        
        task = Task(
            title=title,
            description=description,
            priority=priority,
            due_date=due_date,
            created_by=current_user.id,
            assigned_to=assigned_to_id
        )
        
        db.session.add(task)
        db.session.commit()
        
        flash('تم إضافة المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/update_task_status/<int:task_id>', methods=['POST'])
@login_required
def update_task_status(task_id):
    """Update task status"""
    try:
        task = Task.query.get_or_404(task_id)
        new_status = request.form['status']
        
        task.status = new_status
        
        # If marking as completed, set completion time
        if new_status == 'مكتمل':
            task.completed_at = datetime.utcnow()
        else:
            task.completed_at = None
        
        db.session.commit()
        
        flash('تم تحديث حالة المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/edit_task/<int:task_id>', methods=['POST'])
@login_required
def edit_task(task_id):
    """Edit an existing task"""
    try:
        task = Task.query.get_or_404(task_id)
        
        task.title = request.form['title']
        task.description = request.form.get('description', '')
        task.priority = request.form['priority']
        
        # Parse due date
        due_date_str = request.form.get('due_date')
        if due_date_str:
            task.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        else:
            task.due_date = None
        
        # Update assigned user
        assigned_to = request.form.get('assigned_to')
        if assigned_to and assigned_to != '':
            task.assigned_to = int(assigned_to)
        else:
            task.assigned_to = None
        
        db.session.commit()
        
        flash('تم تحديث المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/delete_task/<int:task_id>', methods=['POST'])
@login_required
def delete_task(task_id):
    """Delete a task"""
    try:
        task = Task.query.get_or_404(task_id)
        db.session.delete(task)
        db.session.commit()
        
        flash('تم حذف المهمة بنجاح!', 'success')
        return redirect(url_for('tasks'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف المهمة: {str(e)}', 'error')
        return redirect(url_for('tasks'))

@app.route('/add_note', methods=['POST'])
@login_required
def add_note():
    """Add a new note"""
    try:
        title = request.form['title']
        content = request.form['content']
        category = request.form['category']
        color = request.form['color']
        is_pinned = 'is_pinned' in request.form
        
        current_user = get_current_user()
        
        note = Note(
            title=title,
            content=content,
            category=category,
            color=color,
            is_pinned=is_pinned,
            created_by=current_user.id
        )
        
        db.session.add(note)
        db.session.commit()
        
        flash('تم إضافة الملاحظة بنجاح!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

@app.route('/edit_note/<int:note_id>', methods=['POST'])
@login_required
def edit_note(note_id):
    """Edit an existing note"""
    try:
        note = Note.query.get_or_404(note_id)
        
        note.title = request.form['title']
        note.content = request.form['content']
        note.category = request.form['category']
        note.color = request.form['color']
        note.is_pinned = 'is_pinned' in request.form
        note.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('تم تحديث الملاحظة بنجاح!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

@app.route('/delete_note/<int:note_id>', methods=['POST'])
@login_required
def delete_note(note_id):
    """Delete a note"""
    try:
        note = Note.query.get_or_404(note_id)
        db.session.delete(note)
        db.session.commit()
        
        flash('تم حذف الملاحظة بنجاح!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

@app.route('/toggle_pin_note/<int:note_id>', methods=['POST'])
@login_required
def toggle_pin_note(note_id):
    """Toggle pin status of a note"""
    try:
        note = Note.query.get_or_404(note_id)
        note.is_pinned = not note.is_pinned
        note.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash('تم تحديث تثبيت الملاحظة!', 'success')
        return redirect(url_for('tasks') + '#notes')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث الملاحظة: {str(e)}', 'error')
        return redirect(url_for('tasks') + '#notes')

# Health check endpoints for monitoring
@app.route('/health')
def health_check():
    """Basic health check endpoint"""
    try:
        # Test database connection
        db.session.execute('SELECT 1')
        db_status = 'healthy'
    except Exception:
        db_status = 'unhealthy'
    
    return jsonify({
        'status': 'healthy' if db_status == 'healthy' else 'unhealthy',
        'database': db_status,
        'timestamp': datetime.utcnow().isoformat(),
        'version': '1.0.0'
    }), 200 if db_status == 'healthy' else 503

@app.route('/ping')
def ping():
    """Simple ping endpoint"""
    return 'pong', 200

@app.route('/status')
def status():
    """Detailed status information"""
    try:
        # Check database
        db.session.execute('SELECT 1')
        db_status = 'connected'
        
        # Count records
        users_count = User.query.count()
        students_count = Student.query.count()
        
    except Exception as e:
        db_status = f'error: {str(e)}'
        users_count = -1
        students_count = -1
    
    return jsonify({
        'app_name': 'Tafra Student Management System',
        'version': '1.0.0',
        'status': 'running',
        'environment': os.environ.get('FLASK_ENV', 'development'),
        'database': {
            'status': db_status,
            'users_count': users_count,
            'students_count': students_count
        },
        'timestamp': datetime.utcnow().isoformat()
    }), 200

def get_instructor_groups(user):
    """Get groups assigned to a specific instructor user"""
    if user.role == 'admin':
        return Group.query.all()
    elif user.role == 'instructor' and user.linked_instructor:
        return user.linked_instructor.groups
    return []

def get_instructor_students(user):
    """Get students assigned to a specific instructor user"""
    if user.role == 'admin':
        return Student.query.all()
    elif user.role == 'instructor' and user.linked_instructor:
        # Get all students in instructor's groups
        instructor_groups = user.linked_instructor.groups
        students = set()
        for group in instructor_groups:
            students.update(group.students)
        return list(students)
    return []

@app.route('/instructor_attendance')
@instructor_required
def instructor_attendance():
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    
    # Get attendance for instructor's groups only
    today = datetime.now().date()
    attendance_records = []
    
    for group in instructor_groups:
        group_attendance = Attendance.query.filter_by(
            group_id=group.id,
            date=today
        ).all()
        attendance_records.extend(group_attendance)
    
    return render_template('instructor_attendance.html',
                         groups=instructor_groups,
                         attendance_records=attendance_records,
                         today=today)

@app.route('/instructor_mark_attendance', methods=['POST'])
@instructor_required
def instructor_mark_attendance():
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    group_ids = [g.id for g in instructor_groups]
    
    group_id = int(request.form['group_id'])
    
    # Verify instructor has access to this group
    if group_id not in group_ids:
        flash('ليس لديك صلاحية لأخذ حضور هذه المجموعة', 'error')
        return redirect(url_for('instructor_attendance'))
    
    date_str = request.form['date']
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    group = Group.query.get(group_id)
    students = group.students
    
    for student in students:
        student_id = str(student.id)
        status = request.form.get(f'attendance_{student_id}')
        
        if status:
            # Check if attendance already exists
            existing = Attendance.query.filter_by(
                student_id=student.id,
                date=date_obj,
                group_id=group_id
            ).first()
            
            if existing:
                existing.status = status
            else:
                attendance = Attendance(
                    student_id=student.id,
                    date=date_obj,
                    status=status,
                    group_id=group_id
                )
                db.session.add(attendance)
    
    try:
        db.session.commit()
        flash('تم حفظ الحضور بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء حفظ الحضور', 'error')
    
    return redirect(url_for('instructor_attendance'))

@app.route('/instructor_notes')
@instructor_required
def instructor_notes():
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    instructor_students = get_instructor_students(current_user)
    
    # Get instructor's notes
    notes = InstructorNote.query.filter_by(created_by=current_user.id)\
                               .order_by(InstructorNote.created_at.desc()).all()
    
    return render_template('instructor_notes.html',
                         notes=notes,
                         instructor_groups=instructor_groups,
                         instructor_students=instructor_students)

@app.route('/add_instructor_note', methods=['POST'])
@instructor_required
def add_instructor_note():
    current_user = get_current_user()
    
    title = request.form['title']
    content = request.form['content']
    priority = request.form['priority']
    student_id = request.form.get('student_id') if request.form.get('student_id') else None
    group_id = request.form.get('group_id') if request.form.get('group_id') else None
    
    # Verify instructor has access to selected student/group
    if student_id:
        instructor_students = get_instructor_students(current_user)
        student_ids = [s.id for s in instructor_students]
        if int(student_id) not in student_ids:
            flash('ليس لديك صلاحية لإضافة ملاحظة لهذا الطالب', 'error')
            return redirect(url_for('instructor_notes'))
    
    if group_id:
        instructor_groups = get_instructor_groups(current_user)
        group_ids = [g.id for g in instructor_groups]
        if int(group_id) not in group_ids:
            flash('ليس لديك صلاحية لإضافة ملاحظة لهذه المجموعة', 'error')
            return redirect(url_for('instructor_notes'))
    
    note = InstructorNote(
        title=title,
        content=content,
        priority=priority,
        student_id=int(student_id) if student_id else None,
        group_id=int(group_id) if group_id else None,
        created_by=current_user.id
    )
    
    try:
        db.session.add(note)
        db.session.commit()
        flash('تم إرسال الملاحظة للإدارة بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء إضافة الملاحظة', 'error')
    
    return redirect(url_for('instructor_notes'))

@app.route('/instructor_todos')
@instructor_required
def instructor_todos():
    """Display instructor's personal todo list"""
    current_user = get_current_user()
    instructor_groups = get_instructor_groups(current_user)
    instructor_students = get_instructor_students(current_user)
    
    # Get filter parameters
    filter_status = request.args.get('status', 'all')
    filter_priority = request.args.get('priority', 'all')
    filter_category = request.args.get('category', 'all')
    
    # Build query based on filters
    query = InstructorTodo.query.filter_by(created_by=current_user.id)
    
    if filter_status != 'all':
        query = query.filter_by(status=filter_status)
    
    if filter_priority != 'all':
        query = query.filter_by(priority=filter_priority)
    
    if filter_category != 'all':
        query = query.filter_by(category=filter_category)
    
    # Order by priority and creation date
    priority_order = {'عالي': 3, 'متوسط': 2, 'منخفض': 1}
    todos = query.all()
    todos.sort(key=lambda x: (priority_order.get(x.priority, 0), x.created_at), reverse=True)
    
    # Get statistics
    total_todos = InstructorTodo.query.filter_by(created_by=current_user.id).count()
    open_todos = InstructorTodo.query.filter_by(created_by=current_user.id, status='مفتوح').count()
    completed_todos = InstructorTodo.query.filter_by(created_by=current_user.id, status='مكتمل').count()
    overdue_todos = len([t for t in InstructorTodo.query.filter_by(created_by=current_user.id).all() if t.is_overdue])
    
    return render_template('instructor_todos.html',
                         todos=todos,
                         instructor_groups=instructor_groups,
                         instructor_students=instructor_students,
                         total_todos=total_todos,
                         open_todos=open_todos,
                         completed_todos=completed_todos,
                         overdue_todos=overdue_todos,
                         filter_status=filter_status,
                         filter_priority=filter_priority,
                         filter_category=filter_category)

@app.route('/add_instructor_todo', methods=['POST'])
@instructor_required
def add_instructor_todo():
    """Add a new todo item for instructor"""
    try:
        current_user = get_current_user()
        
        title = request.form['title']
        description = request.form.get('description', '')
        category = request.form['category']
        priority = request.form['priority']
        due_date_str = request.form.get('due_date')
        group_id = request.form.get('group_id')
        student_id = request.form.get('student_id')
        
        # Parse due date
        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        
        # Verify instructor has access to selected group/student
        if group_id:
            instructor_groups = get_instructor_groups(current_user)
            group_ids = [g.id for g in instructor_groups]
            if int(group_id) not in group_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذه المجموعة', 'error')
                return redirect(url_for('instructor_todos'))
        
        if student_id:
            instructor_students = get_instructor_students(current_user)
            student_ids = [s.id for s in instructor_students]
            if int(student_id) not in student_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذا الطالب', 'error')
                return redirect(url_for('instructor_todos'))
        
        todo = InstructorTodo(
            title=title,
            description=description,
            category=category,
            priority=priority,
            due_date=due_date,
            group_id=int(group_id) if group_id else None,
            student_id=int(student_id) if student_id else None,
            created_by=current_user.id
        )
        
        db.session.add(todo)
        db.session.commit()
        
        flash('تم إضافة المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إضافة المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/update_instructor_todo_status/<int:todo_id>', methods=['POST'])
@instructor_required
def update_instructor_todo_status(todo_id):
    """Update todo status"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        new_status = request.form['status']
        todo.status = new_status
        
        # If marking as completed, set completion time
        if new_status == 'مكتمل':
            todo.completed_at = datetime.utcnow()
        else:
            todo.completed_at = None
        
        todo.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('تم تحديث حالة المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/edit_instructor_todo/<int:todo_id>', methods=['POST'])
@instructor_required
def edit_instructor_todo(todo_id):
    """Edit an existing todo"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        todo.title = request.form['title']
        todo.description = request.form.get('description', '')
        todo.category = request.form['category']
        todo.priority = request.form['priority']
        
        # Parse due date
        due_date_str = request.form.get('due_date')
        if due_date_str:
            todo.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        else:
            todo.due_date = None
        
        # Update group and student
        group_id = request.form.get('group_id')
        student_id = request.form.get('student_id')
        
        # Verify access
        if group_id:
            instructor_groups = get_instructor_groups(current_user)
            group_ids = [g.id for g in instructor_groups]
            if int(group_id) not in group_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذه المجموعة', 'error')
                return redirect(url_for('instructor_todos'))
            todo.group_id = int(group_id)
        else:
            todo.group_id = None
        
        if student_id:
            instructor_students = get_instructor_students(current_user)
            student_ids = [s.id for s in instructor_students]
            if int(student_id) not in student_ids:
                flash('ليس لديك صلاحية لربط المهمة بهذا الطالب', 'error')
                return redirect(url_for('instructor_todos'))
            todo.student_id = int(student_id)
        else:
            todo.student_id = None
        
        todo.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('تم تحديث المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/delete_instructor_todo/<int:todo_id>', methods=['POST'])
@instructor_required
def delete_instructor_todo(todo_id):
    """Delete a todo"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        db.session.delete(todo)
        db.session.commit()
        
        flash('تم حذف المهمة بنجاح!', 'success')
        return redirect(url_for('instructor_todos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف المهمة: {str(e)}', 'error')
        return redirect(url_for('instructor_todos'))

@app.route('/get_instructor_todo/<int:todo_id>')
@instructor_required
def get_instructor_todo(todo_id):
    """Get todo details for editing"""
    try:
        current_user = get_current_user()
        todo = InstructorTodo.query.filter_by(id=todo_id, created_by=current_user.id).first_or_404()
        
        return jsonify({
            'id': todo.id,
            'title': todo.title,
            'description': todo.description,
            'category': todo.category,
            'priority': todo.priority,
            'due_date': todo.due_date.strftime('%Y-%m-%d') if todo.due_date else '',
            'group_id': todo.group_id,
            'student_id': todo.student_id,
            'status': todo.status
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/download_groups_template')
@login_required
def download_groups_template():
    """Download Excel template for groups import"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "قالب المجموعات"
        
        # Set RTL direction for Arabic support
        ws.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'الرقم', 'اسم المجموعة', 'المستوى', 'اسم المدرس', 'السعر الكلي', 
            'السعر الشهري', 'الحد الأقصى للطلاب', 'تاريخ استحقاق الدفع (يوم من الشهر)',
            'تفعيل الدفع الشهري', 'الحالة', 'ملاحظات'
        ]
        
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
            
            # Auto-adjust column width
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add sample data rows with instructions
        sample_data = [
            [1, 'مجموعة الرياضيات المتقدمة', 'متقدم', 'أحمد محمد', 500.0, 100.0, 15, 1, 'نعم', 'نشط', 'مجموعة تجريبية'],
            [2, 'مجموعة الفيزياء', 'متوسط', 'فاطمة علي', 400.0, 80.0, 12, 5, 'نعم', 'نشط', 'مجموعة أساسية'],
            ['', 'اتركها فارغة لحذف هذا السطر', '', '', '', '', '', '', '', '', '']
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if row_idx == 4:  # Instruction row
                    cell.font = Font(italic=True, color="999999")
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet(title="تعليمات الاستخدام")
        ws_instructions.sheet_view.rightToLeft = True
        
        instructions = [
            "تعليمات استخدام قالب المجموعات:",
            "",
            "1. املأ البيانات في الأعمدة المطلوبة فقط",
            "2. اسم المجموعة: مطلوب (لا يمكن أن يكون فارغاً)",
            "3. المستوى: اختياري (مثل: مبتدئ، متوسط، متقدم)",
            "4. اسم المدرس: اختياري (يجب أن يكون موجود في النظام مسبقاً)",
            "5. السعر الكلي: السعر الإجمالي للكورس (رقم)",
            "6. السعر الشهري: السعر الشهري إذا كان مختلف عن السعر الكلي (رقم)",
            "7. الحد الأقصى للطلاب: عدد الطلاب المسموح بهم (رقم، افتراضي: 15)",
            "8. تاريخ استحقاق الدفع: يوم من الشهر (1-28، افتراضي: 1)",
            "9. تفعيل الدفع الشهري: نعم/لا (افتراضي: نعم)",
            "10. الحالة: نشط/مكتمل (افتراضي: نشط)",
            "11. ملاحظات: أي ملاحظات إضافية (اختياري)",
            "",
            "ملاحظات مهمة:",
            "- احذف الصفوف التجريبية قبل التحميل",
            "- تأكد من وجود المدرسين في النظام قبل ربطهم بالمجموعات",
            "- الأعمدة المطلوبة: اسم المجموعة فقط",
            "- سيتم تجاهل الصفوف الفارغة",
            "- في حالة وجود مجموعة بنفس الاسم، سيتم تحديث بياناتها"
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:
                ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        
        # Adjust column width for instructions
        ws_instructions.column_dimensions['A'].width = 60
        
        # Create Excel file in memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with current date
        filename = f"قالب_المجموعات_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير القالب: {str(e)}', 'error')
        return redirect(url_for('groups'))

@app.route('/import_groups', methods=['GET', 'POST'])
@login_required
def import_groups():
    """Import groups from Excel file"""
    if request.method == 'GET':
        return render_template('import_groups.html')
    
    try:
        if 'excel_file' not in request.files:
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_groups'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_groups'))
        
        # Check if file is Excel
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('يرجى رفع ملف Excel صحيح (.xlsx أو .xls)', 'error')
            return redirect(url_for('import_groups'))
        
        # Read the Excel file
        from openpyxl import load_workbook
        import tempfile
        import os
        
        # Create temporary file for upload
        temp_file_path = None
        wb = None
        
        try:
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                file.save(tmp_file.name)
                temp_file_path = tmp_file.name
            
            # Load workbook
            wb = load_workbook(temp_file_path, read_only=False, data_only=True)
            
            import_summary = {
                'groups_added': 0,
                'groups_updated': 0,
                'errors': [],
                'warnings': []
            }
            
            # Find the correct sheet (try different possible names)
            sheet_names = ['قالب المجموعات', 'المجموعات', 'Groups']
            ws = None
            
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    break
            
            if not ws:
                # Use the first sheet if no specific sheet found
                ws = wb.active
            
            # Process each row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Skip completely empty rows
                    if not any(cell for cell in row if cell is not None and str(cell).strip()):
                        continue
                    
                    # Skip rows with empty group name (required field)
                    if not row[1] or not str(row[1]).strip():
                        continue
                    
                    # Extract data with proper handling
                    name = str(row[1]).strip()
                    level = str(row[2]).strip() if row[2] and str(row[2]).strip() else None
                    instructor_name = str(row[3]).strip() if row[3] and str(row[3]).strip() else None
                    
                    # Parse prices
                    def parse_price(value):
                        if value is None:
                            return 0.0
                        try:
                            if isinstance(value, (int, float)):
                                return float(value)
                            value_str = str(value).replace(',', '').replace('ج.م', '').strip()
                            return float(value_str) if value_str else 0.0
                        except:
                            return 0.0
                    
                    price = parse_price(row[4]) if len(row) > 4 else 0.0
                    monthly_price = parse_price(row[5]) if len(row) > 5 else 0.0
                    max_students = int(row[6]) if len(row) > 6 and row[6] else 15
                    payment_due_day = int(row[7]) if len(row) > 7 and row[7] else 1
                    
                    # Parse boolean fields
                    monthly_payment_enabled = True
                    if len(row) > 8 and row[8]:
                        monthly_payment_enabled = str(row[8]).strip().lower() in ['نعم', 'yes', 'true', '1']
                    
                    status = 'active'
                    if len(row) > 9 and row[9]:
                        status_value = str(row[9]).strip().lower()
                        if status_value in ['مكتمل', 'completed', 'complete']:
                            status = 'completed'
                    
                    notes = str(row[10]).strip() if len(row) > 10 and row[10] else None
                    
                    # Validate payment due day
                    payment_due_day = max(1, min(28, payment_due_day))
                    
                    # Find instructor if specified
                    instructor = None
                    if instructor_name:
                        instructor = Instructor.query.filter_by(name=instructor_name).first()
                        if not instructor:
                            import_summary['warnings'].append(
                                f'الصف {row_num}: المدرس "{instructor_name}" غير موجود - سيتم إنشاء المجموعة بدون مدرس'
                            )
                    
                    # Check if group already exists
                    existing_group = Group.query.filter_by(name=name).first()
                    
                    if existing_group:
                        # Update existing group
                        existing_group.instructor_id = instructor.id if instructor else None
                        existing_group.price = price
                        existing_group.monthly_price = monthly_price
                        existing_group.max_students = max_students
                        existing_group.payment_due_day = payment_due_day
                        existing_group.monthly_payment_enabled = monthly_payment_enabled
                        existing_group.status = status
                        
                        import_summary['groups_updated'] += 1
                    else:
                        # Create new group
                        group = Group(
                            name=name,
                            instructor_id=instructor.id if instructor else None,
                            price=price,
                            monthly_price=monthly_price,
                            max_students=max_students,
                            payment_due_day=payment_due_day,
                            monthly_payment_enabled=monthly_payment_enabled,
                            status=status
                        )
                        db.session.add(group)
                        import_summary['groups_added'] += 1
                
                except Exception as e:
                    import_summary['errors'].append(f'خطأ في الصف {row_num}: {str(e)}')
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Create success message
            success_message = f"تم استيراد المجموعات بنجاح! "
            success_message += f"تمت إضافة {import_summary['groups_added']} مجموعة جديدة و "
            success_message += f"تحديث {import_summary['groups_updated']} مجموعة موجودة."
            
            if import_summary['warnings']:
                success_message += f" توجد {len(import_summary['warnings'])} تحذيرات."
            
            flash(success_message, 'success')
            
            # Show warnings if any
            for warning in import_summary['warnings'][:5]:  # Show first 5 warnings
                flash(warning, 'warning')
            
            # Show errors if any
            for error in import_summary['errors'][:5]:  # Show first 5 errors
                flash(error, 'error')
            
            return redirect(url_for('groups'))
            
        finally:
            # Clean up
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except:
                    pass
                
    except Exception as e:
        flash(f'حدث خطأ أثناء استيراد المجموعات: {str(e)}', 'error')
        return redirect(url_for('import_groups'))

@app.route('/add_monthly_payment/<int:group_id>', methods=['POST'])
@login_required
def add_monthly_payment(group_id):
    """Add a monthly payment for a group"""
    try:
        group = Group.query.get_or_404(group_id)
        amount = float(request.form['amount'])
        year = int(request.form['year'])
        month = int(request.form['month'])
        notes = request.form.get('notes', '')
        
        # Get or create monthly payment record
        monthly_payment = group.get_monthly_payment(year, month)
        
        # Add the payment amount
        monthly_payment.total_paid += amount
        monthly_payment.update_payment_status()
        
        if notes:
            if monthly_payment.notes:
                monthly_payment.notes += f"\n{notes}"
            else:
                monthly_payment.notes = notes
        
        db.session.commit()
        
        flash(f'تم إضافة دفعة شهرية بقيمة {amount} ج.م للمجموعة {group.name}', 'success')
        return redirect(url_for('group_details', group_id=group_id))
        
    except Exception as e:
        flash(f'حدث خطأ أثناء إضافة الدفعة: {str(e)}', 'error')
        return redirect(url_for('group_details', group_id=group_id))

@app.route('/monthly_payments/<int:group_id>')
@login_required
def monthly_payments(group_id):
    """View monthly payments for a group"""
    group = Group.query.get_or_404(group_id)
    
    # Get current year or requested year
    year = request.args.get('year', datetime.now().year, type=int)
    
    # Get all monthly payments for this group and year
    monthly_payments = MonthlyPayment.query.filter_by(
        group_id=group_id, year=year
    ).order_by(MonthlyPayment.month).all()
    
    # Create monthly payment records for the year if they don't exist
    arabic_months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    
    existing_months = {mp.month for mp in monthly_payments}
    for month in range(1, 13):
        if month not in existing_months:
            monthly_payment = group.get_monthly_payment(year, month)
            monthly_payments.append(monthly_payment)
    
    # Sort by month
    monthly_payments.sort(key=lambda x: x.month)
    
    # Calculate statistics
    total_expected = sum(mp.monthly_price for mp in monthly_payments)
    total_paid = sum(mp.total_paid for mp in monthly_payments)
    total_remaining = total_expected - total_paid
    
    completion_rate = (total_paid / total_expected * 100) if total_expected > 0 else 0
    
    return render_template('monthly_payments.html',
                         group=group,
                         monthly_payments=monthly_payments,
                         year=year,
                         total_expected=total_expected,
                         total_paid=total_paid,
                         total_remaining=total_remaining,
                         completion_rate=completion_rate,
                         arabic_months=arabic_months)

@app.route('/grades')
@permission_required('manage_students')  # Or view_reports - grades are primarily for student management
def grades():
    """Main grades management page"""
    # Get filter parameters
    group_filter = request.args.get('group_id', type=int)
    subject_filter = request.args.get('subject_id', type=int)
    student_filter = request.args.get('student_id', type=int)
    
    # Get all subjects, groups, and students for filters
    subjects = Subject.query.filter_by(is_active=True).all()
    groups = Group.query.all()
    students = Student.query.all()
    
    # Build grades query with filters
    grades_query = Grade.query.join(Student).join(Subject)
    
    if group_filter:
        grades_query = grades_query.filter(Subject.group_id == group_filter)
    if subject_filter:
        grades_query = grades_query.filter(Grade.subject_id == subject_filter)
    if student_filter:
        grades_query = grades_query.filter(Grade.student_id == student_filter)
    
    grades = grades_query.order_by(Student.name, Subject.name).all()
    
    # Calculate statistics
    total_grades = len(grades)
    if grades:
        average_score = sum(g.score for g in grades if g.score) / len([g for g in grades if g.score])
        passing_grades = len([g for g in grades if g.percentage and g.percentage >= 60])
        passing_rate = (passing_grades / total_grades * 100) if total_grades > 0 else 0
    else:
        average_score = 0
        passing_rate = 0
        passing_grades = 0
    
    return render_template('grades.html',
                         subjects=subjects,
                         groups=groups,
                         students=students,
                         grades=grades,
                         total_grades=total_grades,
                         average_score=average_score,
                         passing_rate=passing_rate,
                         passing_grades=passing_grades,
                         selected_group=group_filter,
                         selected_subject=subject_filter,
                         selected_student=student_filter)

@app.route('/achievements')
@login_required
def achievements():
    """Student achievements and leaderboard page"""
    # Get filter parameters
    grade_level_filter = request.args.get('grade_level', '')
    group_filter = request.args.get('group_id', type=int)
    
    # Build base query
    students_query = Student.query
    
    # Apply filters
    if grade_level_filter:
        students_query = students_query.filter(Student.grade_level == grade_level_filter)
    
    if group_filter:
        group = Group.query.get(group_filter)
        if group:
            students_query = students_query.filter(Student.groups.contains(group))
    
    # Get students and sort by achievement points
    students = students_query.order_by(Student.total_achievement_points.desc()).all()
    
    # Get filter options
    grade_levels = db.session.query(Student.grade_level.distinct()).filter(Student.grade_level.isnot(None)).all()
    grade_levels = [g[0] for g in grade_levels if g[0]]
    
    groups = Group.query.order_by(Group.name).all()
    
    # Calculate achievement statistics
    total_students = len(students)
    
    if students:
        # Achievement level distribution
        level_counts = {}
        level_order = ['نجم', 'متفوق', 'متقدم', 'مبتدئ']
        for level in level_order:
            level_counts[level] = len([s for s in students if s.achievement_level == level])
        
        # Average points
        avg_total_points = sum(s.total_achievement_points for s in students) / total_students
        avg_attendance_points = sum(s.attendance_points for s in students) / total_students
        avg_grade_points = sum(s.grade_points for s in students) / total_students
        
        # Top performers
        top_students = students[:10]  # Top 10 students
        
        stats = {
            'total_students': total_students,
            'level_distribution': level_counts,
            'avg_total_points': round(avg_total_points, 1),
            'avg_attendance_points': round(avg_attendance_points, 1),
            'avg_grade_points': round(avg_grade_points, 1),
            'top_students': top_students
        }
    else:
        stats = {
            'total_students': 0,
            'level_distribution': {},
            'avg_total_points': 0,
            'avg_attendance_points': 0,
            'avg_grade_points': 0,
            'top_students': []
        }
    
    return render_template('achievements.html',
                           students=students,
                           grade_levels=grade_levels,
                           groups=groups,
                           stats=stats,
                           selected_grade_level=grade_level_filter,
                           selected_group=group_filter)

@app.route('/update_achievement_points/<int:student_id>', methods=['POST'])
@login_required
def update_achievement_points_route(student_id):
    """Manually update achievement points for a student"""
    student = Student.query.get_or_404(student_id)
    
    try:
        result = student.update_achievement_points()
        flash(f'تم تحديث نقاط الإنجاز للطالب {student.name} بنجاح', 'success')
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        flash('حدث خطأ أثناء تحديث نقاط الإنجاز', 'error')
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update_all_achievement_points', methods=['POST'])
@admin_required
def update_all_achievement_points():
    """Update achievement points for all students"""
    try:
        students = Student.query.all()
        updated_count = 0
        
        for student in students:
            student.update_achievement_points()
            updated_count += 1
        
        flash(f'تم تحديث نقاط الإنجاز لـ {updated_count} طالب بنجاح', 'success')
        return jsonify({'success': True, 'updated_count': updated_count})
    except Exception as e:
        flash('حدث خطأ أثناء تحديث نقاط الإنجاز', 'error')
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_bonus_points/<int:student_id>', methods=['POST'])
@login_required
def add_bonus_points(student_id):
    """Add bonus points to a student"""
    student = Student.query.get_or_404(student_id)
    points = float(request.form.get('bonus_points', 0))
    reason = request.form.get('reason', '')
    
    try:
        student.bonus_points += points
        student.total_achievement_points = student.attendance_points + student.grade_points + student.bonus_points
        
        # Update achievement level
        student.update_achievement_points()
        
        flash(f'تم إضافة {points} نقطة إضافية للطالب {student.name}', 'success')
        return redirect(request.referrer or url_for('achievements'))
    except Exception as e:
        flash('حدث خطأ أثناء إضافة النقاط الإضافية', 'error')
        return redirect(request.referrer or url_for('achievements'))

@app.route('/download_grades_template')
@login_required
def download_grades_template():
    """Download Excel template for grades import"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "قالب الدرجات"
        
        # Set RTL direction for Arabic support
        ws.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'الرقم', 'اسم الطالب', 'اسم المادة/الاختبار', 'نوع المادة', 'الدرجة المحصلة',
            'الدرجة القصوى', 'تاريخ الاختبار', 'المجموعة', 'ملاحظات'
        ]
        
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
            
            # Auto-adjust column width
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add sample data
        sample_data = [
            [1, 'أحمد محمد', 'الرياضيات - اختبار شهري', 'اختبار', 85, 100, '2024-01-15', 'مجموعة الرياضيات المتقدمة', 'أداء ممتاز'],
            [2, 'فاطمة علي', 'الرياضيات - اختبار شهري', 'اختبار', 92, 100, '2024-01-15', 'مجموعة الرياضيات المتقدمة', 'متفوقة'],
            [3, 'محمد حسن', 'الفيزياء - واجب منزلي', 'واجب', 18, 20, '2024-01-10', 'مجموعة الفيزياء', 'جيد جداً'],
            ['', 'اتركها فارغة لحذف هذا السطر', '', '', '', '', '', '', '']
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if row_idx == 5:  # Instruction row
                    cell.font = Font(italic=True, color="999999")
        
        # Add subjects sheet
        ws_subjects = wb.create_sheet(title="المواد المتاحة")
        ws_subjects.sheet_view.rightToLeft = True
        
        # Get existing subjects for reference
        subjects = Subject.query.filter_by(is_active=True).all()
        
        ws_subjects['A1'] = "المواد المتاحة في النظام:"
        ws_subjects['A1'].font = Font(bold=True, size=14)
        
        for idx, subject in enumerate(subjects, 3):
            ws_subjects[f'A{idx}'] = f"• {subject.name}"
            if subject.group:
                ws_subjects[f'B{idx}'] = f"المجموعة: {subject.group.name}"
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet(title="تعليمات الاستخدام")
        ws_instructions.sheet_view.rightToLeft = True
        
        instructions = [
            "تعليمات استخدام قالب الدرجات:",
            "",
            "1. املأ البيانات في الأعمدة المطلوبة",
            "2. اسم الطالب: يجب أن يكون موجود في النظام مسبقاً",
            "3. اسم المادة/الاختبار: اسم المادة أو الاختبار",
            "4. نوع المادة: مادة، اختبار، واجب، مشروع",
            "5. الدرجة المحصلة: الدرجة التي حصل عليها الطالب (رقم)",
            "6. الدرجة القصوى: أقصى درجة ممكنة (رقم، افتراضي: 100)",
            "7. تاريخ الاختبار: بصيغة YYYY-MM-DD (اختياري)",
            "8. المجموعة: اسم المجموعة (اختياري)",
            "9. ملاحظات: أي ملاحظات إضافية (اختياري)",
            "",
            "ملاحظات مهمة:",
            "- احذف الصفوف التجريبية قبل التحميل",
            "- تأكد من وجود الطلاب في النظام قبل إدخال درجاتهم",
            "- سيتم حساب النسبة المئوية والدرجة بالحروف تلقائياً",
            "- إذا لم تكن المادة موجودة، سيتم إنشاؤها تلقائياً",
            "- الأعمدة المطلوبة: اسم الطالب، اسم المادة، الدرجة المحصلة",
            "- سيتم تجاهل الصفوف الفارغة"
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:
                ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        
        # Adjust column width for instructions
        ws_instructions.column_dimensions['A'].width = 60
        
        # Create Excel file in memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with current date
        filename = f"قالب_الدرجات_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير القالب: {str(e)}', 'error')
        return redirect(url_for('grades'))

@app.route('/import_grades', methods=['GET', 'POST'])
@login_required
def import_grades():
    """Import grades from Excel file"""
    if request.method == 'GET':
        return render_template('import_grades.html')
    
    try:
        if 'excel_file' not in request.files:
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_grades'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_grades'))
        
        # Check if file is Excel
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('يرجى رفع ملف Excel صحيح (.xlsx أو .xls)', 'error')
            return redirect(url_for('import_grades'))
        
        # Read the Excel file
        from openpyxl import load_workbook
        import tempfile
        import os
        
        # Create temporary file for upload
        temp_file_path = None
        wb = None
        
        try:
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                file.save(tmp_file.name)
                temp_file_path = tmp_file.name
            
            # Load workbook
            wb = load_workbook(temp_file_path, read_only=False, data_only=True)
            
            import_summary = {
                'grades_added': 0,
                'grades_updated': 0,
                'subjects_created': 0,
                'errors': [],
                'warnings': []
            }
            
            # Find the correct sheet
            sheet_names = ['قالب الدرجات', 'الدرجات', 'Grades']
            ws = None
            
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    break
            
            if not ws:
                ws = wb.active
            
            # Process each row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Skip completely empty rows
                    if not any(cell for cell in row if cell is not None and str(cell).strip()):
                        continue
                    
                    # Skip rows with missing required fields
                    if not row[1] or not row[2] or row[4] is None:  # student name, subject name, score
                        continue
                    
                    # Extract data
                    student_name = str(row[1]).strip()
                    subject_name = str(row[2]).strip()
                    subject_type = str(row[3]).strip() if row[3] else 'مادة'
                    score = float(row[4])
                    max_score = float(row[5]) if row[5] else 100.0
                    
                    # Parse exam date
                    exam_date = None
                    if row[6]:
                        try:
                            if isinstance(row[6], datetime):
                                exam_date = row[6].date()
                            else:
                                exam_date = datetime.strptime(str(row[6]), '%Y-%m-%d').date()
                        except:
                            pass
                    
                    group_name = str(row[7]).strip() if row[7] else None
                    notes = str(row[8]).strip() if row[8] else None
                    
                    # Find student
                    student = Student.query.filter_by(name=student_name).first()
                    if not student:
                        import_summary['errors'].append(f'الصف {row_num}: الطالب "{student_name}" غير موجود')
                        continue
                    
                    # Find or create subject
                    subject = Subject.query.filter_by(name=subject_name).first()
                    if not subject:
                        # Find group if specified
                        group = None
                        if group_name:
                            group = Group.query.filter_by(name=group_name).first()
                        
                        # Create new subject
                        subject = Subject(
                            name=subject_name,
                            subject_type=subject_type,
                            max_grade=max_score
                        )
                        db.session.add(subject)
                        db.session.flush()  # Get the ID
                        import_summary['subjects_created'] += 1
                    
                    # Check if grade already exists
                    existing_grade = Grade.query.filter_by(
                        student_id=student.id,
                        subject_id=subject.id
                    ).first()
                    
                    if existing_grade:
                        # Update existing grade
                        existing_grade.score = score
                        existing_grade.max_score = max_score
                        existing_grade.exam_date = exam_date
                        existing_grade.notes = notes
                        existing_grade.save_with_calculations()
                        import_summary['grades_updated'] += 1
                    else:
                        # Create new grade
                        grade = Grade(
                            student_id=student.id,
                            subject_id=subject.id,
                            score=score,
                            max_score=max_score,
                            exam_date=exam_date,
                            notes=notes
                        )
                        grade.save_with_calculations()
                        import_summary['grades_added'] += 1
                
                except Exception as e:
                    import_summary['errors'].append(f'خطأ في الصف {row_num}: {str(e)}')
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Create success message
            success_message = f"تم استيراد الدرجات بنجاح! "
            success_message += f"تمت إضافة {import_summary['grades_added']} درجة جديدة و "
            success_message += f"تحديث {import_summary['grades_updated']} درجة موجودة"
            
            if import_summary['subjects_created'] > 0:
                success_message += f" وإنشاء {import_summary['subjects_created']} مادة جديدة"
            
            flash(success_message, 'success')
            
            # Show warnings if any
            for warning in import_summary['warnings'][:5]:
                flash(warning, 'warning')
            
            # Show errors if any
            for error in import_summary['errors'][:5]:
                flash(error, 'error')
            
            return redirect(url_for('grades'))
            
        finally:
            # Clean up
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except:
                    pass
                
    except Exception as e:
        flash(f'حدث خطأ أثناء استيراد الدرجات: {str(e)}', 'error')
        return redirect(url_for('import_grades'))

@app.route('/manage_subjects')
@subjects_required
def manage_subjects():
    """Manage subjects page"""
    subjects = Subject.query.all()
    instructors = Instructor.query.all()
    return render_template('manage_subjects.html', subjects=subjects, instructors=instructors)

@app.route('/add_subject', methods=['POST'])
@login_required
def add_subject():
    """Add new subject"""
    try:
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip()
        description = request.form.get('description', '').strip()
        max_grade = float(request.form.get('max_grade', 100.0))
        min_grade = float(request.form.get('min_grade', 0.0))
        subject_type = request.form.get('subject_type', 'مادة')
        instructor_id = request.form.get('instructor_id')
        
        if not name:
            flash('اسم المادة مطلوب', 'error')
            return redirect(url_for('manage_subjects'))
        
        # Check if subject already exists
        if Subject.query.filter_by(name=name).first():
            flash('المادة موجودة بالفعل', 'warning')
            return redirect(url_for('manage_subjects'))
        
        subject = Subject(
            name=name,
            code=code if code else None,
            description=description if description else None,
            max_grade=max_grade,
            min_grade=min_grade,
            subject_type=subject_type,
            instructor_id=int(instructor_id) if instructor_id else None
        )
        
        db.session.add(subject)
        db.session.commit()
        flash('تم إضافة المادة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في إضافة المادة: {str(e)}', 'error')
    
    return redirect(url_for('manage_subjects'))

@app.route('/edit_subject/<int:subject_id>', methods=['POST'])
@login_required
def edit_subject(subject_id):
    """Edit existing subject"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        
        subject.name = request.form.get('name', '').strip()
        subject.code = request.form.get('code', '').strip() or None
        subject.description = request.form.get('description', '').strip() or None
        subject.max_grade = float(request.form.get('max_grade', 100.0))
        subject.min_grade = float(request.form.get('min_grade', 0.0))
        subject.subject_type = request.form.get('subject_type', 'مادة')
        instructor_id = request.form.get('instructor_id')
        subject.instructor_id = int(instructor_id) if instructor_id else None
        
        if not subject.name:
            flash('اسم المادة مطلوب', 'error')
            return redirect(url_for('manage_subjects'))
        
        db.session.commit()
        flash('تم تحديث المادة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في تحديث المادة: {str(e)}', 'error')
    
    return redirect(url_for('manage_subjects'))

@app.route('/delete_subject/<int:subject_id>', methods=['POST'])
@login_required
def delete_subject(subject_id):
    """Delete subject"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        
        # Check if subject has grades
        if subject.grades:
            flash('لا يمكن حذف المادة لأن لها درجات مسجلة', 'error')
            return redirect(url_for('manage_subjects'))
        
        db.session.delete(subject)
        db.session.commit()
        flash('تم حذف المادة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف المادة: {str(e)}', 'error')
    
    return redirect(url_for('manage_subjects'))

@app.route('/assign_subjects_to_group/<int:group_id>', methods=['GET', 'POST'])
@login_required
def assign_subjects_to_group(group_id):
    """Assign subjects to group"""
    group = Group.query.get_or_404(group_id)
    
    if request.method == 'POST':
        try:
            subject_ids = request.form.getlist('subject_ids')
            
            # Clear existing subjects
            group.subjects.clear()
            
            # Add new subjects
            for subject_id in subject_ids:
                subject = Subject.query.get(int(subject_id))
                if subject:
                    group.subjects.append(subject)
            
            db.session.commit()
            flash(f'تم تحديث مواد المجموعة "{group.name}" بنجاح', 'success')
            return redirect(url_for('groups'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث المواد: {str(e)}', 'error')
    
    subjects = Subject.query.filter_by(is_active=True).all()
    return render_template('assign_subjects.html', group=group, subjects=subjects)

@app.route('/get_group_subjects/<int:group_id>')
@login_required
def get_group_subjects(group_id):
    """Get subjects for a group (API endpoint)"""
    group = Group.query.get_or_404(group_id)
    subjects = [{'id': s.id, 'name': s.name, 'type': s.subject_type} for s in group.subjects]
    return jsonify({'subjects': subjects})

@app.route('/student_profile/<int:student_id>')
@login_required
def student_profile(student_id):
    """Student comprehensive profile page"""
    student = Student.query.get_or_404(student_id)
    
    # Get student's grades
    grades = Grade.query.filter_by(student_id=student_id).join(Subject).order_by(Grade.created_at.desc()).all()
    
    # Get student's attendance records
    attendance_records = Attendance.query.filter_by(student_id=student_id).order_by(Attendance.date.desc()).limit(30).all()
    
    # Get student's payment history
    payments = Payment.query.filter_by(student_id=student_id).order_by(Payment.date.desc()).all()
    
    # Calculate attendance statistics
    total_sessions = len(attendance_records)
    present_sessions = len([a for a in attendance_records if a.status == 'حاضر'])
    attendance_percentage = (present_sessions / total_sessions * 100) if total_sessions > 0 else 0
    
    # Calculate grade statistics
    total_grades = len(grades)
    if total_grades > 0:
        average_score = sum(g.score for g in grades if g.score) / len([g for g in grades if g.score])
        latest_grades = grades[:5]  # Show latest 5 grades
    else:
        average_score = 0
        latest_grades = []
    
    # Calculate financial summary
    total_paid = sum(p.amount for p in payments)
    total_fees = sum(group.price for group in student.groups)
    remaining_balance = total_fees - total_paid
    
    # Get recent activities (last 30 days)
    from datetime import datetime, timedelta
    thirty_days_ago = datetime.now() - timedelta(days=30)
    
    recent_grades = Grade.query.filter(
        Grade.student_id == student_id,
        Grade.created_at >= thirty_days_ago
    ).order_by(Grade.created_at.desc()).all()
    
    recent_attendance = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date >= thirty_days_ago.date()
    ).order_by(Attendance.date.desc()).all()
    
    recent_payments = Payment.query.filter(
        Payment.student_id == student_id,
        Payment.date >= thirty_days_ago
    ).order_by(Payment.date.desc()).all()
    
    return render_template('student_profile.html',
                         student=student,
                         grades=grades,
                         latest_grades=latest_grades,
                         attendance_records=attendance_records,
                         payments=payments,
                         total_grades=total_grades,
                         average_score=average_score,
                         attendance_percentage=attendance_percentage,
                         present_sessions=present_sessions,
                         total_sessions=total_sessions,
                         total_paid=total_paid,
                         total_fees=total_fees,
                         remaining_balance=remaining_balance,
                         recent_grades=recent_grades,
                         recent_attendance=recent_attendance,
                         recent_payments=recent_payments)

@app.route('/generate_monthly_report/<int:student_id>')
@login_required
def generate_monthly_report(student_id):
    """Generate monthly performance report for student"""
    from datetime import datetime, timedelta
    from calendar import monthrange
    
    student = Student.query.get_or_404(student_id)
    
    # Get month and year from request or use current
    month = int(request.args.get('month', datetime.now().month))
    year = int(request.args.get('year', datetime.now().year))
    
    # Calculate date range for the month
    start_date = datetime(year, month, 1).date()
    last_day = monthrange(year, month)[1]
    end_date = datetime(year, month, last_day).date()
    
    # Get monthly data
    monthly_grades = Grade.query.filter(
        Grade.student_id == student_id,
        Grade.exam_date >= start_date,
        Grade.exam_date <= end_date
    ).all()
    
    monthly_attendance = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()
    
    monthly_payments = Payment.query.filter(
        Payment.student_id == student_id,
        Payment.date >= datetime.combine(start_date, datetime.min.time()),
        Payment.date <= datetime.combine(end_date, datetime.max.time())
    ).all()
    
    # Calculate statistics
    total_classes = len(monthly_attendance)
    present_classes = len([a for a in monthly_attendance if a.status == 'حاضر'])
    attendance_rate = (present_classes / total_classes * 100) if total_classes > 0 else 0
    
    if monthly_grades:
        month_average = sum(g.score for g in monthly_grades if g.score) / len([g for g in monthly_grades if g.score])
    else:
        month_average = 0
    
    monthly_paid = sum(p.amount for p in monthly_payments)
    
    # Arabic month names
    arabic_months = [
        '', 'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
        'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
    ]
    
    report_data = {
        'student': student,
        'month': arabic_months[month],
        'year': year,
        'month_num': month,
        'monthly_grades': monthly_grades,
        'monthly_attendance': monthly_attendance,
        'monthly_payments': monthly_payments,
        'total_classes': total_classes,
        'present_classes': present_classes,
        'attendance_rate': attendance_rate,
        'month_average': month_average,
        'monthly_paid': monthly_paid,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render_template('monthly_report.html', **report_data)

@app.route('/send_whatsapp_report/<int:student_id>')
@login_required
def send_whatsapp_report(student_id):
    """Generate WhatsApp message for monthly report"""
    from datetime import datetime
    from urllib.parse import quote
    
    student = Student.query.get_or_404(student_id)
    
    # Get month and year
    month = int(request.args.get('month', datetime.now().month))
    year = int(request.args.get('year', datetime.now().year))
    
    # Get report data (reuse the logic from generate_monthly_report)
    from calendar import monthrange
    start_date = datetime(year, month, 1).date()
    last_day = monthrange(year, month)[1]
    end_date = datetime(year, month, last_day).date()
    
    monthly_grades = Grade.query.filter(
        Grade.student_id == student_id,
        Grade.exam_date >= start_date,
        Grade.exam_date <= end_date
    ).all()
    
    monthly_attendance = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()
    
    total_classes = len(monthly_attendance)
    present_classes = len([a for a in monthly_attendance if a.status == 'حاضر'])
    attendance_rate = (present_classes / total_classes * 100) if total_classes > 0 else 0
    
    if monthly_grades:
        month_average = sum(g.score for g in monthly_grades if g.score) / len([g for g in monthly_grades if g.score])
    else:
        month_average = 0
    
    arabic_months = [
        '', 'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
        'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
    ]
    
    # Create WhatsApp message
    message = f"""🎓 *تقرير شهري - {student.name}*
📅 *شهر {arabic_months[month]} {year}*

📊 *الحضور والغياب:*
• إجمالي الحصص: {total_classes}
• الحصص المحضورة: {present_classes}
• نسبة الحضور: {attendance_rate:.1f}%

📝 *الدرجات:*
• عدد الاختبارات: {len(monthly_grades)}
• متوسط الدرجات: {month_average:.1f}

"""
    
    if monthly_grades:
        message += "📋 *تفاصيل الدرجات:*\n"
        for grade in monthly_grades[:5]:  # Show top 5 grades
            message += f"• {grade.subject.name}: {grade.score}/{grade.max_score} ({grade.percentage:.1f}%)\n"
    
    message += f"""
👥 *المجموعات المسجل بها:*
"""
    for group in student.groups:
        message += f"• {group.name}\n"
    
    message += f"""
🏫 *مؤسسة طفرة التعليمية*
📞 للاستفسار: {student.phone or 'غير محدد'}
"""
    
    # Create WhatsApp URL
    phone = student.phone.replace('+', '').replace(' ', '') if student.phone else ''
    whatsapp_url = f"https://wa.me/{phone}?text={quote(message)}"
    
    return render_template('whatsapp_report.html', 
                         student=student, 
                         message=message, 
                         whatsapp_url=whatsapp_url,
                         phone=phone)

@app.route('/download_students_template')
@login_required
def download_students_template():
    """Download Excel template for bulk student import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "قالب بيانات الطلاب"
        
        # Define headers
        headers = [
            'اسم الطالب',
            'رقم الهاتف',
            'العمر',
            'المنطقة',
            'المرحلة الدراسية',
            'المجموعات (أرقام مفصولة بفاصلة)',
            'الخصم',
            'ملاحظات'
        ]
        
        # Create styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add sample data
        sample_data = [
            ['أحمد محمد علي', '01234567890', 16, 'القاهرة', 'الثانوية الأولى', '1,2', 0, 'طالب متفوق'],
            ['فاطمة سامي', '01987654321', 15, 'الجيزة', 'الإعدادية الثالثة', '3', 50, 'تحتاج متابعة'],
            ['محمد أحمد', '01122334455', 12, 'الإسكندرية', 'الابتدائية السادسة', '1', 0, ''],
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        column_widths = [20, 15, 8, 15, 20, 30, 10, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("تعليمات الاستخدام")
        
        instructions = [
            "تعليمات استخدام قالب إضافة الطلاب:",
            "",
            "1. اسم الطالب: أدخل الاسم كاملاً",
            "2. رقم الهاتف: أدخل رقم الهاتف (اختياري)",
            "3. العمر: أدخل العمر بالسنوات",
            "4. المنطقة: أدخل منطقة السكن (اختياري)",
            "5. المرحلة الدراسية: اختر من القائمة التالية:",
            "   - KG1, KG2",
            "   - الابتدائية الأولى, الابتدائية الثانية, ... الابتدائية السادسة",
            "   - الإعدادية الأولى, الإعدادية الثانية, الإعدادية الثالثة",
            "   - الثانوية الأولى, الثانوية الثانية, الثانوية الثالثة",
            "   - الجامعة, أخرى",
            "6. المجموعات: أدخل أرقام المجموعات مفصولة بفاصلة (مثال: 1,2,3)",
            "7. الخصم: أدخل قيمة الخصم بالجنيه (افتراضي: 0)",
            "8. ملاحظات: أدخل أي ملاحظات إضافية (اختياري)",
            "",
            "ملاحظات مهمة:",
            "- احذف الصفوف النموذجية قبل رفع الملف",
            "- تأكد من صحة أرقام المجموعات",
            "- اسم الطالب مطلوب، باقي الحقول اختيارية",
            "- يمكن ترك الخانات فارغة للحقول الاختيارية",
            "",
            f"تاريخ إنشاء القالب: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row, column=1, value=instruction)
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif instruction.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.')):
                cell.font = Font(bold=True)
        
        ws_instructions.column_dimensions['A'].width = 60
        
        # Get available groups for reference
        groups = Group.query.all()
        if groups:
            ws_groups = wb.create_sheet("المجموعات المتاحة")
            ws_groups.cell(row=1, column=1, value="رقم المجموعة").font = Font(bold=True)
            ws_groups.cell(row=1, column=2, value="اسم المجموعة").font = Font(bold=True)
            ws_groups.cell(row=1, column=3, value="المدرس").font = Font(bold=True)
            ws_groups.cell(row=1, column=4, value="السعر").font = Font(bold=True)
            
            for row, group in enumerate(groups, 2):
                ws_groups.cell(row=row, column=1, value=group.id)
                ws_groups.cell(row=row, column=2, value=group.name)
                ws_groups.cell(row=row, column=3, value=group.instructor_ref.name if group.instructor_ref else 'غير محدد')
                ws_groups.cell(row=row, column=4, value=f"{group.price} ج.م")
            
            for col in range(1, 5):
                ws_groups.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename="students_template.xlsx"'
        
        return response
        
    except Exception as e:
        flash(f'خطأ في إنشاء القالب: {str(e)}', 'error')
        return redirect(url_for('students'))

@app.route('/import_students', methods=['GET', 'POST'])
@login_required
def import_students():
    """Import students from Excel file"""
    if request.method == 'GET':
        return render_template('import_students.html', groups=Group.query.all())
    
    if 'file' not in request.files:
        flash('لم يتم اختيار ملف', 'error')
        return redirect(url_for('students'))
    
    file = request.files['file']
    if file.filename == '':
        flash('لم يتم اختيار ملف', 'error')
        return redirect(url_for('students'))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls)', 'error')
        return redirect(url_for('students'))
    
    try:
        from openpyxl import load_workbook
        import tempfile
        import os
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            
            # Load workbook
            wb = load_workbook(tmp_file.name)
            ws = wb.active
            
            # Process data
            students_data = []
            errors = []
            success_count = 0
            
            # Skip header row
            for row_num in range(2, ws.max_row + 1):
                try:
                    row_data = []
                    for col in range(1, 9):  # 8 columns
                        cell_value = ws.cell(row=row_num, column=col).value
                        row_data.append(cell_value if cell_value is not None else '')
                    
                    name, phone, age, location, grade_level, group_ids_str, discount, notes = row_data
                    
                    # Skip empty rows
                    if not name or str(name).strip() == '':
                        continue
                    
                    # Validate required fields
                    if not name:
                        errors.append(f'الصف {row_num}: اسم الطالب مطلوب')
                        continue
                    
                    # Parse age
                    try:
                        age = int(age) if age else 18
                    except (ValueError, TypeError):
                        age = 18
                    
                    # Parse discount
                    try:
                        discount = float(discount) if discount else 0
                    except (ValueError, TypeError):
                        discount = 0
                    
                    # Parse group IDs
                    group_ids = []
                    if group_ids_str:
                        try:
                            group_ids = [int(gid.strip()) for gid in str(group_ids_str).split(',') if gid.strip()]
                        except ValueError:
                            errors.append(f'الصف {row_num}: أرقام المجموعات غير صحيحة')
                            continue
                    
                    # Validate groups exist
                    valid_groups = []
                    for group_id in group_ids:
                        group = Group.query.get(group_id)
                        if group:
                            valid_groups.append(group)
                        else:
                            errors.append(f'الصف {row_num}: المجموعة رقم {group_id} غير موجودة')
                    
                    # Create student
                    student = Student(
                        name=str(name).strip(),
                        phone=str(phone).strip() if phone else None,
                        age=age,
                        location=str(location).strip() if location else None,
                        grade_level=str(grade_level).strip() if grade_level else None,
                        discount=discount,
                        registration_date=datetime.now().date()
                    )
                    
                    # Add to database
                    db.session.add(student)
                    db.session.flush()  # Get student ID
                    
                    # Add to groups
                    for group in valid_groups:
                        student.groups.append(group)
                    
                    success_count += 1
                    students_data.append({
                        'name': student.name,
                        'phone': student.phone,
                        'groups': [g.name for g in valid_groups]
                    })
                    
                except Exception as e:
                    errors.append(f'الصف {row_num}: خطأ في معالجة البيانات - {str(e)}')
                    continue
            
            # Commit if successful
            if success_count > 0:
                db.session.commit()
                flash(f'تم إضافة {success_count} طالب بنجاح!', 'success')
            else:
                db.session.rollback()
                flash('لم يتم إضافة أي طالب', 'warning')
            
            # Show errors if any
            if errors:
                for error in errors[:10]:  # Show only first 10 errors
                    flash(error, 'error')
                if len(errors) > 10:
                    flash(f'وجد {len(errors) - 10} أخطاء إضافية...', 'warning')
            
            # Clean up temp file
            os.unlink(tmp_file.name)
            
            return render_template('import_students_result.html', 
                                 success_count=success_count,
                                 errors=errors,
                                 students_data=students_data)
            
    except Exception as e:
        flash(f'خطأ في معالجة الملف: {str(e)}', 'error')
        return redirect(url_for('students'))

@app.route('/export_full_backup')
@admin_required
def export_full_backup():
    """Export complete system backup with all data"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Set RTL direction for Arabic support
        ws_overview = wb.active
        ws_overview.title = "نظرة عامة"
        ws_overview.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        sub_header_font = Font(size=12, bold=True, color="2F5F8F")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # Title and backup info
        ws_overview.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws_overview[f'A{current_row}']
        title_cell.value = f"نسخة احتياطية شاملة - نظام تفرا لإدارة الطلاب - {format_arabic_date(datetime.now())}"
        title_cell.font = Font(size=16, bold=True, color="2F5F8F")
        title_cell.alignment = center_alignment
        current_row += 2
        
        # System overview
        ws_overview[f'A{current_row}'] = "معلومات النظام"
        ws_overview[f'A{current_row}'].font = sub_header_font
        current_row += 1
        
        overview_data = [
            ['البيان', 'القيمة'],
            ['تاريخ النسخة الاحتياطية', format_arabic_date(datetime.now())],
            ['وقت النسخة الاحتياطية', datetime.now().strftime('%H:%M:%S')],
            ['إجمالي الطلاب', Student.query.count()],
            ['إجمالي المدرسين', Instructor.query.count()],
            ['إجمالي المجموعات', Group.query.count()],
            ['إجمالي المستخدمين', User.query.count()],
            ['إجمالي المدفوعات', Payment.query.count()],
            ['إجمالي المصروفات', Expense.query.count()],
            ['إجمالي سجلات الحضور', Attendance.query.count()],
            ['إجمالي المهام', Task.query.count()],
            ['إجمالي الملاحظات', Note.query.count()],
            ['إجمالي ملاحظات المدرسين', InstructorNote.query.count()],
            ['إجمالي مهام المدرسين', InstructorTodo.query.count()],
        ]
        
        for row_data in overview_data:
            for col, value in enumerate(row_data, 1):
                cell = ws_overview.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                if current_row == len(overview_data) + current_row - len(overview_data):  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
            current_row += 1
        
        current_row += 2
        
        # Users data sheet
        ws_users = wb.create_sheet(title="المستخدمين")
        ws_users.sheet_view.rightToLeft = True
        
        user_headers = ['#', 'اسم المستخدم', 'الاسم الكامل', 'الدور', 'مخفي', 'تاريخ الإنشاء', 'آخر دخول', 'آخر نشاط', 'نشط الآن']
        for col, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        users = User.query.all()
        for idx, user in enumerate(users, 1):
            user_data = [
                idx,
                user.username,
                user.full_name,
                user.role,
                'نعم' if user.is_hidden else 'لا',
                user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else 'غير محدد',
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'لم يسجل دخول',
                user.last_activity.strftime('%Y-%m-%d %H:%M') if user.last_activity else 'غير محدد',
                'نعم' if user.is_active_now() else 'لا'
            ]
            
            for col, value in enumerate(user_data, 1):
                cell = ws_users.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Students data sheet
        ws_students = wb.create_sheet(title="الطلاب")
        ws_students.sheet_view.rightToLeft = True
        
        student_headers = ['#', 'اسم الطالب', 'الهاتف', 'العمر', 'الموقع', 'المدرس', 'المجموعات', 'إجمالي السعر', 'الخصم', 'السعر بعد الخصم', 'المدفوع', 'المتبقي', 'تاريخ التسجيل']
        for col, header in enumerate(student_headers, 1):
            cell = ws_students.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        students = Student.query.all()
        for idx, student in enumerate(students, 1):
            instructor_name = student.instructor_ref.name if student.instructor_ref else 'غير محدد'
            groups_names = ', '.join([group.name for group in student.groups])
            
            student_data = [
                idx,
                student.name,
                student.phone or 'غير محدد',
                student.age or 'غير محدد',
                student.location or 'غير محدد',
                instructor_name,
                groups_names or 'لا توجد مجموعات',
                f"{student.total_course_price:,.0f}",
                f"{student.discount:,.0f}",
                f"{student.total_course_price_after_discount:,.0f}",
                f"{student.total_paid:,.0f}",
                f"{student.remaining_balance:,.0f}",
                student.registration_date.strftime('%Y-%m-%d') if student.registration_date else 'غير محدد'
            ]
            
            for col, value in enumerate(student_data, 1):
                cell = ws_students.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Instructors data sheet
        ws_instructors = wb.create_sheet(title="المدرسين")
        ws_instructors.sheet_view.rightToLeft = True
        
        instructor_headers = ['#', 'اسم المدرس', 'الهاتف', 'التخصص', 'عدد الطلاب', 'عدد المجموعات', 'مرتبط بمستخدم']
        for col, header in enumerate(instructor_headers, 1):
            cell = ws_instructors.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        instructors = Instructor.query.all()
        for idx, instructor in enumerate(instructors, 1):
            linked_user = 'نعم' if instructor.user_account else 'لا'
            
            instructor_data = [
                idx,
                instructor.name,
                instructor.phone or 'غير محدد',
                instructor.specialization or 'غير محدد',
                len(instructor.students),
                len(instructor.groups),
                linked_user
            ]
            
            for col, value in enumerate(instructor_data, 1):
                cell = ws_instructors.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Groups data sheet
        ws_groups = wb.create_sheet(title="المجموعات")
        ws_groups.sheet_view.rightToLeft = True
        
        group_headers = ['#', 'اسم المجموعة', 'المستوى', 'المدرس', 'عدد الطلاب', 'الحد الأقصى', 'السعر', 'أيام الدروس', 'أوقات الدروس']
        for col, header in enumerate(group_headers, 1):
            cell = ws_groups.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        groups = Group.query.all()
        for idx, group in enumerate(groups, 1):
            instructor_name = group.instructor_ref.name if group.instructor_ref else 'غير محدد'
            
            # Get schedule details
            schedules = []
            for schedule in group.schedules:
                start_12 = convert_24_to_12_hour(schedule.start_time)
                end_12 = convert_24_to_12_hour(schedule.end_time)
                schedules.append(f"{schedule.day_of_week}: {start_12['hour']}:{start_12['minute']} {start_12['period']} - {end_12['hour']}:{end_12['minute']} {end_12['period']}")
            
            days = ', '.join([s.day_of_week for s in group.schedules])
            times = ' | '.join(schedules)
            
            group_data = [
                idx,
                group.name,
                ', '.join([s.name for s in group.subjects]) if group.subjects else 'غير محدد',
                instructor_name,
                group.students.count(),
                group.max_students,
                f"{group.price:,.0f}",
                days or 'غير محدد',
                times or 'غير محدد'
            ]
            
            for col, value in enumerate(group_data, 1):
                cell = ws_groups.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Schedules data sheet
        ws_schedules = wb.create_sheet(title="الجداول الزمنية")
        ws_schedules.sheet_view.rightToLeft = True
        
        schedule_headers = ['#', 'المجموعة', 'المدرس', 'اليوم', 'وقت البداية', 'وقت النهاية', 'المدة']
        for col, header in enumerate(schedule_headers, 1):
            cell = ws_schedules.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        schedules = Schedule.query.all()
        for idx, schedule in enumerate(schedules, 1):
            group = Group.query.get(schedule.group_id)
            instructor_name = group.instructor_ref.name if group and group.instructor_ref else 'غير محدد'
            group_name = group.name if group else 'مجموعة محذوفة'
            
            # Calculate duration
            try:
                start_time = datetime.strptime(schedule.start_time, '%H:%M').time()
                end_time = datetime.strptime(schedule.end_time, '%H:%M').time()
                start_datetime = datetime.combine(datetime.today(), start_time)
                end_datetime = datetime.combine(datetime.today(), end_time)
                duration = end_datetime - start_datetime
                duration_str = f"{duration.seconds // 3600}:{(duration.seconds % 3600) // 60:02d}"
            except:
                duration_str = 'غير محدد'
            
            schedule_data = [
                idx,
                group_name,
                instructor_name,
                schedule.day_of_week,
                schedule.start_time,
                schedule.end_time,
                duration_str
            ]
            
            for col, value in enumerate(schedule_data, 1):
                cell = ws_schedules.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Payments data sheet
        ws_payments = wb.create_sheet(title="المدفوعات")
        ws_payments.sheet_view.rightToLeft = True
        
        payment_headers = ['#', 'اسم الطالب', 'المبلغ', 'الشهر', 'التاريخ', 'ملاحظات']
        for col, header in enumerate(payment_headers, 1):
            cell = ws_payments.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        payments = Payment.query.order_by(Payment.date.desc()).all()
        for idx, payment in enumerate(payments, 1):
            student = Student.query.get(payment.student_id)
            student_name = student.name if student else 'طالب محذوف'
            
            payment_data = [
                idx,
                student_name,
                f"{payment.amount:,.0f}",
                payment.month or 'غير محدد',
                payment.date.strftime('%Y-%m-%d %H:%M') if payment.date else 'غير محدد',
                payment.notes or 'لا توجد ملاحظات'
            ]
            
            for col, value in enumerate(payment_data, 1):
                cell = ws_payments.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Expenses data sheet
        ws_expenses = wb.create_sheet(title="المصروفات")
        ws_expenses.sheet_view.rightToLeft = True
        
        expense_headers = ['#', 'الوصف', 'المبلغ', 'الفئة', 'التاريخ', 'ملاحظات']
        for col, header in enumerate(expense_headers, 1):
            cell = ws_expenses.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        expenses = Expense.query.order_by(Expense.date.desc()).all()
        for idx, expense in enumerate(expenses, 1):
            expense_data = [
                idx,
                expense.description,
                f"{expense.amount:,.0f}",
                expense.category or 'غير محدد',
                expense.date.strftime('%Y-%m-%d %H:%M') if expense.date else 'غير محدد',
                expense.notes or 'لا توجد ملاحظات'
            ]
            
            for col, value in enumerate(expense_data, 1):
                cell = ws_expenses.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Attendance data sheet (last 30 days)
        ws_attendance = wb.create_sheet(title="الحضور")
        ws_attendance.sheet_view.rightToLeft = True
        
        attendance_headers = ['#', 'اسم الطالب', 'المجموعة', 'التاريخ', 'الحالة']
        for col, header in enumerate(attendance_headers, 1):
            cell = ws_attendance.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Get attendance for last 30 days
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        attendance_records = Attendance.query.filter(Attendance.date >= thirty_days_ago).order_by(Attendance.date.desc()).all()
        
        for idx, record in enumerate(attendance_records, 1):
            student = Student.query.get(record.student_id)
            group = Group.query.get(record.group_id)
            student_name = student.name if student else 'طالب محذوف'
            group_name = group.name if group else 'مجموعة محذوفة'
            
            attendance_data = [
                idx,
                student_name,
                group_name,
                record.date.strftime('%Y-%m-%d') if record.date else 'غير محدد',
                record.status
            ]
            
            for col, value in enumerate(attendance_data, 1):
                cell = ws_attendance.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Tasks data sheet
        ws_tasks = wb.create_sheet(title="المهام")
        ws_tasks.sheet_view.rightToLeft = True
        
        task_headers = ['#', 'العنوان', 'الوصف', 'الأولوية', 'الحالة', 'تاريخ الاستحقاق', 'منشئ المهمة', 'المُكلف', 'تاريخ الإنشاء', 'تاريخ الإكمال']
        for col, header in enumerate(task_headers, 1):
            cell = ws_tasks.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        tasks = Task.query.order_by(Task.created_at.desc()).all()
        for idx, task in enumerate(tasks, 1):
            creator = User.query.get(task.created_by)
            assignee = User.query.get(task.assigned_to) if task.assigned_to else None
            
            task_data = [
                idx,
                task.title,
                task.description or 'لا يوجد وصف',
                task.priority,
                task.status,
                task.due_date.strftime('%Y-%m-%d') if task.due_date else 'غير محدد',
                creator.full_name if creator else 'مستخدم محذوف',
                assignee.full_name if assignee else 'غير مُكلف',
                task.created_at.strftime('%Y-%m-%d %H:%M') if task.created_at else 'غير محدد',
                task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else 'غير مكتمل'
            ]
            
            for col, value in enumerate(task_data, 1):
                cell = ws_tasks.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Notes data sheet
        ws_notes = wb.create_sheet(title="الملاحظات")
        ws_notes.sheet_view.rightToLeft = True
        
        note_headers = ['#', 'العنوان', 'المحتوى', 'الفئة', 'اللون', 'مثبت', 'منشئ الملاحظة', 'تاريخ الإنشاء', 'تاريخ التحديث']
        for col, header in enumerate(note_headers, 1):
            cell = ws_notes.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        notes = Note.query.order_by(Note.updated_at.desc()).all()
        for idx, note in enumerate(notes, 1):
            creator = User.query.get(note.created_by)
            
            note_data = [
                idx,
                note.title,
                note.content[:100] + '...' if len(note.content) > 100 else note.content,
                note.category,
                note.color,
                'نعم' if note.is_pinned else 'لا',
                creator.full_name if creator else 'مستخدم محذوف',
                note.created_at.strftime('%Y-%m-%d %H:%M') if note.created_at else 'غير محدد',
                note.updated_at.strftime('%Y-%m-%d %H:%M') if note.updated_at else 'غير محدد'
            ]
            
            for col, value in enumerate(note_data, 1):
                cell = ws_notes.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Instructor Notes data sheet
        ws_instructor_notes = wb.create_sheet(title="ملاحظات المدرسين")
        ws_instructor_notes.sheet_view.rightToLeft = True
        
        instructor_note_headers = ['#', 'العنوان', 'المحتوى', 'الطالب', 'المجموعة', 'الأولوية', 'الحالة', 'منشئ الملاحظة', 'مراجع من الإدارة', 'تاريخ الإنشاء', 'تاريخ المراجعة', 'رد الإدارة']
        for col, header in enumerate(instructor_note_headers, 1):
            cell = ws_instructor_notes.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        instructor_notes = InstructorNote.query.order_by(InstructorNote.created_at.desc()).all()
        for idx, note in enumerate(instructor_notes, 1):
            creator = User.query.get(note.created_by)
            reviewer = User.query.get(note.reviewed_by) if note.reviewed_by else None
            student = Student.query.get(note.student_id) if note.student_id else None
            group = Group.query.get(note.group_id) if note.group_id else None
            
            note_data = [
                idx,
                note.title,
                note.content[:100] + '...' if len(note.content) > 100 else note.content,
                student.name if student else 'غير محدد',
                group.name if group else 'غير محدد',
                note.priority,
                note.status,
                creator.full_name if creator else 'مستخدم محذوف',
                reviewer.full_name if reviewer else 'لم تتم المراجعة',
                note.created_at.strftime('%Y-%m-%d %H:%M') if note.created_at else 'غير محدد',
                note.reviewed_at.strftime('%Y-%m-%d %H:%M') if note.reviewed_at else 'لم تتم المراجعة',
                note.admin_response or 'لا يوجد رد'
            ]
            
            for col, value in enumerate(note_data, 1):
                cell = ws_instructor_notes.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Instructor Todos data sheet
        ws_instructor_todos = wb.create_sheet(title="مهام المدرسين")
        ws_instructor_todos.sheet_view.rightToLeft = True
        
        instructor_todo_headers = ['#', 'العنوان', 'الوصف', 'الفئة', 'الأولوية', 'الحالة', 'الطالب', 'المجموعة', 'تاريخ الاستحقاق', 'منشئ المهمة', 'تاريخ الإنشاء', 'تاريخ التحديث', 'تاريخ الإكمال']
        for col, header in enumerate(instructor_todo_headers, 1):
            cell = ws_instructor_todos.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        instructor_todos = InstructorTodo.query.order_by(InstructorTodo.created_at.desc()).all()
        for idx, todo in enumerate(instructor_todos, 1):
            creator = User.query.get(todo.created_by)
            student = Student.query.get(todo.student_id) if todo.student_id else None
            group = Group.query.get(todo.group_id) if todo.group_id else None
            
            todo_data = [
                idx,
                todo.title,
                todo.description[:100] + '...' if todo.description and len(todo.description) > 100 else todo.description or 'لا يوجد وصف',
                todo.category,
                todo.priority,
                todo.status,
                student.name if student else 'غير محدد',
                group.name if group else 'غير محدد',
                todo.due_date.strftime('%Y-%m-%d') if todo.due_date else 'غير محدد',
                creator.full_name if creator else 'مستخدم محذوف',
                todo.created_at.strftime('%Y-%m-%d %H:%M') if todo.created_at else 'غير محدد',
                todo.updated_at.strftime('%Y-%m-%d %H:%M') if todo.updated_at else 'غير محدد',
                todo.completed_at.strftime('%Y-%m-%d %H:%M') if todo.completed_at else 'غير مكتمل'
            ]
            
            for col, value in enumerate(todo_data, 1):
                cell = ws_instructor_todos.cell(row=idx + 1, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
        
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"نسخة_احتياطية_شاملة_نظام_تفرا_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء إنشاء النسخة الاحتياطية: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/import_system_data', methods=['GET', 'POST'])
@admin_required
def import_system_data():
    """Import complete system data from Excel file"""
    if request.method == 'GET':
        return render_template('import_data.html')
    
    try:
        if 'excel_file' not in request.files:
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_system_data'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('يرجى اختيار ملف Excel للاستيراد', 'error')
            return redirect(url_for('import_system_data'))
        
        # Check if file is Excel
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('يرجى رفع ملف Excel صحيح (.xlsx أو .xls)', 'error')
            return redirect(url_for('import_system_data'))
        
        # Read the Excel file
        from openpyxl import load_workbook
        import tempfile
        import os
        import time
        
        # Create temporary file for upload
        temp_file_path = None
        wb = None
        
        try:
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                file.save(tmp_file.name)
                temp_file_path = tmp_file.name
            
            # Load workbook
            wb = load_workbook(temp_file_path, read_only=False, data_only=True)
            
            import_summary = {
                'users': 0, 'instructors': 0, 'students': 0, 'groups': 0,
                'schedules': 0, 'payments': 0, 'expenses': 0, 'errors': []
            }
            
            # Clear existing data if requested
            if request.form.get('clear_existing') == 'yes':
                # Clear all tables (except admin user)
                db.session.query(Attendance).delete()
                db.session.query(Payment).delete()
                db.session.query(Expense).delete()
                db.session.query(InstructorTodo).delete()
                db.session.query(InstructorNote).delete()
                db.session.query(Note).delete()
                db.session.query(Task).delete()
                db.session.query(Schedule).delete()
                
                # Clear many-to-many relationships
                db.session.execute(student_groups.delete())
                
                # Clear main entities
                db.session.query(Student).delete()
                db.session.query(Group).delete()
                db.session.query(Instructor).delete()
                
                # Keep only admin users
                db.session.query(User).filter(User.role != 'admin').delete()
                
                db.session.commit()
                flash('تم حذف البيانات الموجودة بنجاح', 'info')
            
            # Import Users (skip admin users to avoid conflicts)
            if 'المستخدمين' in wb.sheetnames:
                ws = wb['المستخدمين']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        username = str(row[1]).strip()
                        full_name = str(row[2]).strip()
                        role = str(row[3]).strip()
                        is_hidden = str(row[4]).strip() == 'نعم'
                        
                        # Skip admin users to avoid conflicts
                        if role == 'admin':
                            continue
                        
                        # Check if user already exists
                        if User.query.filter_by(username=username).first():
                            continue
                        
                        user = User(
                            username=username,
                            full_name=full_name,
                            role=role,
                            is_hidden=is_hidden
                        )
                        user.set_password('123456')  # Default password
                        db.session.add(user)
                        import_summary['users'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد المستخدم {row}: {str(e)}')
                
                db.session.commit()
            
            # Import Instructors
            if 'المدرسين' in wb.sheetnames:
                ws = wb['المدرسين']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        name = str(row[1]).strip()
                        phone = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'غير محدد' else None
                        specialization = str(row[3]).strip() if row[3] and str(row[3]).strip() != 'غير محدد' else None
                        
                        # Check if instructor already exists
                        if Instructor.query.filter_by(name=name).first():
                            continue
                        
                        instructor = Instructor(
                            name=name,
                            phone=phone,
                            specialization=specialization
                        )
                        db.session.add(instructor)
                        import_summary['instructors'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد المدرس {row}: {str(e)}')
                
                db.session.commit()
            
            # Import Groups
            if 'المجموعات' in wb.sheetnames:
                ws = wb['المجموعات']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        name = str(row[1]).strip()
                        level = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'غير محدد' else None
                        instructor_name = str(row[3]).strip() if row[3] and str(row[3]).strip() != 'غير محدد' else None
                        
                        # Enhanced price handling - handle various formats including Excel's numeric formatting
                        price = 0.0
                        if row[4] is not None:
                            try:
                                # Handle Excel numeric values with General format fix
                                if isinstance(row[4], (int, float)):
                                    price = float(row[4])
                                    # Handle Excel General format issue for group prices
                                    if price > 0 and price < 50:
                                        print(f"⚠️ سعر مجموعة صغير مشتبه: {price} للمجموعة {name} - سيتم ضربه في 100")
                                        price = price * 100
                                else:
                                    # Handle string values
                                    price_value = str(row[4]).replace(',', '').replace('ج.م', '').replace('جنيه', '').replace('EGP', '').strip()
                                    # Remove currency symbols and extra spaces
                                    price_value = price_value.replace('$', '').replace('£', '').replace('€', '')
                                    
                                    # Extract numeric value using regex
                                    import re
                                    numeric_match = re.search(r'(\d+(?:\.\d+)?)', price_value)
                                    if numeric_match:
                                        price = float(numeric_match.group(1))
                                        # Apply same fix for string values
                                        if price > 0 and price < 50:
                                            print(f"⚠️ سعر مجموعة نصي صغير مشتبه: {price} للمجموعة {name} - سيتم ضربه في 100")
                                            price = price * 100
                                    else:
                                        price = 0.0
                                
                                # Ensure reasonable price range after corrections
                                if price < 0:
                                    price = 0.0
                                elif price > 100000:  # Sanity check for very large prices
                                    price = 0.0
                                
                                # Debug log for price issues - with more detail
                                if price == 0.0 and row[4] is not None:
                                    import_summary['errors'].append(f'تحذير: لم يتم التعرف على السعر للمجموعة {name}: القيمة الأصلية = {row[4]} (نوع: {type(row[4]).__name__})')
                                else:
                                    # Log successful price parsing for verification
                                    print(f"✓ مجموعة {name}: السعر = {price} (من القيمة الأصلية: {row[4]})")
                                    
                            except (ValueError, TypeError) as e:
                                price = 0.0
                                import_summary['errors'].append(f'خطأ في قراءة السعر للمجموعة {name}: القيمة الأصلية = {row[4]} (نوع: {type(row[4]).__name__}) - {str(e)}')
                        
                        max_students = int(row[5]) if row[5] else 15
                        
                        # Find instructor by name
                        instructor = None
                        if instructor_name:
                            instructor = Instructor.query.filter_by(name=instructor_name).first()
                        
                        # Check if group already exists
                        if Group.query.filter_by(name=name).first():
                            continue
                        
                        group = Group(
                            name=name,
                            instructor_id=instructor.id if instructor else None,
                            price=price,
                            max_students=max_students
                        )
                        db.session.add(group)
                        import_summary['groups'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد المجموعة {row}: {str(e)}')
                
                db.session.commit()
            
            # Import Schedules
            if 'الجداول' in wb.sheetnames or 'الجداول الزمنية' in wb.sheetnames:
                # Try both possible sheet names
                sheet_name = 'الجداول' if 'الجداول' in wb.sheetnames else 'الجداول الزمنية'
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        group_name = str(row[1]).strip()
                        day_of_week = str(row[2]).strip() if row[2] else str(row[3]).strip()  # Try column 2 or 3 for day
                        
                        # Handle different time formats
                        start_time = str(row[3]).strip() if len(row) > 3 and row[3] else str(row[4]).strip() if len(row) > 4 and row[4] else ''
                        end_time = str(row[4]).strip() if len(row) > 4 and row[4] else str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        
                        # If start_time looks like day and end_time looks like time, swap them
                        arabic_days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
                        if start_time in arabic_days and ':' in end_time:
                            day_of_week = start_time
                            start_time = end_time
                            end_time = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        
                        # Clean and validate time format
                        def clean_time(time_str):
                            if not time_str:
                                return ''
                            # Remove any non-time characters
                            time_str = str(time_str).strip()
                            # Handle 12-hour format conversion if needed
                            if 'ص' in time_str or 'م' in time_str:
                                # Convert Arabic 12-hour to 24-hour
                                time_str = time_str.replace('ص', 'AM').replace('م', 'PM')
                                try:
                                    time_obj = datetime.strptime(time_str.replace(' ', ''), '%I:%M%p')
                                    return time_obj.strftime('%H:%M')
                                except:
                                    pass
                            # Extract time pattern HH:MM
                            import re
                            time_match = re.search(r'(\d{1,2}):(\d{2})', time_str)
                            if time_match:
                                hour, minute = time_match.groups()
                                hour = int(hour)
                                minute = int(minute)
                                if 0 <= hour <= 23 and 0 <= minute <= 59:
                                    return f"{hour:02d}:{minute:02d}"
                            return ''
                        
                        start_time = clean_time(start_time)
                        end_time = clean_time(end_time)
                        
                        if not start_time or not end_time:
                            continue
                        
                        # Find group by name
                        group = Group.query.filter_by(name=group_name).first()
                        if not group:
                            continue
                        
                        # Check if schedule already exists
                        existing_schedule = Schedule.query.filter_by(
                            group_id=group.id,
                            day_of_week=day_of_week,
                            start_time=start_time
                        ).first()
                        
                        if existing_schedule:
                            continue
                        
                        schedule = Schedule(
                            group_id=group.id,
                            day_of_week=day_of_week,
                            start_time=start_time,
                            end_time=end_time
                        )
                        db.session.add(schedule)
                        import_summary['schedules'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد الجدول {row}: {str(e)}')
                
                db.session.commit()
            
            # Import Students
            if 'الطلاب' in wb.sheetnames:
                ws = wb['الطلاب']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        name = str(row[1]).strip()
                        phone = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'غير محدد' else None
                        age = int(row[3]) if row[3] and str(row[3]).strip() != 'غير محدد' else None
                        location = str(row[4]).strip() if row[4] and str(row[4]).strip() != 'غير محدد' else None
                        instructor_name = str(row[5]).strip() if row[5] and str(row[5]).strip() != 'غير محدد' else None
                        groups_names = str(row[6]).strip() if row[6] and str(row[6]).strip() != 'لا توجد مجموعات' else ''
                        
                        # Helper function to parse numeric values correctly with Excel General format handling
                        def parse_numeric_value(value, field_name="", expected_range=None):
                            if value is None:
                                return 0.0
                            
                            try:
                                # If it's already a numeric type
                                if isinstance(value, (int, float)):
                                    result = float(value)
                                    
                                    # Handle Excel General format issue where values appear divided by 100
                                    # If we expect prices/payments to be >= 50 and we get a decimal < 50, multiply by 100
                                    if expected_range and expected_range == 'price' and result > 0 and result < 50:
                                        print(f"⚠️ قيمة صغيرة مشتبهة: {result} - سيتم ضربها في 100")
                                        result = result * 100
                                    elif expected_range and expected_range == 'payment' and result > 0 and result < 100:
                                        print(f"⚠️ قيمة دفع صغيرة مشتبهة: {result} - سيتم ضربها في 100")
                                        result = result * 100
                                    
                                    return result
                                
                                # Convert to string and clean
                                value_str = str(value).strip()
                                if not value_str or value_str.lower() in ['none', 'null', '']:
                                    return 0.0
                                
                                # Remove common currency symbols and separators
                                cleaned = value_str.replace(',', '').replace('ج.م', '').replace('جنيه', '')
                                cleaned = cleaned.replace('EGP', '').replace('$', '').replace('£', '').replace('€', '')
                                cleaned = cleaned.replace(' ', '').strip()
                                
                                # Extract numeric value using regex
                                import re
                                numeric_match = re.search(r'(\d+(?:\.\d+)?)', cleaned)
                                if numeric_match:
                                    result = float(numeric_match.group(1))
                                    
                                    # Apply the same logic for string values
                                    if expected_range and expected_range == 'price' and result > 0 and result < 50:
                                        print(f"⚠️ قيمة نصية صغيرة مشتبهة: {result} - سيتم ضربها في 100")
                                        result = result * 100
                                    elif expected_range and expected_range == 'payment' and result > 0 and result < 100:
                                        print(f"⚠️ قيمة دفع نصية صغيرة مشتبهة: {result} - سيتم ضربها في 100")
                                        result = result * 100
                                    
                                    return result
                                else:
                                    return 0.0
                                    
                            except (ValueError, TypeError) as e:
                                if field_name:
                                    import_summary['errors'].append(f'خطأ في قراءة {field_name} للطالب {name}: {value} (نوع: {type(value).__name__}) - {str(e)}')
                                return 0.0
                        
                        # Extract discount if available (column 8)
                        discount = parse_numeric_value(row[8] if len(row) > 8 else None, "الخصم", "price")
                        
                        # Extract total_paid if available (column 10)
                        total_paid = parse_numeric_value(row[10] if len(row) > 10 else None, "المبلغ المدفوع", "payment")
                        
                        # Extract registration_date if available (column 12)
                        registration_date = datetime.now()
                        if len(row) > 12 and row[12]:
                            try:
                                if isinstance(row[12], datetime):
                                    registration_date = row[12]
                                else:
                                    # Try to parse date string
                                    date_str = str(row[12]).strip()
                                    registration_date = datetime.strptime(date_str, '%Y-%m-%d')
                            except:
                                registration_date = datetime.now()
                        
                        # Find instructor by name
                        instructor = None
                        if instructor_name:
                            instructor = Instructor.query.filter_by(name=instructor_name).first()
                        
                        # Check if student already exists
                        if Student.query.filter_by(name=name, phone=phone).first():
                            continue
                        
                        student = Student(
                            name=name,
                            phone=phone,
                            age=age,
                            location=location,
                            instructor_id=instructor.id if instructor else None,
                            total_paid=total_paid,
                            discount=discount,
                            registration_date=registration_date
                        )
                        db.session.add(student)
                        db.session.flush()  # Get student ID
                        
                        # Add student to groups
                        if groups_names:
                            group_names_list = [name.strip() for name in groups_names.split(',')]
                            for group_name in group_names_list:
                                group = Group.query.filter_by(name=group_name).first()
                                if group:
                                    student.groups.append(group)
                        
                        import_summary['students'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد الطالب {row}: {str(e)}')
                
                db.session.commit()
            
            # Import Payments
            if 'المدفوعات' in wb.sheetnames:
                ws = wb['المدفوعات']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        student_name = str(row[1]).strip()
                        
                        # Parse payment amount correctly with General format fix
                        def parse_payment_amount(value):
                            if value is None:
                                return 0.0
                            try:
                                if isinstance(value, (int, float)):
                                    result = float(value)
                                    # Handle Excel General format issue for payments
                                    if result > 0 and result < 100:
                                        print(f"⚠️ مبلغ دفع صغير مشتبه للطالب {student_name}: {result} - سيتم ضربه في 100")
                                        result = result * 100
                                    return result
                                
                                value_str = str(value).strip()
                                if not value_str:
                                    return 0.0
                                
                                # Remove currency symbols and separators
                                cleaned = value_str.replace(',', '').replace('ج.م', '').replace('جنيه', '')
                                cleaned = cleaned.replace('EGP', '').replace('$', '').replace('£', '').replace('€', '')
                                cleaned = cleaned.replace(' ', '').strip()
                                
                                # Extract numeric value
                                import re
                                numeric_match = re.search(r'(\d+(?:\.\d+)?)', cleaned)
                                if numeric_match:
                                    result = float(numeric_match.group(1))
                                    # Apply same fix for string values
                                    if result > 0 and result < 100:
                                        print(f"⚠️ مبلغ دفع نصي صغير مشتبه للطالب {student_name}: {result} - سيتم ضربه في 100")
                                        result = result * 100
                                    return result
                                else:
                                    return 0.0
                            except (ValueError, TypeError):
                                return 0.0
                        
                        amount = parse_payment_amount(row[2])
                        month = str(row[3]).strip() if row[3] else ''
                        notes = str(row[4]).strip() if row[4] else None
                        
                        # Extract date if available (column 5)
                        payment_date = datetime.now()
                        if len(row) > 5 and row[5]:
                            try:
                                if isinstance(row[5], datetime):
                                    payment_date = row[5]
                                else:
                                    date_str = str(row[5]).strip()
                                    payment_date = datetime.strptime(date_str, '%Y-%m-%d')
                            except:
                                payment_date = datetime.now()
                        
                        # Find student by name
                        student = Student.query.filter_by(name=student_name).first()
                        if not student:
                            continue
                        
                        payment = Payment(
                            student_id=student.id,
                            amount=amount,
                            month=month,
                            notes=notes,
                            date=payment_date
                        )
                        db.session.add(payment)
                        import_summary['payments'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد المدفوعات {row}: {str(e)}')
                
                db.session.commit()
            
            # Import Expenses
            if 'المصروفات' in wb.sheetnames:
                ws = wb['المصروفات']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0] or not row[1]:  # Skip empty rows
                        continue
                    try:
                        description = str(row[1]).strip()
                        
                        # Parse expense amount correctly with General format fix
                        def parse_expense_amount(value):
                            if value is None:
                                return 0.0
                            try:
                                if isinstance(value, (int, float)):
                                    result = float(value)
                                    # Handle Excel General format issue for expenses
                                    if result > 0 and result < 100:
                                        print(f"⚠️ مبلغ مصروف صغير مشتبه ({description}): {result} - سيتم ضربه في 100")
                                        result = result * 100
                                    return result
                                
                                value_str = str(value).strip()
                                if not value_str:
                                    return 0.0
                                
                                # Remove currency symbols and separators
                                cleaned = value_str.replace(',', '').replace('ج.م', '').replace('جنيه', '')
                                cleaned = cleaned.replace('EGP', '').replace('$', '').replace('£', '').replace('€', '')
                                cleaned = cleaned.replace(' ', '').strip()
                                
                                # Extract numeric value
                                import re
                                numeric_match = re.search(r'(\d+(?:\.\d+)?)', cleaned)
                                if numeric_match:
                                    result = float(numeric_match.group(1))
                                    # Apply same fix for string values
                                    if result > 0 and result < 100:
                                        print(f"⚠️ مبلغ مصروف نصي صغير مشتبه ({description}): {result} - سيتم ضربه في 100")
                                        result = result * 100
                                    return result
                                else:
                                    return 0.0
                            except (ValueError, TypeError):
                                return 0.0
                        
                        amount = parse_expense_amount(row[2])
                        category = str(row[3]).strip() if row[3] else 'أخرى'
                        notes = str(row[4]).strip() if row[4] else None
                        
                        # Extract date if available (column 5)
                        expense_date = datetime.now()
                        if len(row) > 5 and row[5]:
                            try:
                                if isinstance(row[5], datetime):
                                    expense_date = row[5]
                                else:
                                    date_str = str(row[5]).strip()
                                    expense_date = datetime.strptime(date_str, '%Y-%m-%d')
                            except:
                                expense_date = datetime.now()
                        
                        expense = Expense(
                            description=description,
                            amount=amount,
                            category=category,
                            notes=notes,
                            date=expense_date
                        )
                        db.session.add(expense)
                        import_summary['expenses'] += 1
                    except Exception as e:
                        import_summary['errors'].append(f'خطأ في استيراد المصروفات {row}: {str(e)}')
                
                db.session.commit()
            
            # Validate imported data and provide detailed feedback
            validation_issues = []
            
            # Check for groups with zero prices
            zero_price_groups = Group.query.filter_by(price=0.0).all()
            if zero_price_groups:
                group_names = [g.name for g in zero_price_groups[:3]]
                if len(zero_price_groups) > 3:
                    group_names.append(f'و {len(zero_price_groups) - 3} مجموعات أخرى')
                validation_issues.append(f'تحذير: {len(zero_price_groups)} مجموعة بسعر صفر: {", ".join(group_names)}')
            
            # Check for schedules without groups
            orphaned_schedules = Schedule.query.filter(~Schedule.group_id.in_(
                db.session.query(Group.id).subquery()
            )).count()
            if orphaned_schedules > 0:
                validation_issues.append(f'تحذير: {orphaned_schedules} جدول زمني بدون مجموعة مرتبطة')
            
            # Check for groups without schedules
            groups_without_schedules = Group.query.filter(~Group.id.in_(
                db.session.query(Schedule.group_id).filter(Schedule.group_id.isnot(None)).subquery()
            )).count()
            if groups_without_schedules > 0:
                validation_issues.append(f'تحذير: {groups_without_schedules} مجموعة بدون جدول زمني')
            
            # Generate success message with detailed statistics
            success_msg = f"تم استيراد البيانات بنجاح! "
            success_msg += f"المستخدمين: {import_summary['users']}, "
            success_msg += f"المدرسين: {import_summary['instructors']}, "
            success_msg += f"المجموعات: {import_summary['groups']}, "
            success_msg += f"الجداول: {import_summary['schedules']}, "
            success_msg += f"الطلاب: {import_summary['students']}, "
            success_msg += f"المدفوعات: {import_summary['payments']}, "
            success_msg += f"المصروفات: {import_summary['expenses']}"
            
            flash(success_msg, 'success')
            
            # Show validation issues
            for issue in validation_issues:
                flash(issue, 'info')
            
            # Show errors if any
            if import_summary['errors']:
                for error in import_summary['errors'][:5]:  # Show first 5 errors
                    flash(error, 'warning')
                if len(import_summary['errors']) > 5:
                    flash(f'وتوجد {len(import_summary["errors"]) - 5} أخطاء أخرى...', 'warning')
            
            return redirect(url_for('reports'))
            
        finally:
            # Close workbook properly to release file handle
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            # Clean up temporary file with multiple attempts
            if temp_file_path:
                cleanup_attempts = 0
                max_attempts = 3
                while cleanup_attempts < max_attempts:
                    try:
                        os.unlink(temp_file_path)
                        break  # Success - exit loop
                    except Exception as cleanup_error:
                        cleanup_attempts += 1
                        if cleanup_attempts < max_attempts:
                            # Wait and try again
                            time.sleep(0.1)
                        else:
                            # Log warning for last attempt
                            print(f"Warning: Could not delete temporary file {temp_file_path} after {max_attempts} attempts: {cleanup_error}")
                            # File will be cleaned up by system later
                            break
                
    except Exception as e:
        flash(f'حدث خطأ أثناء استيراد البيانات: {str(e)}', 'error')
        return redirect(url_for('import_system_data'))

@app.route('/admin_respond_instructor_note/<int:note_id>', methods=['POST'])
@admin_required
def admin_respond_instructor_note(note_id):
    """Admin responds to instructor note"""
    try:
        current_user = get_current_user()
        note = InstructorNote.query.get_or_404(note_id)
        
        response = request.form.get('response', '')
        status = request.form.get('status', 'قيد المراجعة')
        
        note.admin_response = response
        note.status = status
        note.reviewed_by = current_user.id
        note.reviewed_at = datetime.utcnow()
        
        db.session.commit()
        flash('تم الرد على الملاحظة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء الرد على الملاحظة', 'error')
    
    return redirect(url_for('tasks') + '#instructor-notes')

@app.route('/admin_update_instructor_note_status/<int:note_id>', methods=['POST'])
@admin_required
def admin_update_instructor_note_status(note_id):
    """Admin updates instructor note status"""
    try:
        current_user = get_current_user()
        note = InstructorNote.query.get_or_404(note_id)
        
        new_status = request.form['status']
        note.status = new_status
        
        # If not reviewed yet, mark as reviewed
        if not note.reviewed_by:
            note.reviewed_by = current_user.id
            note.reviewed_at = datetime.utcnow()
        
        db.session.commit()
        flash('تم تحديث حالة الملاحظة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء تحديث حالة الملاحظة', 'error')
    
    return redirect(url_for('tasks') + '#instructor-notes')

@app.route('/diagnose_import_data', methods=['GET'])
@admin_required
def diagnose_import_data():
    """Diagnose and show details about imported data"""
    try:
        diagnosis = {
            'groups_with_issues': [],
            'students_with_issues': [],
            'payments_with_issues': [],
            'expenses_with_issues': []
        }
        
        # Check groups with unusual prices
        groups = Group.query.all()
        for group in groups:
            issue = None
            if group.price == 0.0:
                issue = 'سعر صفر'
            elif group.price > 0 and group.price < 50:
                issue = 'سعر صغير مشتبه (ربما Excel General format)'
            elif group.price < 10:
                issue = 'سعر صغير جداً'
            elif group.price > 10000:
                issue = 'سعر مرتفع جداً'
                
            if issue:
                diagnosis['groups_with_issues'].append({
                    'id': group.id,
                    'name': group.name,
                    'price': group.price,
                    'issue': issue
                })
        
        # Check students with unusual amounts
        students = Student.query.all()
        for student in students:
            issues = []
            if student.discount < 0:
                issues.append('خصم سالب')
            if student.total_paid < 0:
                issues.append('مبلغ مدفوع سالب')
            if student.discount > 0 and student.discount < 50:
                issues.append('خصم صغير مشتبه (ربما Excel General format)')
            if student.total_paid > 0 and student.total_paid < 100:
                issues.append('مبلغ مدفوع صغير مشتبه (ربما Excel General format)')
            if student.discount > 5000:
                issues.append('خصم مرتفع جداً')
            if student.total_paid > 50000:
                issues.append('مبلغ مدفوع مرتفع جداً')
                
            if issues:
                diagnosis['students_with_issues'].append({
                    'id': student.id,
                    'name': student.name,
                    'discount': student.discount,
                    'total_paid': student.total_paid,
                    'issues': ', '.join(issues)
                })
        
        # Check payments with unusual amounts
        payments = Payment.query.all()
        for payment in payments:
            issue = None
            if payment.amount <= 0:
                issue = 'مبلغ صفر أو سالب'
            elif payment.amount > 0 and payment.amount < 100:
                issue = 'مبلغ صغير مشتبه (ربما Excel General format)'
            elif payment.amount > 50000:
                issue = 'مبلغ مرتفع جداً'
                
            if issue:
                diagnosis['payments_with_issues'].append({
                    'id': payment.id,
                    'student_name': payment.student.name if payment.student else 'غير محدد',
                    'amount': payment.amount,
                    'date': payment.date,
                    'issue': issue
                })
        
        # Check expenses with unusual amounts
        expenses = Expense.query.all()
        for expense in expenses:
            issue = None
            if expense.amount <= 0:
                issue = 'مبلغ صفر أو سالب'
            elif expense.amount > 0 and expense.amount < 100:
                issue = 'مبلغ صغير مشتبه (ربما Excel General format)'
            elif expense.amount > 100000:
                issue = 'مبلغ مرتفع جداً'
                
            if issue:
                diagnosis['expenses_with_issues'].append({
                    'id': expense.id,
                    'description': expense.description,
                    'amount': expense.amount,
                    'date': expense.date,
                    'issue': issue
                })
        
        return jsonify({
            'success': True,
            'diagnosis': diagnosis
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في تشخيص البيانات: {str(e)}'
        })



@app.route('/fix_import_data', methods=['POST'])
@admin_required
def fix_import_data():
    """Fix common issues after data import"""
    try:
        fixed_count = 0
        
        # Fix groups with zero prices OR suspiciously small prices (Excel General format issue)
        zero_or_small_price_groups = Group.query.filter(
            db.or_(Group.price == 0.0, db.and_(Group.price > 0, Group.price < 50))
        ).all()
        
        for group in zero_or_small_price_groups:
            original_price = group.price
            
            # If the price looks like it was divided by 100 (Excel General format issue)
            if group.price > 0 and group.price < 50:
                group.price = group.price * 100
                print(f"🔧 إصلاح سعر المجموعة {group.name}: {original_price} → {group.price}")
                fixed_count += 1
                continue
            
            # Set default prices for zero-price groups based on level and name
            price = 300.0  # Default price
            
            # Check subjects for pricing hints
            if group.subjects:
                subject_names = [s.name.lower() for s in group.subjects]
                subject_text = ' '.join(subject_names)
                if 'متقدم' in subject_text or 'advanced' in subject_text:
                    price = 500.0
                elif 'متوسط' in subject_text or 'intermediate' in subject_text:
                    price = 350.0
                elif 'مبتدئ' in subject_text or 'beginner' in subject_text:
                    price = 250.0
            
            # Check group name for more specific pricing
            if group.name:
                name_lower = group.name.lower()
                if any(keyword in name_lower for keyword in ['انجليزي', 'english', 'ielts', 'toefl']):
                    price = max(price, 400.0)  # English courses tend to be higher
                elif any(keyword in name_lower for keyword in ['رياضيات', 'math', 'calculus']):
                    price = max(price, 350.0)
                elif any(keyword in name_lower for keyword in ['فيزياء', 'physics', 'كيمياء', 'chemistry']):
                    price = max(price, 380.0)
                elif any(keyword in name_lower for keyword in ['برمجة', 'programming', 'كمبيوتر', 'computer']):
                    price = max(price, 450.0)  # Programming courses tend to be higher
            
            group.price = price
            print(f"🔧 تحديد سعر افتراضي للمجموعة {group.name}: {price}")
            fixed_count += 1
        
        # Fix students with suspiciously small amounts (Excel General format issue)
        small_discount_students = Student.query.filter(
            db.and_(Student.discount > 0, Student.discount < 50)
        ).all()
        
        for student in small_discount_students:
            original_discount = student.discount
            student.discount = student.discount * 100
            print(f"🔧 إصلاح خصم الطالب {student.name}: {original_discount} → {student.discount}")
            fixed_count += 1
        
        small_paid_students = Student.query.filter(
            db.and_(Student.total_paid > 0, Student.total_paid < 100)
        ).all()
        
        for student in small_paid_students:
            original_paid = student.total_paid
            student.total_paid = student.total_paid * 100
            print(f"🔧 إصلاح مبلغ مدفوع للطالب {student.name}: {original_paid} → {student.total_paid}")
            fixed_count += 1
        
        # Fix payments with suspiciously small amounts
        small_payments = Payment.query.filter(
            db.and_(Payment.amount > 0, Payment.amount < 100)
        ).all()
        
        for payment in small_payments:
            original_amount = payment.amount
            payment.amount = payment.amount * 100
            student_name = payment.student.name if payment.student else 'غير محدد'
            print(f"🔧 إصلاح مبلغ دفع للطالب {student_name}: {original_amount} → {payment.amount}")
            fixed_count += 1
        
        # Fix expenses with suspiciously small amounts
        small_expenses = Expense.query.filter(
            db.and_(Expense.amount > 0, Expense.amount < 100)
        ).all()
        
        for expense in small_expenses:
            original_amount = expense.amount
            expense.amount = expense.amount * 100
            print(f"🔧 إصلاح مبلغ مصروف ({expense.description}): {original_amount} → {expense.amount}")
            fixed_count += 1
        
        # Remove orphaned schedules (schedules without groups)
        orphaned_schedules = Schedule.query.filter(~Schedule.group_id.in_(
            db.session.query(Group.id).subquery()
        )).all()
        for schedule in orphaned_schedules:
            db.session.delete(schedule)
            fixed_count += 1
        
        db.session.commit()
        
        if fixed_count > 0:
            flash(f'تم إصلاح {fixed_count} عنصر بنجاح', 'success')
        else:
            flash('لا توجد مشاكل تحتاج للإصلاح', 'info')
            
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الإصلاح: {str(e)}', 'error')
    
    return redirect(url_for('reports'))

@app.route('/diagnose_financial_calculations', methods=['GET'])
@admin_required
def diagnose_financial_calculations():
    """Diagnose financial calculations to check for logical errors"""
    try:
        # Get all financial data
        total_revenue = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
        
        # Calculate pending payments manually with detailed logging
        pending_payments = 0
        students = Student.query.all()
        student_details = []
        
        for student in students:
            student_groups = [group.name for group in student.groups]
            total_group_prices = sum(group.price for group in student.groups)
            
            student_detail = {
                'name': student.name,
                'groups': student_groups,
                'group_prices': [group.price for group in student.groups],
                'total_group_price': total_group_prices,
                'discount': student.discount,
                'total_paid': student.total_paid,
                'calculated_price_after_discount': total_group_prices - student.discount,
                'remaining_balance': student.remaining_balance,
                'expected_remaining': max(0, (total_group_prices - student.discount) - student.total_paid)
            }
            
            student_details.append(student_detail)
            
            if student.remaining_balance > 0:
                pending_payments += student.remaining_balance
        
        # Calculate expected revenue manually
        total_groups_revenue = sum(student.total_course_price_after_discount for student in students)
        
        # Manual calculation for verification
        manual_expected_revenue = sum(max(0, sum(group.price for group in student.groups) - student.discount) for student in students)
        
        # Check logical consistency
        calculated_total = total_revenue + pending_payments
        
        diagnosis = {
            'total_revenue': total_revenue,
            'pending_payments': pending_payments,
            'total_groups_revenue': total_groups_revenue,
            'manual_expected_revenue': manual_expected_revenue,
            'calculated_total_should_equal_expected': calculated_total,
            'logical_consistency_check': abs(calculated_total - total_groups_revenue) < 0.01,
            'manual_vs_property_consistency': abs(manual_expected_revenue - total_groups_revenue) < 0.01,
            'difference': calculated_total - total_groups_revenue,
            'students_count': len(students),
            'student_details': student_details[:10],  # First 10 students for detailed view
            'payments_count': Payment.query.count(),
            'groups_with_zero_price': [group.name for group in Group.query.filter_by(price=0.0).all()],
            'students_with_negative_balance': [s.name for s in students if s.remaining_balance < 0],
            'students_with_issues': [
                s.name for s in students 
                if abs(s.remaining_balance - max(0, (sum(g.price for g in s.groups) - s.discount) - s.total_paid)) > 0.01
            ]
        }
        
        return jsonify({
            'success': True,
            'diagnosis': diagnosis
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في التشخيص المالي: {str(e)}'
        })

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 8009))
    # Enable debug mode for development by default
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=True)
else:
    # Production mode: Initialize database when imported
    init_db() 